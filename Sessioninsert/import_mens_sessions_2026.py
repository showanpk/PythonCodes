from __future__ import annotations

"""
Import the workbook "Mens Sessions 2026 (1).xlsx" into:
    dbo.Sessions
    dbo.SessionAttendance

Safety:
    - The script runs as a dry run by default and rolls back all SQL changes.
    - Add --apply only after checking the dry-run summary.
    - Add --inspect-only to check the Excel parsing without connecting to SQL Server.

Required packages:
    pip install openpyxl pyodbc

Required environment variables (recommended):
    CRM_SQL_USER
    CRM_SQL_PASSWORD

Optional environment variables:
    CRM_SQL_SERVER      default: tcp:sahelihub.database.windows.net,1433
    CRM_SQL_DATABASE    default: SahelihubCRM
    CRM_SQL_CONNECTION_STRING  full override if required
"""

import argparse
import os
import re
import sys
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

try:
    import pyodbc
except ModuleNotFoundError:
    pyodbc = None  # type: ignore[assignment]

from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

DEFAULT_INPUT_FILE = Path(r"C:\Users\shonk\Downloads\Mens Sessions 2026 (1).xlsx")
IMPORT_SOURCE_NAME = "Mens Sessions 2026"

REUSE_EXISTING_SESSIONS = True
SKIP_DUPLICATE_ATTENDANCE = True
DEFAULT_SESSION_DURATION_MINUTES = 60

# IMPORTANT:
# The workbook itself does not contain a dedicated VenueName column.
# Review these mappings before using --apply.
# The ARCC mappings are clear from the worksheet names.
# Lunch Club, Digital Skills and Mens Walk are assumed to be ARCC here.
SHEET_CONFIG: dict[str, dict[str, str]] = {
    "Men's Multisport Calthorpe": {
        "activity_name": "Men's Multisport",
        "venue_name": "Calthorpe Wellbeing Hub",
        "category": "Fitness",
        "sub_category": "Men's Health",
        "activity_category": "Fitness",
    },
    "Men's Circuit ARCC": {
        "activity_name": "Men's Circuit",
        "venue_name": "Alum Rock Community Centre",
        "category": "Fitness",
        "sub_category": "Men's Health",
        "activity_category": "Fitness",
    },
    "Men's Exercise ARCC": {
        "activity_name": "Men's Exercise",
        "venue_name": "Alum Rock Community Centre",
        "category": "Fitness",
        "sub_category": "Men's Health",
        "activity_category": "Fitness",
    },
    "Over 50's Health Club ARCC": {
        "activity_name": "Over 50's Health Club",
        "venue_name": "Alum Rock Community Centre",
        "category": "Health & Wellbeing",
        "sub_category": "Men's Health",
        "activity_category": "Health & Wellbeing",
    },
    "Lunch Club": {
        "activity_name": "Lunch Club",
        "venue_name": "Alum Rock Community Centre",
        "category": "Community",
        "sub_category": "Men's Health",
        "activity_category": "Social",
    },
    "Digital Skills": {
        "activity_name": "Digital Skills",
        "venue_name": "Alum Rock Community Centre",
        "category": "Skills",
        "sub_category": "Men's Health",
        "activity_category": "Skills",
    },
    "Mens Walk": {
        "activity_name": "Men's Walk",
        "venue_name": "Alum Rock Community Centre",
        "category": "Fitness",
        "sub_category": "Men's Health",
        "activity_category": "Fitness",
    },
}

SKIP_SHEETS = {"Template", "Sheet1"}

NULL_TEXT_VALUES = {
    "",
    "#N/A",
    "N/A",
    "NA",
    "NONE",
    "NULL",
    "NAN",
    "(BLANK)",
    "-",
    "--",
}

NO_ATTENDANCE_WORDS = {
    "no attended",
    "no attendance",
    "no one attended",
    "none attended",
    "nobody attended",
    "no participant",
    "no participants",
}

CANCELLED_WORDS = {
    "cancelled",
    "canceled",
    "bad weather got cancelled",
}

# Worksheet columns
COL_SESSION = 1
COL_DAY = 2
COL_DATE = 3
COL_MONTH = 4
COL_TIME = 5
COL_CARD = 6
COL_NAME = 7
COL_EMERGENCY_NAME = 8
COL_EMERGENCY_PHONE = 9
COL_RISK = 10


# ============================================================
# DATA CLASSES
# ============================================================

@dataclass
class ParsedRow:
    source_sheet: str
    source_row: int
    activity_name: str
    venue_name: str
    category: str
    sub_category: str
    activity_category: str
    session_day: str
    session_date: date
    session_month: str
    start_time: str
    end_time: str
    raw_time: Optional[str]
    status: str
    original_card_value: Optional[str]
    card_number: Optional[str]
    member_name: Optional[str]
    emergency_name: Optional[str]
    emergency_phone: Optional[str]
    risk_stratification: Optional[str]
    has_attendance_person: bool

    @property
    def session_key(self) -> tuple[str, str, str, str, str]:
        return (
            self.venue_name.casefold().strip(),
            self.activity_name.casefold().strip(),
            self.session_date.isoformat(),
            self.start_time,
            self.end_time,
        )


@dataclass
class SheetStats:
    worksheet: str
    activity_name: str
    venue_name: str
    session_keys: set[tuple[str, str, str, str, str]] = field(default_factory=set)
    attendance_candidates: int = 0
    blank_rows_skipped: int = 0
    invalid_datetime_rows: int = 0
    cancelled_or_no_attendance_rows: int = 0


# ============================================================
# BASIC CLEANING
# ============================================================

def clean_text(value: Any, *, zero_is_null: bool = True) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()

    text = re.sub(r"\s+", " ", text).strip()

    if text.upper() in NULL_TEXT_VALUES:
        return None

    if zero_is_null and text in {"0", "00", "0.0"}:
        return None

    return text or None


def normalise_phone(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None

    # Preserve a leading + but remove ordinary formatting separators.
    if text.startswith("+"):
        return "+" + re.sub(r"\D", "", text[1:])

    digits = re.sub(r"\D", "", text)
    return digits or None


def normalise_risk(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None

    fixes = {
        "low": "Low",
        "medium": "Medium",
        "high": "High",
    }

    return fixes.get(text.casefold(), text)


def normalise_day(value: Any, fallback_date: date) -> str:
    text = clean_text(value)
    if not text:
        return fallback_date.strftime("%A")

    fixes = {
        "mon": "Monday",
        "monday": "Monday",
        "tue": "Tuesday",
        "tues": "Tuesday",
        "tuesday": "Tuesday",
        "wed": "Wednesday",
        "wednesday": "Wednesday",
        "thu": "Thursday",
        "thur": "Thursday",
        "thurs": "Thursday",
        "thursday": "Thursday",
        "fri": "Friday",
        "friday": "Friday",
        "sat": "Saturday",
        "saturday": "Saturday",
        "sun": "Sunday",
        "sunday": "Sunday",
    }

    return fixes.get(text.casefold(), text.title())


def excel_value_to_date(value: Any) -> Optional[date]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel date epoch used by Windows workbooks.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    text = clean_text(value, zero_is_null=False)
    if not text:
        return None

    for fmt in (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d/%m/%y",
        "%d-%m-%y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    return None


# ============================================================
# TIME PARSING
# ============================================================

def minutes_to_sql_time(total_minutes: int) -> str:
    total_minutes %= 24 * 60
    return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}:00"


def sql_time_to_minutes(value: str) -> int:
    return int(value[0:2]) * 60 + int(value[3:5])


def _parse_one_time_token(token: str, fallback_suffix: Optional[str] = None) -> Optional[str]:
    text = token.strip().lower()
    text = text.strip(" .,-–—")

    if not text:
        return None

    if text == "noon":
        return "12:00:00"

    if text == "midnight":
        return "00:00:00"

    # Normalise am/pm variations such as a.m., a.m, a m, pm and p.m.
    text = re.sub(r"a\s*\.?\s*m\.?", "am", text)
    text = re.sub(r"p\s*\.?\s*m\.?", "pm", text)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"(?<=\d)\.(?=\d)", ":", text)

    suffix: Optional[str] = None
    if text.endswith("am"):
        suffix = "am"
        text = text[:-2]
    elif text.endswith("pm"):
        suffix = "pm"
        text = text[:-2]
    elif fallback_suffix in {"am", "pm"}:
        suffix = fallback_suffix

    # Handles compact forms such as 930 and 1030.
    if ":" not in text and text.isdigit():
        if len(text) == 4:
            text = f"{text[:2]}:{text[2:]}"
        elif len(text) == 3:
            text = f"{text[0]}:{text[1:]}"

    match = re.fullmatch(r"(\d{1,2})(?::(\d{1,2}))?", text)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or 0)

    if suffix == "pm" and hour != 12:
        hour += 12
    elif suffix == "am" and hour == 12:
        hour = 0

    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None

    return f"{hour:02d}:{minute:02d}:00"


def parse_time_or_range(value: Any) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Return (start_time, end_time, raw_time_text)."""
    if value is None:
        return None, None, None

    if isinstance(value, datetime):
        start = f"{value.hour:02d}:{value.minute:02d}:00"
        return start, minutes_to_sql_time(sql_time_to_minutes(start) + DEFAULT_SESSION_DURATION_MINUTES), str(value)

    if isinstance(value, time):
        start = f"{value.hour:02d}:{value.minute:02d}:00"
        return start, minutes_to_sql_time(sql_time_to_minutes(start) + DEFAULT_SESSION_DURATION_MINUTES), str(value)

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if 0 <= numeric < 1:
            start_minutes = round(numeric * 24 * 60)
            start = minutes_to_sql_time(start_minutes)
            end = minutes_to_sql_time(start_minutes + DEFAULT_SESSION_DURATION_MINUTES)
            return start, end, str(value)

    raw = str(value).strip()
    if not raw:
        return None, None, None

    text = raw.lower().strip()
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\s+", " ", text)

    # Examples: 12 noon, noon, 12 midnight, midnight.
    if re.fullmatch(r"(?:12\s*)?noon", text):
        start = "12:00:00"
        return start, minutes_to_sql_time(sql_time_to_minutes(start) + DEFAULT_SESSION_DURATION_MINUTES), raw

    if re.fullmatch(r"(?:12\s*)?midnight", text):
        start = "00:00:00"
        return start, minutes_to_sql_time(sql_time_to_minutes(start) + DEFAULT_SESSION_DURATION_MINUTES), raw

    # Split explicit ranges without using a regex character-class dash bug.
    parts = re.split(r"\s*(?:-|\bto\b)\s*", text, maxsplit=1)

    if len(parts) == 2:
        start_token, end_token = parts

        start_has_am = bool(re.search(r"a\s*\.?\s*m\.?", start_token))
        start_has_pm = bool(re.search(r"p\s*\.?\s*m\.?", start_token))
        end_has_am = bool(re.search(r"a\s*\.?\s*m\.?", end_token))
        end_has_pm = bool(re.search(r"p\s*\.?\s*m\.?", end_token))

        start_suffix = "am" if start_has_am else "pm" if start_has_pm else None
        end_suffix = "am" if end_has_am else "pm" if end_has_pm else None

        start = _parse_one_time_token(start_token, fallback_suffix=start_suffix or end_suffix)
        end = _parse_one_time_token(end_token, fallback_suffix=end_suffix or start_suffix)

        if start and end:
            return start, end, raw

        return None, None, raw

    start = _parse_one_time_token(text)
    if not start:
        return None, None, raw

    end = minutes_to_sql_time(sql_time_to_minutes(start) + DEFAULT_SESSION_DURATION_MINUTES)
    return start, end, raw


# ============================================================
# WORKBOOK PARSING
# ============================================================

def row_status(*values: Any) -> str:
    combined = " ".join(clean_text(value) or "" for value in values).casefold()

    if any(word in combined for word in CANCELLED_WORDS):
        return "cancelled"

    if any(word in combined for word in NO_ATTENDANCE_WORDS):
        return "no_attendance"

    return "normal"


def normalise_original_card(value: Any) -> Optional[str]:
    return clean_text(value)


def normalise_numeric_card(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None

    if re.fullmatch(r"\d+(?:\.0)?", text):
        return str(int(float(text)))

    # Values such as L-14 are kept in notes but are not treated as FULL card numbers.
    return None


def parse_workbook(input_file: Path) -> tuple[list[ParsedRow], dict[str, SheetStats]]:
    workbook = load_workbook(input_file, data_only=True, read_only=False)

    parsed_rows: list[ParsedRow] = []
    stats: dict[str, SheetStats] = {}

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title.strip()

        if sheet_name in SKIP_SHEETS:
            continue

        if sheet_name not in SHEET_CONFIG:
            print(f"Skipping worksheet without configuration: {sheet_name}")
            continue

        config = SHEET_CONFIG[sheet_name]
        sheet_stats = SheetStats(
            worksheet=sheet_name,
            activity_name=config["activity_name"],
            venue_name=config["venue_name"],
        )
        stats[sheet_name] = sheet_stats

        current_day: Optional[str] = None
        current_date: Optional[date] = None
        current_month: Optional[str] = None
        current_start_time: Optional[str] = None
        current_end_time: Optional[str] = None
        current_raw_time: Optional[str] = None

        for source_row in range(2, worksheet.max_row + 1):
            raw_session = clean_text(worksheet.cell(source_row, COL_SESSION).value)
            raw_day = clean_text(worksheet.cell(source_row, COL_DAY).value)
            raw_date = excel_value_to_date(worksheet.cell(source_row, COL_DATE).value)
            raw_month = clean_text(worksheet.cell(source_row, COL_MONTH).value)
            raw_time_value = worksheet.cell(source_row, COL_TIME).value

            raw_card_value = worksheet.cell(source_row, COL_CARD).value
            raw_name_value = worksheet.cell(source_row, COL_NAME).value
            raw_emergency_name = worksheet.cell(source_row, COL_EMERGENCY_NAME).value
            raw_emergency_phone = worksheet.cell(source_row, COL_EMERGENCY_PHONE).value
            raw_risk = worksheet.cell(source_row, COL_RISK).value

            raw_card_text = normalise_original_card(raw_card_value)
            card_number = normalise_numeric_card(raw_card_value)
            member_name = clean_text(raw_name_value)
            emergency_name = clean_text(raw_emergency_name)
            emergency_phone = normalise_phone(raw_emergency_phone)
            risk_stratification = normalise_risk(raw_risk)

            if raw_day:
                current_day = raw_day

            if raw_date:
                current_date = raw_date

            if raw_month:
                current_month = raw_month

            parsed_start, parsed_end, raw_time_text = parse_time_or_range(raw_time_value)
            if parsed_start and parsed_end:
                current_start_time = parsed_start
                current_end_time = parsed_end
                current_raw_time = raw_time_text

            has_any_row_content = any(
                value is not None
                for value in (
                    raw_session,
                    raw_day,
                    raw_date,
                    raw_month,
                    clean_text(raw_time_value, zero_is_null=False),
                    raw_card_text,
                    member_name,
                    emergency_name,
                    emergency_phone,
                    risk_stratification,
                )
            )

            if not has_any_row_content:
                sheet_stats.blank_rows_skipped += 1
                continue

            if not current_date or not current_start_time or not current_end_time:
                sheet_stats.invalid_datetime_rows += 1
                print(
                    "Skipping row: missing date/time context | "
                    f"Sheet={sheet_name} | Row={source_row} | "
                    f"Date={current_date} | Time={current_start_time} | "
                    f"RawTime={raw_time_value!r} | Card={raw_card_text!r} | Name={member_name!r}"
                )
                continue

            session_day = normalise_day(current_day, current_date)
            # Date is the source of truth. Some workbook Month cells are inconsistent.
            session_month = current_date.strftime("%B")
            status = row_status(raw_session, raw_card_text, member_name, emergency_name, emergency_phone)
            has_attendance_person = bool(raw_card_text or member_name)

            parsed = ParsedRow(
                source_sheet=sheet_name,
                source_row=source_row,
                activity_name=config["activity_name"],
                venue_name=config["venue_name"],
                category=config["category"],
                sub_category=config["sub_category"],
                activity_category=config["activity_category"],
                session_day=session_day,
                session_date=current_date,
                session_month=session_month,
                start_time=current_start_time,
                end_time=current_end_time,
                raw_time=current_raw_time,
                status=status,
                original_card_value=raw_card_text,
                card_number=card_number,
                member_name=member_name,
                emergency_name=emergency_name,
                emergency_phone=emergency_phone,
                risk_stratification=risk_stratification,
                has_attendance_person=has_attendance_person,
            )

            parsed_rows.append(parsed)
            sheet_stats.session_keys.add(parsed.session_key)

            if status in {"cancelled", "no_attendance"}:
                sheet_stats.cancelled_or_no_attendance_rows += 1
            elif has_attendance_person:
                sheet_stats.attendance_candidates += 1

    return parsed_rows, stats


def print_excel_inspection(stats: dict[str, SheetStats]) -> None:
    print("\nExcel parsing summary")
    print("=" * 90)
    print(f"{'Worksheet':34} {'Venue':30} {'Sessions':>8} {'Attendance rows':>16} {'Invalid date/time':>18}")
    print("-" * 90)

    for item in stats.values():
        print(
            f"{item.worksheet:34} "
            f"{item.venue_name:30} "
            f"{len(item.session_keys):>8} "
            f"{item.attendance_candidates:>16} "
            f"{item.invalid_datetime_rows:>18}"
        )

    print("-" * 90)
    print(f"{'TOTAL':34} {'':30} {sum(len(s.session_keys) for s in stats.values()):>8} {sum(s.attendance_candidates for s in stats.values()):>16} {sum(s.invalid_datetime_rows for s in stats.values()):>18}")
    print()


# ============================================================
# SQL CONNECTION AND HELPERS
# ============================================================

def build_connection_string() -> str:
    override = os.getenv("CRM_SQL_CONNECTION_STRING")
    if override:
        return override

    sql_user = os.getenv("CRM_SQL_USER")
    sql_password = os.getenv("CRM_SQL_PASSWORD")

    if not sql_user or not sql_password:
        raise RuntimeError(
            "Missing SQL credentials. Set CRM_SQL_USER and CRM_SQL_PASSWORD environment variables, "
            "or set CRM_SQL_CONNECTION_STRING."
        )

    server = os.getenv("CRM_SQL_SERVER", "tcp:sahelihub.database.windows.net,1433")
    database = os.getenv("CRM_SQL_DATABASE", "SahelihubCRM")

    return (
        "DRIVER={ODBC Driver 18 for SQL Server};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={sql_user};"
        f"PWD={sql_password};"
        "Encrypt=yes;"
        "TrustServerCertificate=yes;"
    )


def is_identity_column(cursor: pyodbc.Cursor, table_name: str, column_name: str) -> bool:
    sql = """
    SELECT c.is_identity
    FROM sys.columns c
    WHERE c.object_id = OBJECT_ID(?)
      AND c.name = ?;
    """
    row = cursor.execute(sql, f"dbo.{table_name}", column_name).fetchone()

    if not row:
        raise RuntimeError(f"Column not found: dbo.{table_name}.{column_name}")

    return bool(row.is_identity)


def get_next_manual_id(cursor: pyodbc.Cursor, table_name: str, column_name: str) -> int:
    sql = f"SELECT ISNULL(MAX([{column_name}]), 0) + 1 AS NextId FROM dbo.[{table_name}];"
    row = cursor.execute(sql).fetchone()
    return int(row.NextId)


def get_next_lite_display_number(cursor: pyodbc.Cursor) -> int:
    sql = """
    SELECT ISNULL(MAX(TRY_CONVERT(int, REPLACE(MemberDisplayId, 'LITE-', ''))), 0) + 1 AS NextLiteNumber
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND MemberDisplayId LIKE 'LITE-%';
    """
    row = cursor.execute(sql).fetchone()
    return int(row.NextLiteNumber)


def find_existing_session(cursor: pyodbc.Cursor, row: ParsedRow) -> Optional[int]:
    sql = """
    SELECT TOP 1 SessionId
    FROM dbo.Sessions
    WHERE LOWER(LTRIM(RTRIM(ISNULL(VenueName, '')))) = LOWER(LTRIM(RTRIM(?)))
      AND LOWER(LTRIM(RTRIM(ISNULL(ActivityName, '')))) = LOWER(LTRIM(RTRIM(?)))
      AND CAST(SessionDate AS date) = CAST(? AS date)
      AND CONVERT(varchar(8), StartTime, 108) = ?
      AND CONVERT(varchar(8), EndTime, 108) = ?
    ORDER BY SessionId;
    """
    found = cursor.execute(
        sql,
        row.venue_name,
        row.activity_name,
        row.session_date.isoformat(),
        row.start_time,
        row.end_time,
    ).fetchone()

    return int(found.SessionId) if found else None


def create_session(
    cursor: pyodbc.Cursor,
    row: ParsedRow,
    created_at: datetime,
    session_identity: bool,
    manual_session_id: Optional[int],
) -> int:
    columns = [
        "Frequency",
        "Category",
        "SubCategory",
        "ActivityCategory",
        "VenueName",
        "ActivityName",
        "Notes",
        "IsRecurringWeekly",
        "DayOfWeek",
        "SessionDate",
        "ArrivalTime",
        "StartTime",
        "EndTime",
        "Capacity",
        "IsBookingRequired",
        "IsCancelled",
        "CreatedAtUtc",
        "AssignedStaffId",
        "RecurringSeriesId",
    ]

    values: list[Any] = [
        "WEEKLY",
        row.category,
        row.sub_category,
        row.activity_category,
        row.venue_name,
        row.activity_name,
        f"Imported from {IMPORT_SOURCE_NAME}; Source sheet: {row.source_sheet}; Raw time: {row.raw_time}",
        0,
        None,
        row.session_date.isoformat(),
        None,
        row.start_time,
        row.end_time,
        None,
        0,
        1 if row.status == "cancelled" else 0,
        created_at,
        None,
        None,
    ]

    if not session_identity:
        if manual_session_id is None:
            raise RuntimeError("manual_session_id is required because SessionId is not IDENTITY.")
        columns.insert(0, "SessionId")
        values.insert(0, manual_session_id)

    column_sql = ", ".join(f"[{column}]" for column in columns)
    placeholders = ", ".join("?" for _ in values)

    sql = f"""
    INSERT INTO dbo.Sessions ({column_sql})
    OUTPUT INSERTED.SessionId
    VALUES ({placeholders});
    """

    inserted = cursor.execute(sql, *values).fetchone()
    if not inserted:
        raise RuntimeError("Session INSERT did not return a SessionId.")

    return int(inserted.SessionId)


def update_session_cancelled(cursor: pyodbc.Cursor, session_id: int) -> None:
    cursor.execute("UPDATE dbo.Sessions SET IsCancelled = 1 WHERE SessionId = ?;", session_id)


def find_participant_by_card(cursor: pyodbc.Cursor, card_number: Optional[str]) -> Optional[dict[str, Any]]:
    if not card_number:
        return None

    sql = """
    SELECT TOP 1
        ParticipantId,
        CONVERT(varchar(100), SaheliCardNumber) AS SaheliCardNumber,
        FullName,
        MobileNumber
    FROM dbo.Participants
    WHERE LTRIM(RTRIM(CONVERT(varchar(100), SaheliCardNumber))) = ?
    ORDER BY ParticipantId;
    """
    found = cursor.execute(sql, card_number).fetchone()

    if not found:
        return None

    return {
        "ParticipantId": int(found.ParticipantId),
        "SaheliCardNumber": str(found.SaheliCardNumber).strip(),
        "FullName": clean_text(found.FullName),
        "MobileNumber": normalise_phone(found.MobileNumber),
    }


def find_existing_lite_member_by_original_card(cursor: pyodbc.Cursor, original_card: Optional[str]) -> Optional[dict[str, Any]]:
    if not original_card:
        return None

    search_marker = f"Original Excel card value: [{original_card}]"
    sql = """
    SELECT TOP 1
        LiteMemberId,
        MemberDisplayId,
        MemberName,
        Phone,
        EmergencyName,
        EmergencyPhone
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND LiteMemberId IS NOT NULL
      AND Notes LIKE '%' + ? + '%'
    ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
    """
    found = cursor.execute(sql, search_marker).fetchone()

    if not found:
        return None

    return {
        "LiteMemberId": str(found.LiteMemberId),
        "MemberDisplayId": clean_text(found.MemberDisplayId),
        "MemberName": clean_text(found.MemberName),
        "Phone": normalise_phone(found.Phone),
        "EmergencyName": clean_text(found.EmergencyName),
        "EmergencyPhone": normalise_phone(found.EmergencyPhone),
    }


def find_existing_lite_member_by_name(
    cursor: pyodbc.Cursor,
    member_name: Optional[str],
    emergency_phone: Optional[str],
) -> Optional[dict[str, Any]]:
    if not member_name:
        return None

    if emergency_phone:
        sql = """
        SELECT TOP 1
            LiteMemberId,
            MemberDisplayId,
            MemberName,
            Phone,
            EmergencyName,
            EmergencyPhone
        FROM dbo.SessionAttendance
        WHERE AttendanceMemberKind = 'LITE'
          AND LiteMemberId IS NOT NULL
          AND LOWER(LTRIM(RTRIM(ISNULL(MemberName, '')))) = LOWER(LTRIM(RTRIM(?)))
          AND LTRIM(RTRIM(ISNULL(EmergencyPhone, ''))) = LTRIM(RTRIM(?))
        ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
        """
        found = cursor.execute(sql, member_name, emergency_phone).fetchone()
        if found:
            return {
                "LiteMemberId": str(found.LiteMemberId),
                "MemberDisplayId": clean_text(found.MemberDisplayId),
                "MemberName": clean_text(found.MemberName),
                "Phone": normalise_phone(found.Phone),
                "EmergencyName": clean_text(found.EmergencyName),
                "EmergencyPhone": normalise_phone(found.EmergencyPhone),
            }

    sql = """
    SELECT TOP 1
        LiteMemberId,
        MemberDisplayId,
        MemberName,
        Phone,
        EmergencyName,
        EmergencyPhone
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND LiteMemberId IS NOT NULL
      AND LOWER(LTRIM(RTRIM(ISNULL(MemberName, '')))) = LOWER(LTRIM(RTRIM(?)))
    ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
    """
    found = cursor.execute(sql, member_name).fetchone()

    if not found:
        return None

    return {
        "LiteMemberId": str(found.LiteMemberId),
        "MemberDisplayId": clean_text(found.MemberDisplayId),
        "MemberName": clean_text(found.MemberName),
        "Phone": normalise_phone(found.Phone),
        "EmergencyName": clean_text(found.EmergencyName),
        "EmergencyPhone": normalise_phone(found.EmergencyPhone),
    }


def attendance_exists(
    cursor: pyodbc.Cursor,
    session_id: int,
    participant_id: Optional[int],
    lite_member_id: Optional[str],
) -> bool:
    if participant_id is not None:
        sql = """
        SELECT TOP 1 AttendanceId
        FROM dbo.SessionAttendance
        WHERE SessionId = ? AND ParticipantId = ?;
        """
        return cursor.execute(sql, session_id, participant_id).fetchone() is not None

    if lite_member_id:
        sql = """
        SELECT TOP 1 AttendanceId
        FROM dbo.SessionAttendance
        WHERE SessionId = ? AND LiteMemberId = ?;
        """
        return cursor.execute(sql, session_id, lite_member_id).fetchone() is not None

    return False


def build_attendance_notes(row: ParsedRow, member_kind: str) -> str:
    notes = [
        f"Imported from {IMPORT_SOURCE_NAME}",
        f"Source sheet: {row.source_sheet}",
        f"Source row: {row.source_row}",
        f"Raw time: {row.raw_time}",
    ]

    if row.original_card_value and member_kind == "LITE":
        notes.append(f"Original Excel card value: [{row.original_card_value}]")

    return "; ".join(notes)


def create_attendance(
    cursor: pyodbc.Cursor,
    row: ParsedRow,
    session_id: int,
    participant_id: Optional[int],
    saheli_card_number: Optional[str],
    attendance_member_kind: str,
    lite_member_id: Optional[str],
    member_display_id: str,
    member_name: str,
    phone: Optional[str],
    emergency_name: Optional[str],
    emergency_phone: Optional[str],
    created_at: datetime,
    attendance_identity: bool,
    manual_attendance_id: Optional[int],
) -> int:
    columns = [
        "SessionId",
        "ParticipantId",
        "SessionName",
        "SessionDay",
        "SessionDate",
        "SessionMonth",
        "SessionStartTime",
        "SessionEndTime",
        "SaheliCardNumber",
        "RiskStratification",
        "Attended",
        "CheckInTime",
        "CheckOutTime",
        "Notes",
        "CreatedAtUtc",
        "UpdatedAtUtc",
        "AttendanceMemberKind",
        "LiteMemberId",
        "MemberDisplayId",
        "MemberName",
        "Phone",
        "EmergencyName",
        "EmergencyPhone",
        "AF",
        "BP",
        "HeightCm",
        "WeightKg",
    ]

    values: list[Any] = [
        session_id,
        participant_id,
        row.activity_name,
        row.session_day,
        row.session_date.isoformat(),
        row.session_month,
        row.start_time,
        row.end_time,
        saheli_card_number,
        row.risk_stratification,
        1,
        row.start_time,
        row.end_time,
        build_attendance_notes(row, attendance_member_kind),
        created_at,
        created_at,
        attendance_member_kind,
        lite_member_id,
        member_display_id,
        member_name,
        phone,
        emergency_name,
        emergency_phone,
        None,
        None,
        None,
        None,
    ]

    if not attendance_identity:
        if manual_attendance_id is None:
            raise RuntimeError("manual_attendance_id is required because AttendanceId is not IDENTITY.")
        columns.insert(0, "AttendanceId")
        values.insert(0, manual_attendance_id)

    column_sql = ", ".join(f"[{column}]" for column in columns)
    placeholders = ", ".join("?" for _ in values)

    sql = f"""
    INSERT INTO dbo.SessionAttendance ({column_sql})
    OUTPUT INSERTED.AttendanceId
    VALUES ({placeholders});
    """

    inserted = cursor.execute(sql, *values).fetchone()
    if not inserted:
        raise RuntimeError("Attendance INSERT did not return an AttendanceId.")

    return int(inserted.AttendanceId)


# ============================================================
# DATABASE IMPORT
# ============================================================

def import_into_database(parsed_rows: list[ParsedRow], *, apply_changes: bool) -> None:
    if pyodbc is None:
        raise RuntimeError("pyodbc is not installed. Run: pip install pyodbc")

    connection_string = build_connection_string()
    created_at = datetime.now(timezone.utc).replace(microsecond=0, tzinfo=None)

    counters: Counter[str] = Counter()
    sheet_counters: dict[str, Counter[str]] = defaultdict(Counter)
    session_id_cache: dict[tuple[str, str, str, str, str], int] = {}
    lite_member_cache: dict[str, dict[str, Any]] = {}

    with pyodbc.connect(connection_string) as conn:
        conn.autocommit = False
        cursor = conn.cursor()

        try:
            session_identity = is_identity_column(cursor, "Sessions", "SessionId")
            attendance_identity = is_identity_column(cursor, "SessionAttendance", "AttendanceId")

            next_manual_session_id = None if session_identity else get_next_manual_id(cursor, "Sessions", "SessionId")
            next_manual_attendance_id = None if attendance_identity else get_next_manual_id(cursor, "SessionAttendance", "AttendanceId")
            next_lite_display_number = get_next_lite_display_number(cursor)

            print(f"SessionId IDENTITY: {session_identity}")
            print(f"AttendanceId IDENTITY: {attendance_identity}")
            print(f"Next LITE MemberDisplayId: LITE-{next_lite_display_number}")

            for row in parsed_rows:
                per_sheet = sheet_counters[row.source_sheet]

                # --------------------------------------------------------
                # Create or reuse the session
                # --------------------------------------------------------
                session_id = session_id_cache.get(row.session_key)

                if session_id is None:
                    existing_session_id = find_existing_session(cursor, row) if REUSE_EXISTING_SESSIONS else None

                    if existing_session_id is not None:
                        session_id = existing_session_id
                        counters["sessions_reused"] += 1
                        per_sheet["sessions_reused"] += 1
                    else:
                        manual_session_id = None
                        if not session_identity:
                            manual_session_id = next_manual_session_id
                            assert next_manual_session_id is not None
                            next_manual_session_id += 1

                        session_id = create_session(
                            cursor=cursor,
                            row=row,
                            created_at=created_at,
                            session_identity=session_identity,
                            manual_session_id=manual_session_id,
                        )
                        counters["sessions_created"] += 1
                        per_sheet["sessions_created"] += 1

                    session_id_cache[row.session_key] = session_id

                if row.status == "cancelled":
                    update_session_cancelled(cursor, session_id)
                    counters["sessions_marked_cancelled"] += 1
                    per_sheet["sessions_marked_cancelled"] += 1
                    continue

                if row.status == "no_attendance" or not row.has_attendance_person:
                    counters["non_attendance_rows_skipped"] += 1
                    per_sheet["non_attendance_rows_skipped"] += 1
                    continue

                # --------------------------------------------------------
                # Resolve FULL or LITE member key
                # --------------------------------------------------------
                participant = find_participant_by_card(cursor, row.card_number)

                if participant:
                    participant_id = participant["ParticipantId"]
                    saheli_card_number = participant["SaheliCardNumber"]
                    attendance_member_kind = "FULL"
                    lite_member_id = None
                    member_display_id = saheli_card_number
                    member_name = row.member_name or participant["FullName"] or f"Card {saheli_card_number}"
                    phone = participant["MobileNumber"]
                    emergency_name = row.emergency_name
                    emergency_phone = row.emergency_phone
                    counters["full_rows"] += 1
                    per_sheet["full_rows"] += 1
                else:
                    participant_id = None
                    saheli_card_number = None
                    attendance_member_kind = "LITE"

                    member_name = row.member_name
                    if not member_name:
                        if row.original_card_value:
                            member_name = f"Card {row.original_card_value}"
                        else:
                            counters["attendance_rows_skipped_no_name"] += 1
                            per_sheet["attendance_rows_skipped_no_name"] += 1
                            continue

                    if row.original_card_value:
                        lite_cache_key = f"CARD|{row.original_card_value.casefold()}"
                    else:
                        lite_cache_key = f"NAME|{member_name.casefold()}|{row.emergency_phone or ''}"

                    lite_member = lite_member_cache.get(lite_cache_key)

                    if lite_member is None:
                        lite_member = find_existing_lite_member_by_original_card(cursor, row.original_card_value)

                    if lite_member is None:
                        lite_member = find_existing_lite_member_by_name(cursor, member_name, row.emergency_phone)

                    if lite_member is None:
                        lite_member = {
                            "LiteMemberId": str(uuid.uuid4()).upper(),
                            "MemberDisplayId": f"LITE-{next_lite_display_number}",
                            "MemberName": member_name,
                            "Phone": None,
                            "EmergencyName": row.emergency_name,
                            "EmergencyPhone": row.emergency_phone,
                        }
                        next_lite_display_number += 1
                        counters["lite_members_created"] += 1
                        per_sheet["lite_members_created"] += 1
                    else:
                        counters["lite_members_reused"] += 1
                        per_sheet["lite_members_reused"] += 1

                    lite_member_cache[lite_cache_key] = lite_member

                    lite_member_id = lite_member["LiteMemberId"]
                    member_display_id = lite_member["MemberDisplayId"]
                    member_name = member_name or lite_member["MemberName"]
                    phone = lite_member.get("Phone")
                    emergency_name = row.emergency_name or lite_member.get("EmergencyName")
                    emergency_phone = row.emergency_phone or lite_member.get("EmergencyPhone")

                    counters["lite_rows"] += 1
                    per_sheet["lite_rows"] += 1

                # --------------------------------------------------------
                # Avoid duplicate attendance rows
                # --------------------------------------------------------
                if SKIP_DUPLICATE_ATTENDANCE and attendance_exists(
                    cursor=cursor,
                    session_id=session_id,
                    participant_id=participant_id,
                    lite_member_id=lite_member_id,
                ):
                    counters["duplicate_attendance_skipped"] += 1
                    per_sheet["duplicate_attendance_skipped"] += 1
                    continue

                manual_attendance_id = None
                if not attendance_identity:
                    manual_attendance_id = next_manual_attendance_id
                    assert next_manual_attendance_id is not None
                    next_manual_attendance_id += 1

                create_attendance(
                    cursor=cursor,
                    row=row,
                    session_id=session_id,
                    participant_id=participant_id,
                    saheli_card_number=saheli_card_number,
                    attendance_member_kind=attendance_member_kind,
                    lite_member_id=lite_member_id,
                    member_display_id=member_display_id,
                    member_name=member_name,
                    phone=phone,
                    emergency_name=emergency_name,
                    emergency_phone=emergency_phone,
                    created_at=created_at,
                    attendance_identity=attendance_identity,
                    manual_attendance_id=manual_attendance_id,
                )

                counters["attendance_created"] += 1
                per_sheet["attendance_created"] += 1

            if apply_changes:
                conn.commit()
                mode_message = "COMMITTED: database changes were saved."
            else:
                conn.rollback()
                mode_message = "DRY RUN: transaction rolled back. No database changes were saved."

        except Exception:
            conn.rollback()
            raise

    print("\nDatabase import summary")
    print("=" * 100)
    print(mode_message)
    print("-" * 100)

    ordered_counter_keys = [
        "sessions_created",
        "sessions_reused",
        "sessions_marked_cancelled",
        "attendance_created",
        "duplicate_attendance_skipped",
        "full_rows",
        "lite_rows",
        "lite_members_created",
        "lite_members_reused",
        "attendance_rows_skipped_no_name",
        "non_attendance_rows_skipped",
    ]

    for key in ordered_counter_keys:
        print(f"{key.replace('_', ' ').title():42}: {counters[key]}")

    print("\nBy worksheet")
    print("-" * 100)
    for sheet_name, per_sheet in sheet_counters.items():
        print(f"\n{sheet_name}")
        for key in ordered_counter_keys:
            if per_sheet[key]:
                print(f"  {key.replace('_', ' ').title():40}: {per_sheet[key]}")


# ============================================================
# COMMAND-LINE ENTRY POINT
# ============================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Mens Sessions 2026 into CRM Sessions and SessionAttendance tables.")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_FILE,
        help=f"Excel workbook path. Default: {DEFAULT_INPUT_FILE}",
    )
    parser.add_argument(
        "--inspect-only",
        action="store_true",
        help="Parse Excel and show worksheet/session counts without connecting to SQL Server.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Commit inserts and updates. Without this option the SQL transaction is rolled back.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_file: Path = args.input

    if not input_file.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_file}")

    print(f"Reading workbook: {input_file}")
    parsed_rows, stats = parse_workbook(input_file)
    print_excel_inspection(stats)

    print("Venue mappings used")
    print("=" * 90)
    for worksheet, config in SHEET_CONFIG.items():
        print(f"{worksheet:34} -> {config['activity_name']:26} | {config['venue_name']}")
    print()

    if args.inspect_only:
        print("Inspection complete. No SQL connection was made.")
        return

    import_into_database(parsed_rows, apply_changes=args.apply)

    if not args.apply:
        print("\nReview the summary. To save the records, run the same command with --apply.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        raise
