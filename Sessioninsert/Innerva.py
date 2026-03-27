import re
from datetime import datetime
from copy import copy
from openpyxl import load_workbook

INPUT_FILE = r"C:\Users\shonk\Downloads\Innerva Booking Sheet.xlsx"
OUTPUT_FILE = r"C:\Users\shonk\Downloads\Innerva Booking Sheet_2026_fixed_dates.xlsx"
SHEET_NAME = "2026"
YEAR_VALUE = 2026

MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

FILL_DOWN_HEADERS = [
    "Lead",
    "Session Type",
    "Day",
    "date",
    "Month",
    "Induction Time",
    "Session",
]

def clean_text(v):
    if v is None:
        return ""
    return str(v).strip()

def extract_day_number(day_text):
    text = clean_text(day_text).lower()
    if not text:
        return None
    m = re.search(r"(\d{1,2})", text)
    if not m:
        return None
    d = int(m.group(1))
    return d if 1 <= d <= 31 else None

def extract_month_number(month_text):
    text = clean_text(month_text).lower()
    if not text:
        return None
    if text in MONTHS:
        return MONTHS[text]
    for month_name, month_num in MONTHS.items():
        if month_name.startswith(text):
            return month_num
    return None

def is_blank(value):
    return value is None or str(value).strip() == ""

wb = load_workbook(INPUT_FILE)
ws = wb[SHEET_NAME]

# ----------------------------
# read headers
# ----------------------------
header_row = 1
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(header_row, col).value
    if val is not None:
        headers[str(val).strip().lower()] = col

required = ["date", "month"]
for h in required:
    if h not in headers:
        raise ValueError(f"Missing required column: {h}")

# ----------------------------
# fill down grouped columns
# ----------------------------
for header_name in FILL_DOWN_HEADERS:
    key = header_name.strip().lower()
    if key not in headers:
        continue

    col_idx = headers[key]
    last_value = None

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, col_idx)
        val = cell.value

        if not is_blank(val):
            last_value = val
        else:
            # only fill down if we already have something above
            if last_value is not None:
                cell.value = last_value

# ----------------------------
# add Actual Date column
# ----------------------------
actual_date_col = None
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if str(val).strip().lower() == "actual date":
        actual_date_col = col
        break

if actual_date_col is None:
    actual_date_col = ws.max_column + 1
    ws.cell(1, actual_date_col).value = "Actual Date"

date_col = headers["date"]
month_col = headers["month"]

for row in range(2, ws.max_row + 1):
    day_text = ws.cell(row, date_col).value
    month_text = ws.cell(row, month_col).value

    day_num = extract_day_number(day_text)
    month_num = extract_month_number(month_text)

    out_cell = ws.cell(row, actual_date_col)

    if day_num is None or month_num is None:
        out_cell.value = None
        continue

    try:
        real_date = datetime(YEAR_VALUE, month_num, day_num)
        out_cell.value = real_date
        out_cell.number_format = "dd/mm/yyyy"
    except ValueError:
        out_cell.value = None

# ----------------------------
# make new column a bit wider
# ----------------------------
from openpyxl.utils import get_column_letter
ws.column_dimensions[get_column_letter(actual_date_col)].width = 15

wb.save(OUTPUT_FILE)
print(f"Done. Saved to:\n{OUTPUT_FILE}")