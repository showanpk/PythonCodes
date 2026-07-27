from __future__ import annotations

"""
Import the workbook "2025 Register for Exercise - ARCC (1).xlsx" into:
    dbo.Sessions
    dbo.SessionAttendance

Safety:
    - The script runs as a dry run by default and rolls back all SQL changes.
    - Add --apply only after checking the dry-run summary.
    - Add --inspect-only to check the Excel extraction without connecting to SQL Server.

Required packages:
    pip install openpyxl pyodbc

Required environment variables:
    CRM_SQL_USER
    CRM_SQL_PASSWORD

Optional environment variables:
    CRM_SQL_SERVER      default: tcp:sahelihub.database.windows.net,1433
    CRM_SQL_DATABASE    default: SahelihubCRM
    CRM_SQL_CONNECTION_STRING  full override if required
"""

import argparse
import os
import re
import sys
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

try:
    import pyodbc
except ModuleNotFoundError:
    pyodbc = None  # type: ignore[assignment]

from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

DEFAULT_INPUT_FILE = Path(r"C:\Users\shonk\Downloads\2025 Register for Exercise - ARCC  (1).xlsx")
IMPORT_SOURCE_NAME = "ARCC Exercise Register 2025"
DEFAULT_VENUE_NAME = "Alum Rock Community Centre"

REUSE_EXISTING_SESSIONS = True
SKIP_DUPLICATE_ATTENDANCE = True

NULL_TEXT_VALUES = {"", "#N/A", "N/A", "NA", "NONE", "NULL", "NAN", "(BLANK)", "-", "--"}
YES_VALUES = {"yes", "y", "true", "1", "present", "attended", "✓"}

ACTIVITY_DEFAULTS: dict[str, tuple[time, time]] = {
    "Men's Exercise": (time(9, 15), time(10, 15)),
    "Chair Based Exercise": (time(12, 30), time(13, 30)),
    "Chair Pilates": (time(12, 30), time(13, 30)),
    "Yoga": (time(12, 30), time(13, 30)),
    "Circuit": (time(10, 0), time(11, 0)),
    "Art Social": (time(9, 30), time(12, 0)),
    "Crochet": (time(9, 30), time(10, 30)),
    "Salsa": (time(11, 0), time(12, 0)),
    "Strength & Stretch": (time(9, 45), time(10, 45)),
    "Pilates": (time(11, 0), time(12, 0)),
    "Body Conditioning": (time(9, 45), time(10, 45)),
    "Saheli Social": (time(10, 30), time(12, 0)),
    "Walk & Talk": (time(10, 0), time(11, 0)),
    "Self Defence": (time(17, 0), time(18, 0)),
}


# ============================================================
# DATA MODEL
# ============================================================

@dataclass(frozen=True)
class ParsedAttendanceRow:
    source_sheet: str
    source_sheets: str
    source_row: int
    source_col: int

    venue_name: str
    activity_name: str
    raw_activity_header: str
    category: str
    sub_category: str
    activity_category: str

    session_date: date
    start_time: str
    end_time: str

    saheli_card_number: Optional[str]
    original_card_value: Optional[str]
    wellbeing_card_number: Optional[str]
    member_name: Optional[str]
    gender: Optional[str]
    dob: Optional[str]
    postcode: Optional[str]
    emergency_name: Optional[str]
    emergency_phone: Optional[str]
    risk_stratification: Optional[str]
    comments: Optional[str]

    @property
    def session_key(self) -> tuple[str, str, str, str, str]:
        return (
            self.venue_name.casefold().strip(),
            self.activity_name.casefold().strip(),
            self.session_date.isoformat(),
            self.start_time,
            self.end_time,
        )

    @property
    def session_day(self) -> str:
        return self.session_date.strftime("%A")

    @property
    def session_month(self) -> str:
        return self.session_date.strftime("%B")


# ============================================================
# CLEANING / PARSING HELPERS
# ============================================================

def clean_text(value: Any, *, zero_is_null: bool = True) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        text = value.date().isoformat()
    elif isinstance(value, date):
        text = value.isoformat()
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()

    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()

    if text.upper() in NULL_TEXT_VALUES:
        return None

    if zero_is_null and text in {"0", "00", "0.0"}:
        return None

    return text or None


def normalise_phone(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None

    if text.startswith("+"):
        return "+" + re.sub(r"\D", "", text[1:])

    digits = re.sub(r"\D", "", text)
    return digits or None


def is_attended(value: Any) -> bool:
    text = clean_text(value)
    return bool(text and text.strip().casefold() in YES_VALUES)


def normalise_card(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None

    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()

    # The workbook sometimes records SAH-1, while the CRM usually stores 1.
    if text.casefold().startswith("sah-"):
        digits = re.sub(r"\D", "", text)
        return digits or text

    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]

    return text


def canonical_activity(raw: Any = None, sheet_title: Optional[str] = None) -> str:
    source = clean_text(raw) or clean_text(sheet_title) or "ARCC Activity"
    text = source.casefold()

    if "mens exercise" in text or "men's exercise" in text:
        return "Men's Exercise"
    if "chair pilates" in text:
        return "Chair Pilates"
    if "chair" in text and ("exercise" in text or "based" in text or "class" in text):
        return "Chair Based Exercise"
    if "pilates" in text:
        return "Pilates"
    if "body conditioning" in text:
        return "Body Conditioning"
    if "social" in text:
        if "saheli" in text or "morning" in text or re.search(r"\d", text):
            return "Saheli Social"
        return "Art Social"
    if "art social" in text:
        return "Art Social"
    if "circuit" in text:
        return "Circuit"
    if "strength" in text and "stretch" in text:
        return "Strength & Stretch"
    if "yoga" in text:
        return "Yoga"
    if "walk" in text:
        return "Walk & Talk"
    if "self defence" in text or "self defense" in text:
        return "Self Defence"
    if "salsa" in text:
        return "Salsa"
    if "crochet" in text:
        return "Crochet"

    return re.sub(r"\s+", " ", source).strip().title()


def category_for_activity(activity_name: str) -> tuple[str, str, str]:
    text = activity_name.casefold()

    if any(token in text for token in [
        "exercise",
        "pilates",
        "conditioning",
        "circuit",
        "stretch",
        "yoga",
        "walk",
        "self defence",
        "salsa",
    ]):
        return "Fitness", "Health & Wellbeing", "Physical Activity"

    if any(token in text for token in ["social", "crochet", "art"]):
        return "Community", "Health & Wellbeing", "Social / Wellbeing"

    return "Health & Wellbeing", "Health & Wellbeing", "Activity"


def parse_date_value(value: Any, *, last_date: Optional[date] = None) -> Optional[date]:
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = clean_text(value, zero_is_null=False)
        if not text:
            return None

        iso = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
        if iso:
            year, month, day = map(int, iso.groups())
            try:
                parsed = date(year, month, day)
            except ValueError:
                return None
        else:
            slash = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
            if slash:
                month, day, year_text = slash.groups()
                if len(year_text) == 2:
                    year_text = "20" + year_text
                try:
                    parsed = date(int(year_text), int(month), int(day))
                except ValueError:
                    return None
            elif isinstance(value, (int, float)) and 30000 < value < 60000:
                parsed = datetime(1899, 12, 30).date() + timedelta(days=int(value))
            else:
                return None

    # Correct obvious workbook typos in the late-2025 activity sheets.
    if parsed.year == 2026 and parsed.month in {9, 10, 11, 12}:
        parsed = date(2025, parsed.month, parsed.day)

    # In the Yoga sheet, 04/12/2025 appears inside the Nov/Dec sequence and is treated as 04 Dec 2025.
    if parsed.year == 2025 and parsed.month == 4 and parsed.day == 12 and last_date and last_date.month == 11:
        parsed = date(2025, 12, 4)

    return parsed


def parse_one_time(token: str, suffix_hint: Optional[str] = None) -> Optional[time]:
    text = token.strip().lower()
    text = text.replace("a.m.", "am").replace("p.m.", "pm")
    text = text.replace("a.m", "am").replace("p.m", "pm")
    text = text.replace(".", ":").replace(" ", "")

    match = re.match(r"^(\d{1,2})(?::(\d{1,2}))?(am|pm)?$", text)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    suffix = match.group(3) or suffix_hint

    if hour > 23 or minute > 59:
        return None

    if suffix == "pm" and hour < 12:
        hour += 12
    if suffix == "am" and hour == 12:
        hour = 0

    if hour > 23:
        return None

    return time(hour, minute)


def parse_time_range(
    text_value: Any,
    default_start: time,
    default_end: time,
    *,
    default_minutes: int = 60,
) -> tuple[time, time]:
    if isinstance(text_value, (datetime, date)):
        return default_start, default_end

    text = clean_text(text_value, zero_is_null=False)
    if not text:
        return default_start, default_end

    text = text.casefold()
    text = re.sub(r"\d{4}-\d{1,2}-\d{1,2}", " ", text)
    text = re.sub(r"\d{1,2}/\d{1,2}/\d{2,4}", " ", text)

    text = text.replace("a.m.", "am").replace("p.m.", "pm")
    text = text.replace("a.m", "am").replace("p.m", "pm")
    text = text.replace("noon", "12pm")
    text = text.replace("–", "-").replace("—", "-")

    range_match = re.search(
        r"(\d{1,2}(?:[.:]\d{1,2})?\s*(?:am|pm)?)\s*-\s*(\d{1,2}(?:[.:]\d{1,2})?\s*(?:am|pm)?)",
        text,
    )

    if range_match:
        left = range_match.group(1)
        right = range_match.group(2)

        suffix_hint = None
        if "pm" in right:
            suffix_hint = "pm"
        elif "am" in right:
            suffix_hint = "am"

        start_hint = suffix_hint if suffix_hint == "pm" and re.search(r"^\s*[1-7](?:[.:]\d{1,2})?\s*$", left) else None
        start = parse_one_time(left, start_hint)
        end = parse_one_time(right, suffix_hint)

        if start and end:
            return start, end

    single_time_matches = list(
        re.finditer(r"(?<!\d)(\d{1,2}(?:[.:]\d{1,2})?\s*(?:am|pm))(?!\d)", text)
    )

    if single_time_matches:
        start = parse_one_time(single_time_matches[-1].group(1))
        if start:
            end_dt = datetime.combine(date(2000, 1, 1), start) + timedelta(minutes=default_minutes)
            return start, end_dt.time()

    return default_start, default_end


# ============================================================
# WORKBOOK EXTRACTION
# ============================================================

def is_month_sheet(ws: Any) -> bool:
    cell_text = clean_text(ws.cell(3, 1).value, zero_is_null=False)
    return bool(cell_text and "sessions attended" in cell_text.casefold())


def is_activity_sheet(ws: Any) -> bool:
    cell_text = clean_text(ws.cell(2, 1).value, zero_is_null=False)
    return bool(cell_text and "sessions attended" in cell_text.casefold())


def actual_max_col(ws: Any, rows: tuple[int, ...], max_cap: int = 300) -> int:
    max_col = 0

    for col in range(1, min(ws.max_column, max_cap) + 1):
        if any(clean_text(ws.cell(row, col).value, zero_is_null=False) is not None for row in rows):
            max_col = col

    return max_col


def build_meta_map(ws: Any, header_row: int, start_col: int) -> dict[str, int]:
    mapping: dict[str, int] = {}

    for col in range(1, start_col):
        header = clean_text(ws.cell(header_row, col).value, zero_is_null=False)
        if not header:
            continue

        text = header.casefold().strip()

        if "saheli" in text and "card" in text:
            mapping["saheli_card_number"] = col
        elif text == "name":
            mapping["name"] = col
        elif "wellbeing" in text and "card" in text:
            mapping["wellbeing_card_number"] = col
        elif text == "gender":
            mapping["gender"] = col
        elif "d.o.b" in text or "dob" in text:
            mapping["dob"] = col
        elif "post" in text and "code" in text:
            mapping["postcode"] = col
        elif "emergency contact name" in text:
            mapping["emergency_name"] = col
        elif "emergency number" in text:
            mapping["emergency_phone"] = col
        elif "risk" in text:
            mapping["risk_stratification"] = col
        elif "comments" in text or "quotes" in text:
            mapping["comments"] = col

    return mapping


def read_participant_from_row(ws: Any, row: int, meta_map: dict[str, int]) -> dict[str, Optional[str]]:
    def get_value(key: str) -> Any:
        col = meta_map.get(key)
        return ws.cell(row, col).value if col else None

    return {
        "saheli_card_number": normalise_card(get_value("saheli_card_number")),
        "original_card_value": clean_text(get_value("saheli_card_number")),
        "wellbeing_card_number": clean_text(get_value("wellbeing_card_number")),
        "member_name": clean_text(get_value("name")),
        "gender": clean_text(get_value("gender")),
        "dob": clean_text(get_value("dob")),
        "postcode": clean_text(get_value("postcode")),
        "emergency_name": clean_text(get_value("emergency_name")),
        "emergency_phone": normalise_phone(get_value("emergency_phone")),
        "risk_stratification": clean_text(get_value("risk_stratification")),
        "comments": clean_text(get_value("comments")),
    }


def participant_quality_score(row: dict[str, Any]) -> int:
    fields = [
        "member_name",
        "gender",
        "dob",
        "postcode",
        "emergency_name",
        "emergency_phone",
        "risk_stratification",
        "wellbeing_card_number",
    ]
    return sum(1 for field in fields if row.get(field))


def merge_attendance(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    if participant_quality_score(incoming) > participant_quality_score(existing):
        merged = incoming.copy()
        other = existing
    else:
        merged = existing.copy()
        other = incoming

    for key, value in other.items():
        if not merged.get(key) and value:
            merged[key] = value

    sources = set()
    for item in (existing, incoming):
        if item.get("source_sheets_set"):
            sources.update(item["source_sheets_set"])
        elif item.get("source_sheet"):
            sources.add(item["source_sheet"])
    merged["source_sheets_set"] = sources
    return merged


def parse_workbook(input_path: Path) -> tuple[list[ParsedAttendanceRow], Counter[str], list[dict[str, Any]]]:
    workbook = load_workbook(input_path, data_only=True)

    sessions: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    attendance: dict[tuple[str, str, str, str, str, str], dict[str, Any]] = {}
    issues: list[dict[str, Any]] = []

    def add_attendance(session: dict[str, Any], person: dict[str, Any], marker: Any, source_row: int) -> None:
        participant_key = person.get("saheli_card_number") or (person.get("member_name") or "").casefold().strip()

        if not participant_key:
            issues.append({
                "sheet": session["source_sheet"],
                "row": source_row,
                "col": session["source_col"],
                "issue": "Attendance marked yes but no card/name found",
            })
            return

        session_key = (
            session["venue_name"].casefold().strip(),
            session["activity_name"].casefold().strip(),
            session["session_date"].isoformat(),
            session["start_time"],
            session["end_time"],
        )

        if session_key not in sessions:
            sessions[session_key] = {
                **session,
                "source_sheets_set": {session["source_sheet"]},
                "attendance_count": 0,
            }
        else:
            sessions[session_key]["source_sheets_set"].add(session["source_sheet"])

        attendance_key = session_key + (participant_key,)

        new_row = {
            **session,
            **person,
            "source_row": source_row,
            "attendance_marker": clean_text(marker, zero_is_null=False),
            "source_sheets_set": {session["source_sheet"]},
        }

        if attendance_key in attendance:
            attendance[attendance_key] = merge_attendance(attendance[attendance_key], new_row)
        else:
            attendance[attendance_key] = new_row
            sessions[session_key]["attendance_count"] += 1

    for ws in workbook.worksheets:
        sheet_title = ws.title.strip()

        if sheet_title == "Sheet1":
            continue

        if is_month_sheet(ws):
            header_row = 3
            date_row = 2
            max_col = actual_max_col(ws, (1, 2, 3))

            start_col = None
            for col in range(1, max_col + 1):
                raw_date = ws.cell(date_row, col).value
                raw_header = clean_text(ws.cell(header_row, col).value, zero_is_null=False)

                if col >= 7 and (parse_date_value(raw_date) or (raw_header and re.search(r"\d", raw_header))):
                    start_col = col
                    break

            if not start_col:
                issues.append({"sheet": sheet_title, "issue": "No session columns detected"})
                continue

            meta_map = build_meta_map(ws, header_row, start_col)
            last_date: Optional[date] = None

            for col in range(start_col, max_col + 1):
                raw_header = clean_text(ws.cell(header_row, col).value, zero_is_null=False)

                if not raw_header:
                    continue

                session_date = parse_date_value(ws.cell(date_row, col).value, last_date=last_date)

                # Some month sheets have a blank date in a repeated date group; carry forward the last date.
                if session_date is None and last_date is not None:
                    session_date = last_date

                if session_date is None:
                    issues.append({
                        "sheet": sheet_title,
                        "col": col,
                        "header": raw_header,
                        "issue": "Missing or invalid date",
                    })
                    continue

                last_date = session_date

                if session_date.year != 2025:
                    continue

                activity_name = canonical_activity(raw_header)
                default_start, default_end = ACTIVITY_DEFAULTS.get(activity_name, (time(10, 0), time(11, 0)))
                start_time, end_time = parse_time_range(raw_header, default_start, default_end)
                category, sub_category, activity_category = category_for_activity(activity_name)

                session = {
                    "source_sheet": sheet_title,
                    "source_col": col,
                    "venue_name": DEFAULT_VENUE_NAME,
                    "activity_name": activity_name,
                    "raw_activity_header": raw_header,
                    "category": category,
                    "sub_category": sub_category,
                    "activity_category": activity_category,
                    "session_date": session_date,
                    "start_time": start_time.strftime("%H:%M:%S"),
                    "end_time": end_time.strftime("%H:%M:%S"),
                }

                for row in range(header_row + 1, min(ws.max_row, 1200) + 1):
                    marker = ws.cell(row, col).value

                    if not is_attended(marker):
                        continue

                    person = read_participant_from_row(ws, row, meta_map)
                    add_attendance(session, person, marker, row)

        elif is_activity_sheet(ws):
            header_row = 2
            start_col = 7
            max_col = actual_max_col(ws, (1, 2))
            meta_map = build_meta_map(ws, header_row, start_col)

            activity_name = canonical_activity(sheet_title=sheet_title)
            default_start, default_end = ACTIVITY_DEFAULTS.get(activity_name, (time(10, 0), time(11, 0)))
            last_date = None

            for col in range(start_col, max_col + 1):
                raw_header = ws.cell(header_row, col).value
                session_date = parse_date_value(raw_header, last_date=last_date)

                if session_date is None:
                    continue

                if session_date.year != 2025:
                    continue

                last_date = session_date
                start_time, end_time = parse_time_range(raw_header, default_start, default_end)
                category, sub_category, activity_category = category_for_activity(activity_name)

                session = {
                    "source_sheet": sheet_title,
                    "source_col": col,
                    "venue_name": DEFAULT_VENUE_NAME,
                    "activity_name": activity_name,
                    "raw_activity_header": clean_text(raw_header, zero_is_null=False) or activity_name,
                    "category": category,
                    "sub_category": sub_category,
                    "activity_category": activity_category,
                    "session_date": session_date,
                    "start_time": start_time.strftime("%H:%M:%S"),
                    "end_time": end_time.strftime("%H:%M:%S"),
                }

                for row in range(header_row + 1, min(ws.max_row, 1200) + 1):
                    marker = ws.cell(row, col).value

                    if not is_attended(marker):
                        continue

                    person = read_participant_from_row(ws, row, meta_map)
                    add_attendance(session, person, marker, row)

        else:
            issues.append({"sheet": sheet_title, "issue": "Unrecognised sheet structure"})

    parsed_rows: list[ParsedAttendanceRow] = []

    for _, row in sorted(
        attendance.items(),
        key=lambda item: (
            item[1]["session_date"],
            item[1]["start_time"],
            item[1]["activity_name"],
            item[1].get("member_name") or item[1].get("saheli_card_number") or "",
        ),
    ):
        parsed_rows.append(
            ParsedAttendanceRow(
                source_sheet=row["source_sheet"],
                source_sheets=", ".join(sorted(row.get("source_sheets_set", {row["source_sheet"]}))),
                source_row=int(row["source_row"]),
                source_col=int(row["source_col"]),
                venue_name=row["venue_name"],
                activity_name=row["activity_name"],
                raw_activity_header=row["raw_activity_header"],
                category=row["category"],
                sub_category=row["sub_category"],
                activity_category=row["activity_category"],
                session_date=row["session_date"],
                start_time=row["start_time"],
                end_time=row["end_time"],
                saheli_card_number=row.get("saheli_card_number"),
                original_card_value=row.get("original_card_value"),
                wellbeing_card_number=row.get("wellbeing_card_number"),
                member_name=row.get("member_name"),
                gender=row.get("gender"),
                dob=row.get("dob"),
                postcode=row.get("postcode"),
                emergency_name=row.get("emergency_name"),
                emergency_phone=row.get("emergency_phone"),
                risk_stratification=row.get("risk_stratification"),
                comments=row.get("comments"),
            )
        )

    session_counter = Counter()
    for row in parsed_rows:
        session_counter[row.session_key] += 1

    return parsed_rows, session_counter, issues


def print_excel_inspection(parsed_rows: list[ParsedAttendanceRow], session_counter: Counter, issues: list[dict[str, Any]]) -> None:
    by_sheet = Counter(row.source_sheet for row in parsed_rows)
    by_activity = Counter(row.activity_name for row in parsed_rows)
    by_month = Counter(row.session_date.strftime("%Y-%m") for row in parsed_rows)

    print("\nExcel extraction summary")
    print("=" * 100)
    print(f"Unique sessions detected : {len(session_counter)}")
    print(f"Attendance rows detected : {len(parsed_rows)}")
    print(f"Issue rows skipped       : {len(issues)}")

    print("\nAttendance by month")
    print("-" * 100)
    for month, count in sorted(by_month.items()):
        print(f"{month:10} {count:5}")

    print("\nAttendance by activity")
    print("-" * 100)
    for activity, count in by_activity.most_common():
        print(f"{activity:30} {count:5}")

    print("\nRows used by source worksheet")
    print("-" * 100)
    for sheet, count in by_sheet.most_common():
        print(f"{sheet:30} {count:5}")

    if issues:
        print("\nIssues skipped")
        print("-" * 100)
        for issue in issues[:20]:
            print(issue)
        if len(issues) > 20:
            print(f"... plus {len(issues) - 20} more")


# ============================================================
# DATABASE HELPERS
# ============================================================

def build_connection_string() -> str:
    override = os.getenv("CRM_SQL_CONNECTION_STRING")
    if override:
        return override

    server = os.getenv("CRM_SQL_SERVER", "tcp:sahelihub.database.windows.net,1433")
    database = os.getenv("CRM_SQL_DATABASE", "SahelihubCRM")
    username = os.getenv("CRM_SQL_USER")
    password = os.getenv("CRM_SQL_PASSWORD")

    if not username or not password:
        raise RuntimeError(
            "Missing CRM_SQL_USER or CRM_SQL_PASSWORD environment variables. "
            "Set them before running without --inspect-only."
        )

    return (
        "DRIVER={ODBC Driver 18 for SQL Server};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
        "Encrypt=yes;"
        "TrustServerCertificate=no;"
        "Connection Timeout=30;"
    )


def is_identity_column(cursor: Any, table_name: str, column_name: str) -> bool:
    sql = """
    SELECT COLUMNPROPERTY(OBJECT_ID(?), ?, 'IsIdentity') AS IsIdentity;
    """
    row = cursor.execute(sql, f"dbo.{table_name}", column_name).fetchone()
    return bool(row and row.IsIdentity == 1)


def get_next_manual_id(cursor: Any, table_name: str, column_name: str) -> int:
    sql = f"SELECT ISNULL(MAX([{column_name}]), 0) + 1 AS NextId FROM dbo.[{table_name}];"
    row = cursor.execute(sql).fetchone()
    return int(row.NextId)


def get_next_lite_display_number(cursor: Any) -> int:
    sql = """
    SELECT MemberDisplayId
    FROM dbo.SessionAttendance
    WHERE MemberDisplayId LIKE 'LITE-%';
    """
    max_number = 0

    for row in cursor.execute(sql).fetchall():
        value = clean_text(row.MemberDisplayId)
        if not value:
            continue

        match = re.search(r"LITE-(\d+)", value, flags=re.IGNORECASE)
        if match:
            max_number = max(max_number, int(match.group(1)))

    return max_number + 1


def find_existing_session(cursor: Any, row: ParsedAttendanceRow) -> Optional[int]:
    sql = """
    SELECT TOP 1 SessionId
    FROM dbo.Sessions
    WHERE LTRIM(RTRIM(ISNULL(VenueName, ''))) = LTRIM(RTRIM(?))
      AND LTRIM(RTRIM(ISNULL(ActivityName, ''))) = LTRIM(RTRIM(?))
      AND CAST(SessionDate AS date) = ?
      AND CONVERT(varchar(8), TRY_CONVERT(time, StartTime), 108) = ?
      AND CONVERT(varchar(8), TRY_CONVERT(time, EndTime), 108) = ?
    ORDER BY SessionId;
    """

    found = cursor.execute(
        sql,
        row.venue_name,
        row.activity_name,
        row.session_date.isoformat(),
        row.start_time,
        row.end_time,
    ).fetchone()

    return int(found.SessionId) if found else None


def create_session(
    cursor: Any,
    row: ParsedAttendanceRow,
    created_at: datetime,
    session_identity: bool,
    manual_session_id: Optional[int],
) -> int:
    columns = [
        "Frequency",
        "Category",
        "SubCategory",
        "ActivityCategory",
        "VenueName",
        "ActivityName",
        "Notes",
        "IsRecurringWeekly",
        "DayOfWeek",
        "SessionDate",
        "ArrivalTime",
        "StartTime",
        "EndTime",
        "Capacity",
        "IsBookingRequired",
        "IsCancelled",
        "CreatedAtUtc",
        "AssignedStaffId",
        "RecurringSeriesId",
    ]

    values: list[Any] = [
        "WEEKLY",
        row.category,
        row.sub_category,
        row.activity_category,
        row.venue_name,
        row.activity_name,
        (
            f"Imported from {IMPORT_SOURCE_NAME}; "
            f"Source sheets: {row.source_sheets}; "
            f"Raw activity header: {row.raw_activity_header}"
        ),
        0,
        None,
        row.session_date.isoformat(),
        None,
        row.start_time,
        row.end_time,
        None,
        0,
        0,
        created_at,
        None,
        None,
    ]

    if not session_identity:
        if manual_session_id is None:
            raise RuntimeError("manual_session_id is required because SessionId is not IDENTITY.")
        columns.insert(0, "SessionId")
        values.insert(0, manual_session_id)

    column_sql = ", ".join(f"[{column}]" for column in columns)
    placeholders = ", ".join("?" for _ in values)

    sql = f"""
    INSERT INTO dbo.Sessions ({column_sql})
    OUTPUT INSERTED.SessionId
    VALUES ({placeholders});
    """

    inserted = cursor.execute(sql, *values).fetchone()
    if not inserted:
        raise RuntimeError("Session INSERT did not return a SessionId.")

    return int(inserted.SessionId)


def find_participant_by_card(cursor: Any, card_number: Optional[str]) -> Optional[dict[str, Any]]:
    if not card_number:
        return None

    sql = """
    SELECT TOP 1
        ParticipantId,
        CONVERT(varchar(100), SaheliCardNumber) AS SaheliCardNumber,
        FullName,
        MobileNumber
    FROM dbo.Participants
    WHERE LTRIM(RTRIM(CONVERT(varchar(100), SaheliCardNumber))) = ?
    ORDER BY ParticipantId;
    """

    found = cursor.execute(sql, card_number).fetchone()
    if not found:
        return None

    return {
        "ParticipantId": int(found.ParticipantId),
        "SaheliCardNumber": str(found.SaheliCardNumber).strip(),
        "FullName": clean_text(found.FullName),
        "MobileNumber": normalise_phone(found.MobileNumber),
    }


def find_existing_lite_member_by_original_card(cursor: Any, original_card: Optional[str]) -> Optional[dict[str, Any]]:
    if not original_card:
        return None

    marker = f"Original Excel card value: [{original_card}]"

    sql = """
    SELECT TOP 1
        LiteMemberId,
        MemberDisplayId,
        MemberName,
        Phone,
        EmergencyName,
        EmergencyPhone
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND LiteMemberId IS NOT NULL
      AND Notes LIKE '%' + ? + '%'
    ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
    """

    found = cursor.execute(sql, marker).fetchone()
    if not found:
        return None

    return {
        "LiteMemberId": str(found.LiteMemberId),
        "MemberDisplayId": clean_text(found.MemberDisplayId),
        "MemberName": clean_text(found.MemberName),
        "Phone": normalise_phone(found.Phone),
        "EmergencyName": clean_text(found.EmergencyName),
        "EmergencyPhone": normalise_phone(found.EmergencyPhone),
    }


def find_existing_lite_member_by_name(cursor: Any, member_name: Optional[str], emergency_phone: Optional[str]) -> Optional[dict[str, Any]]:
    if not member_name:
        return None

    if emergency_phone:
        sql = """
        SELECT TOP 1
            LiteMemberId,
            MemberDisplayId,
            MemberName,
            Phone,
            EmergencyName,
            EmergencyPhone
        FROM dbo.SessionAttendance
        WHERE AttendanceMemberKind = 'LITE'
          AND LiteMemberId IS NOT NULL
          AND LOWER(LTRIM(RTRIM(ISNULL(MemberName, '')))) = LOWER(LTRIM(RTRIM(?)))
          AND LTRIM(RTRIM(ISNULL(EmergencyPhone, ''))) = LTRIM(RTRIM(?))
        ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
        """
        found = cursor.execute(sql, member_name, emergency_phone).fetchone()
        if found:
            return {
                "LiteMemberId": str(found.LiteMemberId),
                "MemberDisplayId": clean_text(found.MemberDisplayId),
                "MemberName": clean_text(found.MemberName),
                "Phone": normalise_phone(found.Phone),
                "EmergencyName": clean_text(found.EmergencyName),
                "EmergencyPhone": normalise_phone(found.EmergencyPhone),
            }

    sql = """
    SELECT TOP 1
        LiteMemberId,
        MemberDisplayId,
        MemberName,
        Phone,
        EmergencyName,
        EmergencyPhone
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND LiteMemberId IS NOT NULL
      AND LOWER(LTRIM(RTRIM(ISNULL(MemberName, '')))) = LOWER(LTRIM(RTRIM(?)))
    ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
    """
    found = cursor.execute(sql, member_name).fetchone()

    if not found:
        return None

    return {
        "LiteMemberId": str(found.LiteMemberId),
        "MemberDisplayId": clean_text(found.MemberDisplayId),
        "MemberName": clean_text(found.MemberName),
        "Phone": normalise_phone(found.Phone),
        "EmergencyName": clean_text(found.EmergencyName),
        "EmergencyPhone": normalise_phone(found.EmergencyPhone),
    }


def attendance_exists(cursor: Any, session_id: int, participant_id: Optional[int], lite_member_id: Optional[str]) -> bool:
    if participant_id is not None:
        return cursor.execute(
            "SELECT TOP 1 AttendanceId FROM dbo.SessionAttendance WHERE SessionId = ? AND ParticipantId = ?;",
            session_id,
            participant_id,
        ).fetchone() is not None

    if lite_member_id:
        return cursor.execute(
            "SELECT TOP 1 AttendanceId FROM dbo.SessionAttendance WHERE SessionId = ? AND LiteMemberId = ?;",
            session_id,
            lite_member_id,
        ).fetchone() is not None

    return False


def build_attendance_notes(row: ParsedAttendanceRow, member_kind: str) -> str:
    parts = [
        f"Imported from {IMPORT_SOURCE_NAME}",
        f"Source sheets: {row.source_sheets}",
        f"Source row: {row.source_row}",
        f"Source column: {row.source_col}",
        f"Raw activity header: {row.raw_activity_header}",
    ]

    if row.original_card_value and member_kind == "LITE":
        parts.append(f"Original Excel card value: [{row.original_card_value}]")

    if row.wellbeing_card_number:
        parts.append(f"Wellbeing card number: [{row.wellbeing_card_number}]")

    if row.postcode:
        parts.append(f"Postcode from Excel: [{row.postcode}]")

    if row.comments:
        parts.append(f"Excel comments: {row.comments}")

    return "; ".join(parts)


def create_attendance(
    cursor: Any,
    row: ParsedAttendanceRow,
    session_id: int,
    participant_id: Optional[int],
    saheli_card_number: Optional[str],
    attendance_member_kind: str,
    lite_member_id: Optional[str],
    member_display_id: str,
    member_name: str,
    phone: Optional[str],
    emergency_name: Optional[str],
    emergency_phone: Optional[str],
    created_at: datetime,
    attendance_identity: bool,
    manual_attendance_id: Optional[int],
) -> int:
    columns = [
        "SessionId",
        "ParticipantId",
        "SessionName",
        "SessionDay",
        "SessionDate",
        "SessionMonth",
        "SessionStartTime",
        "SessionEndTime",
        "SaheliCardNumber",
        "RiskStratification",
        "Attended",
        "CheckInTime",
        "CheckOutTime",
        "Notes",
        "CreatedAtUtc",
        "UpdatedAtUtc",
        "AttendanceMemberKind",
        "LiteMemberId",
        "MemberDisplayId",
        "MemberName",
        "Phone",
        "EmergencyName",
        "EmergencyPhone",
        "AF",
        "BP",
        "HeightCm",
        "WeightKg",
    ]

    values: list[Any] = [
        session_id,
        participant_id,
        row.activity_name,
        row.session_day,
        row.session_date.isoformat(),
        row.session_month,
        row.start_time,
        row.end_time,
        saheli_card_number,
        row.risk_stratification,
        1,
        row.start_time,
        row.end_time,
        build_attendance_notes(row, attendance_member_kind),
        created_at,
        created_at,
        attendance_member_kind,
        lite_member_id,
        member_display_id,
        member_name,
        phone,
        emergency_name,
        emergency_phone,
        None,
        None,
        None,
        None,
    ]

    if not attendance_identity:
        if manual_attendance_id is None:
            raise RuntimeError("manual_attendance_id is required because AttendanceId is not IDENTITY.")
        columns.insert(0, "AttendanceId")
        values.insert(0, manual_attendance_id)

    column_sql = ", ".join(f"[{column}]" for column in columns)
    placeholders = ", ".join("?" for _ in values)

    sql = f"""
    INSERT INTO dbo.SessionAttendance ({column_sql})
    OUTPUT INSERTED.AttendanceId
    VALUES ({placeholders});
    """

    inserted = cursor.execute(sql, *values).fetchone()
    if not inserted:
        raise RuntimeError("Attendance INSERT did not return an AttendanceId.")

    return int(inserted.AttendanceId)


# ============================================================
# DATABASE IMPORT
# ============================================================

def import_into_database(parsed_rows: list[ParsedAttendanceRow], *, apply_changes: bool) -> None:
    if pyodbc is None:
        raise RuntimeError("pyodbc is not installed. Run: pip install pyodbc")

    connection_string = build_connection_string()
    created_at = datetime.now(timezone.utc).replace(microsecond=0, tzinfo=None)

    counters: Counter[str] = Counter()
    sheet_counters: dict[str, Counter[str]] = defaultdict(Counter)
    session_id_cache: dict[tuple[str, str, str, str, str], int] = {}
    lite_member_cache: dict[str, dict[str, Any]] = {}

    with pyodbc.connect(connection_string) as conn:
        conn.autocommit = False
        cursor = conn.cursor()

        try:
            session_identity = is_identity_column(cursor, "Sessions", "SessionId")
            attendance_identity = is_identity_column(cursor, "SessionAttendance", "AttendanceId")

            next_manual_session_id = None if session_identity else get_next_manual_id(cursor, "Sessions", "SessionId")
            next_manual_attendance_id = None if attendance_identity else get_next_manual_id(cursor, "SessionAttendance", "AttendanceId")
            next_lite_display_number = get_next_lite_display_number(cursor)

            print(f"SessionId IDENTITY: {session_identity}")
            print(f"AttendanceId IDENTITY: {attendance_identity}")
            print(f"Next LITE MemberDisplayId: LITE-{next_lite_display_number}")

            for row in parsed_rows:
                per_sheet = sheet_counters[row.source_sheet]

                session_id = session_id_cache.get(row.session_key)

                if session_id is None:
                    existing_session_id = find_existing_session(cursor, row) if REUSE_EXISTING_SESSIONS else None

                    if existing_session_id is not None:
                        session_id = existing_session_id
                        counters["sessions_reused"] += 1
                        per_sheet["sessions_reused"] += 1
                    else:
                        manual_session_id = None
                        if not session_identity:
                            manual_session_id = next_manual_session_id
                            assert next_manual_session_id is not None
                            next_manual_session_id += 1

                        session_id = create_session(
                            cursor=cursor,
                            row=row,
                            created_at=created_at,
                            session_identity=session_identity,
                            manual_session_id=manual_session_id,
                        )
                        counters["sessions_created"] += 1
                        per_sheet["sessions_created"] += 1

                    session_id_cache[row.session_key] = session_id

                participant = find_participant_by_card(cursor, row.saheli_card_number)

                if participant:
                    participant_id = participant["ParticipantId"]
                    saheli_card_number = participant["SaheliCardNumber"]
                    attendance_member_kind = "FULL"
                    lite_member_id = None
                    member_display_id = saheli_card_number
                    member_name = row.member_name or participant["FullName"] or f"Card {saheli_card_number}"
                    phone = participant["MobileNumber"]
                    emergency_name = row.emergency_name
                    emergency_phone = row.emergency_phone
                    counters["full_rows"] += 1
                    per_sheet["full_rows"] += 1
                else:
                    participant_id = None
                    saheli_card_number = None
                    attendance_member_kind = "LITE"

                    member_name = row.member_name
                    if not member_name:
                        if row.original_card_value:
                            member_name = f"Card {row.original_card_value}"
                        else:
                            counters["attendance_rows_skipped_no_name"] += 1
                            per_sheet["attendance_rows_skipped_no_name"] += 1
                            continue

                    if row.original_card_value:
                        lite_cache_key = f"CARD|{row.original_card_value.casefold()}"
                    else:
                        lite_cache_key = f"NAME|{member_name.casefold()}|{row.emergency_phone or ''}"

                    lite_member = lite_member_cache.get(lite_cache_key)

                    if lite_member is None:
                        lite_member = find_existing_lite_member_by_original_card(cursor, row.original_card_value)

                    if lite_member is None:
                        lite_member = find_existing_lite_member_by_name(cursor, member_name, row.emergency_phone)

                    if lite_member is None:
                        lite_member = {
                            "LiteMemberId": str(uuid.uuid4()).upper(),
                            "MemberDisplayId": f"LITE-{next_lite_display_number}",
                            "MemberName": member_name,
                            "Phone": None,
                            "EmergencyName": row.emergency_name,
                            "EmergencyPhone": row.emergency_phone,
                        }
                        next_lite_display_number += 1
                        counters["lite_members_created"] += 1
                        per_sheet["lite_members_created"] += 1
                    else:
                        counters["lite_members_reused"] += 1
                        per_sheet["lite_members_reused"] += 1

                    lite_member_cache[lite_cache_key] = lite_member

                    lite_member_id = lite_member["LiteMemberId"]
                    member_display_id = lite_member["MemberDisplayId"]
                    member_name = member_name or lite_member["MemberName"]
                    phone = lite_member.get("Phone")
                    emergency_name = row.emergency_name or lite_member.get("EmergencyName")
                    emergency_phone = row.emergency_phone or lite_member.get("EmergencyPhone")

                    counters["lite_rows"] += 1
                    per_sheet["lite_rows"] += 1

                if SKIP_DUPLICATE_ATTENDANCE and attendance_exists(
                    cursor=cursor,
                    session_id=session_id,
                    participant_id=participant_id,
                    lite_member_id=lite_member_id,
                ):
                    counters["duplicate_attendance_skipped"] += 1
                    per_sheet["duplicate_attendance_skipped"] += 1
                    continue

                manual_attendance_id = None
                if not attendance_identity:
                    manual_attendance_id = next_manual_attendance_id
                    assert next_manual_attendance_id is not None
                    next_manual_attendance_id += 1

                create_attendance(
                    cursor=cursor,
                    row=row,
                    session_id=session_id,
                    participant_id=participant_id,
                    saheli_card_number=saheli_card_number,
                    attendance_member_kind=attendance_member_kind,
                    lite_member_id=lite_member_id,
                    member_display_id=member_display_id,
                    member_name=member_name,
                    phone=phone,
                    emergency_name=emergency_name,
                    emergency_phone=emergency_phone,
                    created_at=created_at,
                    attendance_identity=attendance_identity,
                    manual_attendance_id=manual_attendance_id,
                )

                counters["attendance_created"] += 1
                per_sheet["attendance_created"] += 1

            if apply_changes:
                conn.commit()
                mode_message = "COMMITTED: database changes were saved."
            else:
                conn.rollback()
                mode_message = "DRY RUN: transaction rolled back. No database changes were saved."

        except Exception:
            conn.rollback()
            raise

    print("\nDatabase import summary")
    print("=" * 100)
    print(mode_message)
    print("-" * 100)

    keys = [
        "sessions_created",
        "sessions_reused",
        "attendance_created",
        "duplicate_attendance_skipped",
        "full_rows",
        "lite_rows",
        "lite_members_created",
        "lite_members_reused",
        "attendance_rows_skipped_no_name",
    ]

    for key in keys:
        print(f"{key.replace('_', ' ').title():42}: {counters[key]}")

    print("\nBy worksheet")
    print("-" * 100)
    for sheet_name, per_sheet in sheet_counters.items():
        print(f"\n{sheet_name}")
        for key in keys:
            if per_sheet[key]:
                print(f"  {key.replace('_', ' ').title():40}: {per_sheet[key]}")


# ============================================================
# COMMAND LINE
# ============================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import ARCC 2025 exercise register into CRM Sessions and SessionAttendance.")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_FILE,
        help=f"Excel workbook path. Default: {DEFAULT_INPUT_FILE}",
    )
    parser.add_argument(
        "--inspect-only",
        action="store_true",
        help="Parse Excel and show counts without connecting to SQL Server.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Commit database inserts. Without this option the transaction is rolled back.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if not args.input.exists():
        raise FileNotFoundError(f"Input workbook not found: {args.input}")

    print(f"Reading workbook: {args.input}")
    parsed_rows, session_counter, issues = parse_workbook(args.input)
    print_excel_inspection(parsed_rows, session_counter, issues)

    print("\nImportant")
    print("=" * 100)
    print("The script deduplicates the September monthly sheet against the activity tabs where they overlap.")
    print("It prefers rows with richer participant details, such as name, gender, DOB and postcode.")
    print("It imports only 2025 session dates and ignores 2026 placeholder columns.")

    if args.inspect_only:
        print("\nInspection complete. No SQL connection was made.")
        return

    import_into_database(parsed_rows, apply_changes=args.apply)


if __name__ == "__main__":
    main()
