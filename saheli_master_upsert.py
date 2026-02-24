import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


MASTER_FILE = r"C:\Users\shonk\Downloads\Full Registration for SAHELI.xlsx"
MASTER_SHEET = "Full Register"

GENERATED_FILE = r"C:\Users\shonk\source\PythonCodes\New folder\Saheli_Master_Wide_Output.xlsx"
GENERATED_SHEET = 0

UPDATED_MASTER_FILE = "Full Registration for SAHELI_UPDATED.xlsx"
CHANGELOG_FILE = "Master_Upsert_CHANGELOG.xlsx"

SAHELI_KEY_HEADER = "Saheli Card Number"

PERSONAL_FIELDS_PRIORITY = [
    "Full Name:",
    "Date of Birth:",
    "Address:",
    "Postcode:",
    "Email:",
    "Mobile/Home No:",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    return s


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def saheli_digits(value: Any) -> str:
    s = normalize_text(value)
    digits = re.sub(r"\D+", "", s)
    return digits


def normalize_header(header: Any) -> str:
    s = normalize_text(header).lower()
    if not s:
        return ""
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("：", ":")
    s = re.sub(r"\s*:\s*", ":", s)
    # unify "comment 2"/"comments 2"/"comments:2"
    s = re.sub(r"\bcomments?\s*:?\s*2\b", "comments:2", s)
    return s


def normalize_master_key_for_mapping(header: Any) -> str:
    """Normalize master header to generated key space, applying special WEMWBS rule."""
    raw = normalize_text(header)
    norm = normalize_header(raw)
    if not norm:
        return ""
    # Special rule: map MASTER "<block>  WEMWBS" from GENERATED "<block>  Comments:2"
    m = re.match(r"^(.*?)(?:\s+)?wemwbs$", norm)
    if m:
        prefix = m.group(1).strip()
        if prefix:
            return f"{prefix} comments:2"
        return "comments:2"
    return norm


def flatten_multiindex_columns(columns) -> List[str]:
    flattened = []
    for col in columns:
        if isinstance(col, tuple):
            parts = [normalize_text(x) for x in col if normalize_text(x)]
            flattened.append(" | ".join(parts))
        else:
            flattened.append(normalize_text(col))
    return flattened


def load_generated_dataframe(path: str, sheet: Any) -> Tuple[pd.DataFrame, str]:
    try:
        df = pd.read_excel(path, sheet_name=sheet, header=[0, 1], dtype=object)
        if isinstance(df.columns, pd.MultiIndex):
            flat = flatten_multiindex_columns(df.columns)
            if any("|" in c for c in flat):
                df = df.copy()
                df.columns = flat
                return df, "header=[0,1] flattened"
    except Exception:
        pass
    df = pd.read_excel(path, sheet_name=sheet, header=0, dtype=object)
    df.columns = [normalize_text(c) for c in df.columns]
    return df, "header=0"


@dataclass
class CandidateScore:
    gen_col: str
    gen_idx: int
    global_nonblank: int
    new_keys_nonblank: int
    chosen: bool = False


def build_generated_column_index(df: pd.DataFrame) -> Tuple[Dict[str, List[int]], Dict[int, str], Dict[int, str]]:
    key_to_indices: Dict[str, List[int]] = {}
    idx_to_col: Dict[int, str] = {}
    idx_to_norm: Dict[int, str] = {}
    for idx, col in enumerate(df.columns):
        col_name = normalize_text(col)
        idx_to_col[idx] = col_name
        norm = normalize_header(col_name)
        idx_to_norm[idx] = norm
        if norm:
            key_to_indices.setdefault(norm, []).append(idx)

        # If flattened multi-header (e.g., "Section | Full Name:"), also index each segment.
        # This is critical for mapping master headers like "Full Name:" to the correct generated column.
        if "|" in col_name:
            parts = [normalize_text(p) for p in col_name.split("|")]
            for p in parts:
                p_norm = normalize_header(p)
                if p_norm:
                    key_to_indices.setdefault(p_norm, []).append(idx)
    return key_to_indices, idx_to_col, idx_to_norm


def resolve_generated_saheli_key_candidates(
    df: pd.DataFrame,
    gen_key_to_indices: Dict[str, List[int]],
    gen_idx_to_col: Dict[int, str],
) -> Tuple[List[int], List[Dict[str, Any]]]:
    """
    Find likely GENERATED Saheli key columns.
    Supports exact normalized match and flattened multi-header columns
    where the final label includes the key text.
    """
    target = normalize_header(SAHELI_KEY_HEADER)
    target_no_colon = target.replace(":", "")
    candidates: List[int] = []
    debug_rows: List[Dict[str, Any]] = []

    # 1) Exact normalized key match
    if target in gen_key_to_indices:
        for idx in gen_key_to_indices[target]:
            candidates.append(idx)
            debug_rows.append(
                {
                    "reason": "exact_normalized_match",
                    "idx": idx,
                    "column": gen_idx_to_col[idx],
                }
            )

    # 2) Common variant with colon
    colon_variant = f"{target_no_colon}:"
    if colon_variant in gen_key_to_indices:
        for idx in gen_key_to_indices[colon_variant]:
            if idx not in candidates:
                candidates.append(idx)
                debug_rows.append(
                    {
                        "reason": "colon_variant_match",
                        "idx": idx,
                        "column": gen_idx_to_col[idx],
                    }
                )

    # 3) Flattened multi-header contains the key in one segment
    for idx, raw_col in gen_idx_to_col.items():
        raw = normalize_text(raw_col).lower()
        if not raw:
            continue
        parts = [p.strip() for p in raw.split("|")]
        parts_norm = [normalize_header(p) for p in parts if p.strip()]
        joined_norm = normalize_header(raw)
        matched = False
        if target in parts_norm or target_no_colon in [p.replace(":", "") for p in parts_norm]:
            matched = True
        elif target in joined_norm or target_no_colon in joined_norm.replace(":", ""):
            # fallback substring match for malformed flattened labels
            matched = True
        if matched and idx not in candidates:
            candidates.append(idx)
            debug_rows.append(
                {
                    "reason": "flattened_contains_match",
                    "idx": idx,
                    "column": gen_idx_to_col[idx],
                }
            )

    # Stable left-to-right order
    candidates = sorted(set(candidates))
    debug_rows = sorted(debug_rows, key=lambda r: r["idx"])
    return candidates, debug_rows


def nonblank_count(series: pd.Series) -> int:
    return int(series.apply(lambda v: not is_blank(v)).sum())


def choose_best_duplicate(
    df: pd.DataFrame,
    candidate_indices: List[int],
    new_key_mask: Optional[pd.Series] = None,
) -> Tuple[int, List[CandidateScore]]:
    scores: List[CandidateScore] = []
    for idx in candidate_indices:
        s = df.iloc[:, idx]
        global_count = nonblank_count(s)
        new_count = 0
        if new_key_mask is not None and len(new_key_mask) == len(df):
            new_count = int(s[new_key_mask].apply(lambda v: not is_blank(v)).sum())
        scores.append(
            CandidateScore(
                gen_col=normalize_text(df.columns[idx]),
                gen_idx=idx,
                global_nonblank=global_count,
                new_keys_nonblank=new_count,
            )
        )
    # Base choice: global nonblank, then stable leftmost
    best = max(scores, key=lambda x: (x.global_nonblank, -x.gen_idx))
    # New-keys override if strictly better on new-key population
    if new_key_mask is not None:
        best_new = max(scores, key=lambda x: (x.new_keys_nonblank, x.global_nonblank, -x.gen_idx))
        if best_new.new_keys_nonblank > best.new_keys_nonblank:
            best = best_new
    for sc in scores:
        sc.chosen = sc.gen_idx == best.gen_idx
    return best.gen_idx, scores


def ensure_unique_columns(columns: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out = []
    for c in columns:
        base = c if c else "Unnamed"
        n = seen.get(base, 0)
        out_name = base if n == 0 else f"{base}__dup{n}"
        seen[base] = n + 1
        out.append(out_name)
    return out


def main() -> None:
    print("Loading MASTER workbook with openpyxl...")
    wb = load_workbook(MASTER_FILE)
    ws = wb[MASTER_SHEET]

    master_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    master_header_to_col: Dict[str, int] = {}
    master_norm_to_cols: Dict[str, List[int]] = {}
    for c_idx, header in enumerate(master_headers, start=1):
        h = normalize_text(header)
        if h:
            master_header_to_col[h] = c_idx
            master_norm_to_cols.setdefault(normalize_header(h), []).append(c_idx)

    # Resolve master key column
    master_key_col = None
    for c_idx, header in enumerate(master_headers, start=1):
        if normalize_header(header) == normalize_header(SAHELI_KEY_HEADER):
            master_key_col = c_idx
            break
    if master_key_col is None:
        raise KeyError(f"MASTER key column not found: {SAHELI_KEY_HEADER}")

    print("Loading GENERATED file...")
    gen_df, gen_load_mode = load_generated_dataframe(GENERATED_FILE, GENERATED_SHEET)
    gen_df = gen_df.copy()
    gen_df.columns = ensure_unique_columns([normalize_text(c) for c in gen_df.columns])
    print(f"Generated load mode: {gen_load_mode}")
    print(f"Generated shape: {gen_df.shape}")

    gen_key_to_indices, gen_idx_to_col, _ = build_generated_column_index(gen_df)
    gen_key_norm = normalize_header(SAHELI_KEY_HEADER)
    key_candidate_idxs, key_candidate_debug = resolve_generated_saheli_key_candidates(
        gen_df, gen_key_to_indices, gen_idx_to_col
    )
    if not key_candidate_idxs:
        print("\nCould not find GENERATED Saheli key column by exact match. Header samples:")
        for i, c in enumerate(list(gen_df.columns)[:80]):
            print(f"  [{i}] {c}")
        raise KeyError(f"GENERATED key column not found for normalized header: {gen_key_norm}")

    print("\nGenerated key column candidates:")
    for r in key_candidate_debug[:20]:
        print(f"  idx={r['idx']}: {r['column']} ({r['reason']})")

    # Choose generated key column (duplicate-safe by global fill)
    chosen_gen_key_idx, key_scores = choose_best_duplicate(gen_df, key_candidate_idxs, None)
    gen_key_col_name = gen_idx_to_col[chosen_gen_key_idx]
    gen_df["_saheli_key_digits"] = gen_df.iloc[:, chosen_gen_key_idx].apply(saheli_digits)

    # Build master key index from worksheet values
    master_key_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        key = saheli_digits(ws.cell(row=r, column=master_key_col).value)
        if key and key not in master_key_to_row:
            master_key_to_row[key] = r

    generated_keys = gen_df["_saheli_key_digits"]
    new_key_mask = generated_keys.apply(lambda k: bool(k) and k not in master_key_to_row)
    new_keys = [k for k in generated_keys[new_key_mask].tolist() if k]

    # Build best generated column choice per normalized header
    resolved_gen_col_idx: Dict[str, int] = {}
    duplicate_resolution_rows: List[Dict[str, Any]] = []

    personal_norm_targets = {normalize_header(x) for x in PERSONAL_FIELDS_PRIORITY}
    # Also include common variant without colon in case generated flattened headers differ
    personal_norm_targets |= {normalize_header(x.replace(":", "")) for x in PERSONAL_FIELDS_PRIORITY}

    for norm_key, candidate_idxs in gen_key_to_indices.items():
        use_new_key_mask = new_key_mask if norm_key in personal_norm_targets else None
        chosen_idx, score_list = choose_best_duplicate(gen_df, candidate_idxs, use_new_key_mask)
        resolved_gen_col_idx[norm_key] = chosen_idx
        for score in score_list:
            duplicate_resolution_rows.append(
                {
                    "NormalizedHeaderKey": norm_key,
                    "GeneratedColumn": score.gen_col,
                    "GeneratedColumnIndex0": score.gen_idx,
                    "GlobalNonBlankCount": score.global_nonblank,
                    "NewKeysNonBlankCount": score.new_keys_nonblank,
                    "Chosen": "Y" if score.chosen else "",
                    "UsedNewKeyScoring": "Y" if use_new_key_mask is not None else "",
                }
            )

    # Build mapping from MASTER column -> chosen GENERATED column index (no new columns)
    master_to_generated_map: Dict[int, int] = {}
    column_mapping_rows: List[Dict[str, Any]] = []
    for m_col_idx, m_header in enumerate(master_headers, start=1):
        m_header_txt = normalize_text(m_header)
        if not m_header_txt:
            continue
        map_key = normalize_master_key_for_mapping(m_header_txt)
        chosen_idx = resolved_gen_col_idx.get(map_key)
        if chosen_idx is None:
            # Fallback: try direct normalized header if special conversion misses
            chosen_idx = resolved_gen_col_idx.get(normalize_header(m_header_txt))
        if chosen_idx is not None:
            master_to_generated_map[m_col_idx] = chosen_idx
        column_mapping_rows.append(
            {
                "MasterColumnIndex1": m_col_idx,
                "MasterHeader": m_header_txt,
                "MasterNormalizedHeader": normalize_header(m_header_txt),
                "MappingLookupKey": map_key,
                "MappedGeneratedColumnIndex0": chosen_idx if chosen_idx is not None else "",
                "MappedGeneratedColumn": gen_idx_to_col.get(chosen_idx, "") if chosen_idx is not None else "",
                "IsSaheliKey": "Y" if m_col_idx == master_key_col else "",
                "SpecialWEMWBSRuleApplied": "Y" if normalize_header(m_header_txt).endswith("wemwbs") else "",
            }
        )

    # Diagnostics: personal field mapping + duplicate candidates
    print("\n=== DIAGNOSTIC: Personal field duplicate resolution (top 50 rows) ===")
    personal_diag_rows = [
        r
        for r in duplicate_resolution_rows
        if r["NormalizedHeaderKey"] in personal_norm_targets
    ]
    personal_diag_rows = sorted(
        personal_diag_rows,
        key=lambda r: (
            r["NormalizedHeaderKey"],
            -(1 if r["Chosen"] == "Y" else 0),
            -int(r["NewKeysNonBlankCount"]),
            -int(r["GlobalNonBlankCount"]),
            int(r["GeneratedColumnIndex0"]),
        ),
    )
    for row in personal_diag_rows[:50]:
        marker = "*" if row["Chosen"] == "Y" else " "
        print(
            f"{marker} {row['NormalizedHeaderKey']:<25} "
            f"idx={row['GeneratedColumnIndex0']:<4} "
            f"global={row['GlobalNonBlankCount']:<5} "
            f"new={row['NewKeysNonBlankCount']:<5} "
            f"{row['GeneratedColumn']}"
        )

    # Preview chosen personal fields for sample new keys
    sample_new_preview_rows: List[Dict[str, Any]] = []
    sample_new_keys = []
    for k in new_keys:
        if k not in sample_new_keys:
            sample_new_keys.append(k)
        if len(sample_new_keys) >= 10:
            break

    # Resolve chosen generated columns for requested personal fields (using exact key if possible)
    chosen_personal_cols: Dict[str, Optional[int]] = {}
    for field in PERSONAL_FIELDS_PRIORITY:
        field_norm = normalize_header(field)
        idx = resolved_gen_col_idx.get(field_norm)
        if idx is None and field_norm.endswith(":"):
            idx = resolved_gen_col_idx.get(field_norm[:-1])
        chosen_personal_cols[field] = idx

    print("\n=== DIAGNOSTIC: Sample new Saheli keys preview (first 10) ===")
    for key in sample_new_keys:
        match_rows = gen_df.index[gen_df["_saheli_key_digits"] == key].tolist()
        if not match_rows:
            continue
        i = match_rows[0]
        preview = {"SaheliKeyDigits": key}
        for field, idx in chosen_personal_cols.items():
            val = gen_df.iloc[i, idx] if idx is not None else None
            preview[field] = val
        sample_new_preview_rows.append(preview)
        pretty = " | ".join([f"{k}={normalize_text(v)}" for k, v in preview.items()])
        print(pretty)

    # Upsert
    print("\nApplying upsert...")
    cell_changes: List[Dict[str, Any]] = []
    new_rows_changelog: List[Dict[str, Any]] = []
    summary = {
        "MasterRowsBefore": max(0, ws.max_row - 1),
        "GeneratedRows": int(len(gen_df)),
        "GeneratedRowsWithSaheliKey": int((generated_keys != "").sum()),
        "NewSaheliKeysFound": int(len(set(new_keys))),
        "ExistingRowsUpdated": 0,
        "NewRowsAppended": 0,
        "CellsFilled": 0,
        "MappedMasterColumns": int(len(master_to_generated_map)),
    }

    updated_existing_rows = set()
    appended_rows = 0
    cells_filled = 0

    # Use first non-empty occurrence for a key; if duplicate key in GENERATED exists, keep first row
    gen_first_row_by_key: Dict[str, int] = {}
    for idx, key in enumerate(generated_keys.tolist()):
        if key and key not in gen_first_row_by_key:
            gen_first_row_by_key[key] = idx

    # Process generated rows in source order
    for g_idx, g_row in gen_df.iterrows():
        key = g_row["_saheli_key_digits"]
        if not key:
            continue
        if key in master_key_to_row:
            target_row = master_key_to_row[key]
            row_is_new = False
        else:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=master_key_col).value = g_row.iloc[chosen_gen_key_idx]
            master_key_to_row[key] = target_row
            row_is_new = True
            appended_rows += 1

        row_changed = False
        new_row_record: Dict[str, Any] = {
            "MasterRowNumber": target_row,
            "SaheliKeyDigits": key,
            "GeneratedRowIndex0": int(g_idx),
        }

        for m_col_idx, gen_col_idx in master_to_generated_map.items():
            src_val = g_row.iloc[gen_col_idx]
            if is_blank(src_val):
                continue
            cell = ws.cell(row=target_row, column=m_col_idx)
            old_val = cell.value
            if not is_blank(old_val):
                continue  # never overwrite non-blank master cells
            cell.value = src_val
            row_changed = True
            cells_filled += 1
            cell_changes.append(
                {
                    "MasterRowNumber": target_row,
                    "MasterColumnIndex1": m_col_idx,
                    "MasterHeader": normalize_text(master_headers[m_col_idx - 1]),
                    "SaheliKeyDigits": key,
                    "GeneratedRowIndex0": int(g_idx),
                    "GeneratedColumnIndex0": int(gen_col_idx),
                    "GeneratedColumn": gen_idx_to_col.get(gen_col_idx, ""),
                    "OldValue": old_val,
                    "NewValue": src_val,
                    "ChangeType": "NewRowFill" if row_is_new else "BlankFillExistingRow",
                }
            )
            if row_is_new:
                new_row_record[normalize_text(master_headers[m_col_idx - 1])] = src_val

        if row_is_new:
            # Ensure changelog includes mapped headers even if blank/unmapped
            new_rows_changelog.append(new_row_record)
        if row_changed and not row_is_new:
            updated_existing_rows.add(target_row)

    summary["ExistingRowsUpdated"] = int(len(updated_existing_rows))
    summary["NewRowsAppended"] = int(appended_rows)
    summary["CellsFilled"] = int(cells_filled)
    summary["MasterRowsAfter"] = max(0, ws.max_row - 1)

    # Save updated master workbook
    updated_path = str(Path.cwd() / UPDATED_MASTER_FILE)
    wb.save(updated_path)
    print(f"Updated master saved: {updated_path}")

    # Build changelog workbook
    print("Writing changelog workbook...")
    summary_df = pd.DataFrame(
        [{"Metric": k, "Value": v} for k, v in summary.items()]
    )
    cell_changes_df = pd.DataFrame(cell_changes)
    new_rows_df = pd.DataFrame(new_rows_changelog)
    column_mapping_df = pd.DataFrame(column_mapping_rows)
    duplicate_resolution_df = pd.DataFrame(duplicate_resolution_rows)
    sample_preview_df = pd.DataFrame(sample_new_preview_rows)

    with pd.ExcelWriter(Path.cwd() / CHANGELOG_FILE, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        new_rows_df.to_excel(writer, sheet_name="NewRows", index=False)
        cell_changes_df.to_excel(writer, sheet_name="CellChanges", index=False)
        column_mapping_df.to_excel(writer, sheet_name="ColumnMapping", index=False)
        duplicate_resolution_df.to_excel(writer, sheet_name="DuplicateResolution", index=False)
        sample_preview_df.to_excel(writer, sheet_name="SampleNewKeysPreview", index=False)

    print(f"Changelog saved: {Path.cwd() / CHANGELOG_FILE}")
    print("Done.")


if __name__ == "__main__":
    main()
