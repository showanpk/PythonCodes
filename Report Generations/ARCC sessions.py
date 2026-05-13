from __future__ import annotations

import os
import re
import uuid
from datetime import datetime, date, timedelta, timezone
from pathlib import Path
from typing import Any, Optional, Tuple
from collections import defaultdict

import pandas as pd
import pyodbc
import openpyxl


# ============================================================
# CONFIG
# ============================================================

INPUT_FILE = Path(r"C:\Users\shonk\Downloads\ARCC Activity Register  2026 (2).xlsx")

SQL_USER = os.getenv("CRM_SQL_USER", "sahelihubadmin")
SQL_PASSWORD = os.getenv("CRM_SQL_PASSWORD", "W7WZ7ZaG1YbMZ71gh%2xSFuR")

SQL_CONNECTION_STRING = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=tcp:sahelihub.database.windows.net,1433;"
    "DATABASE=SahelihubCRM;"
    f"UID={SQL_USER};"
    f"PWD={SQL_PASSWORD};"
    "Encrypt=yes;"
    "TrustServerCertificate=yes;"
)

IMPORT_SOURCE_NAME = "Alum Rock Community Centre Activity Register 2026"

SKIP_SHEETS = {
    "Sheet1",
    "Template",
}

REUSE_EXISTING_SESSIONS = True
SKIP_DUPLICATE_ATTENDANCE = True

DEFAULT_VENUE_NAME = "Alum Rock Community Centre"
DEFAULT_FREQUENCY = "WEEKLY"
DEFAULT_CATEGORY = "Fitness"
DEFAULT_SUBCATEGORY = "Alum Rock Community Centre"
DEFAULT_ACTIVITY_CATEGORY = "Fitness"

# Alum Rock Community Centre register usually has start time only, not end time.
# This creates EndTime automatically.
DEFAULT_SESSION_DURATION_MINUTES = 60

# Per-sheet default session start times (use only when no time can be detected)
DEFAULT_SESSION_TIMES_BY_SHEET: dict[str, str] = {
    # "Chair Based": "10:00",
    # "Walk & Talk": "10:00",
}

TEXT_NULLS = {
    "",
    "#N/A",
    "N/A",
    "NA",
    "NONE",
    "NULL",
    "NAN",
    "(BLANK)",
    "0",
    "00",
    "-",
}

NO_ATTENDANCE_WORDS = {
    "no attended",
    "no attendance",
    "no one",
    "none attended",
    "nobody attended",
    "no participant",
    "no participants",
}

CANCELLED_WORDS = {
    "cancelled",
    "canceled",
    "bad weather got cancelled",
}


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_value(value: Any) -> Optional[Any]:
    if pd.isna(value):
        return None

    if isinstance(value, str):
        text = re.sub(r"\s+", " ", value.strip())

        if text.upper() in TEXT_NULLS:
            return None

        return text

    return value


def clean_text(value: Any) -> Optional[str]:
    value = clean_value(value)

    if value is None:
        return None

    # Convert integer-like floats to strings first
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()

    if text.upper() in TEXT_NULLS:
        return None

    return text


def normalise_header(value: Any) -> str:
    text = clean_text(value) or ""
    text = text.lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def excel_date_to_date(value: Any) -> Optional[date]:
    value = clean_value(value)

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()

    parsed = pd.to_datetime(str(value).strip(), dayfirst=True, errors="coerce")

    if pd.isna(parsed):
        return None

    return parsed.date()


def normalise_day(value: Any, fallback_date: Optional[date]) -> Optional[str]:
    text = clean_text(value)

    if text:
        fixes = {
            "mon": "Monday",
            "monday": "Monday",
            "tue": "Tuesday",
            "tues": "Tuesday",
            "tuesdsay": "Tuesday",
            "tuesday": "Tuesday",
            "wed": "Wednesday",
            "wednesday": "Wednesday",
            "thu": "Thursday",
            "thur": "Thursday",
            "thurs": "Thursday",
            "thursday": "Thursday",
            "fri": "Friday",
            "friday": "Friday",
            "sat": "Saturday",
            "saturday": "Saturday",
            "sun": "Sunday",
            "sunday": "Sunday",
        }

        key = text.lower().strip()
        return fixes.get(key, text.capitalize())

    if fallback_date:
        return fallback_date.strftime("%A")

    return None


def month_name(value: Any, fallback_date: Optional[date]) -> Optional[str]:
    text = clean_text(value)

    if text:
        return (
            text
            .replace("Decemeber", "December")
            .replace("Febuary", "February")
        )

    if fallback_date:
        return fallback_date.strftime("%B")

    return None


def normalise_activity_from_sheet(sheet_name: Optional[str]) -> str:
    text = (sheet_name or "").strip()

    mapping = {
        "Chair Based": "Chair Based",
        "Omnia Chair Exercise": "Omnia Chair Exercise",
        "Circuit": "Circuit",
        "Yoga": "Yoga",
        "Strength & Stretch": "Strength & Stretch",
        "Walk & Talk": "Walk & Talk",
        "Saheli Social": "Saheli Social",
    }

    return mapping.get(text, text)


# ============================================================
# TIME HELPERS
# ============================================================

def minutes_to_time_string(total_minutes: int) -> str:
    total_minutes = total_minutes % (24 * 60)
    hour = total_minutes // 60
    minute = total_minutes % 60
    return f"{hour:02d}:{minute:02d}:00"


def time_string_to_minutes(value: str) -> int:
    parts = value.split(":")
    return int(parts[0]) * 60 + int(parts[1])


def parse_single_time(value: Any) -> Optional[str]:
    value = clean_value(value)

    if value is None:
        return None

    # Excel time stored as fraction of a day, e.g. 0.4479166667
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 0 <= float(value) < 1:
            total_minutes = round(float(value) * 24 * 60)
            return minutes_to_time_string(total_minutes)

        # Do not treat normal numbers like 1373 or 2026 as times.
        return None

    # Only treat datetime as time if it actually has a time component.
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return None
        return f"{value.hour:02d}:{value.minute:02d}:00"

    text = str(value).strip().lower()

    if not text:
        return None

    # Avoid parsing dates as times.
    if re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text):
        return None

    text = re.sub(r"\s+", " ", text)

    # Normalise a.m. / p.m. before changing dots.
    text = (
        text
        .replace("a.m.", "am")
        .replace("p.m.", "pm")
        .replace("a.m", "am")
        .replace("p.m", "pm")
        .replace("a m", "am")
        .replace("p m", "pm")
    )

    # If a range exists, take only the first time.
    # Examples: 09:30 - 10:30, 9.30 to 10.30, 12.15 p.m. - 1.15 p.m.
    range_match = re.search(
        r"(\d{1,2}(?::|\.)?\d{0,2}\s*(?:am|pm)?)"
        r"\s*(?:-|–|—|to)\s*"
        r"(\d{1,2}(?::|\.)?\d{0,2}\s*(?:am|pm)?)",
        text,
    )

    if range_match:
        text = range_match.group(1)
    else:
        embedded = re.search(
            r"(\d{1,2}(?::|\.)?\d{0,2}\s*(?:am|pm)?)",
            text,
        )
        if embedded:
            text = embedded.group(1)

    text = text.strip()
    text = re.sub(r"\s+", "", text)

    suffix = None

    if text.endswith("am"):
        suffix = "am"
        text = text[:-2]
    elif text.endswith("pm"):
        suffix = "pm"
        text = text[:-2]

    # 12.15 -> 12:15
    text = text.replace(".", ":")

    # 1230 -> 12:30, 930 -> 9:30
    if ":" not in text and text.isdigit():
        if len(text) == 4:
            text = f"{text[:2]}:{text[2:]}"
        elif len(text) == 3:
            text = f"{text[0]}:{text[1:]}"

    match = re.match(r"^(\d{1,2})(?::(\d{1,2}))?$", text)

    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or 0)

    if suffix == "pm" and hour != 12:
        hour += 12
    elif suffix == "am" and hour == 12:
        hour = 0

    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None

    return f"{hour:02d}:{minute:02d}:00"


def make_end_time(start_time: Optional[str]) -> Optional[str]:
    if not start_time:
        return None

    start_minutes = time_string_to_minutes(start_time)
    end_minutes = start_minutes + DEFAULT_SESSION_DURATION_MINUTES

    return minutes_to_time_string(end_minutes)


# ============================================================
# EXCEL COLUMN MAPPING
# ============================================================

def find_column(columns: list[Any], possible_names: list[str]) -> Optional[str]:
    normalised = {
        normalise_header(column): column
        for column in columns
    }

    for wanted in possible_names:
        wanted_norm = normalise_header(wanted)

        for norm, original in normalised.items():
            if norm == wanted_norm:
                return original

        for norm, original in normalised.items():
            if wanted_norm in norm:
                return original

    return None


def get_sheet_column_map(df: pd.DataFrame) -> dict[str, Optional[str]]:
    columns = list(df.columns)

    return {
        "session": find_column(columns, ["Session", "session"]),
        "day": find_column(columns, ["Day"]),
        "date": find_column(columns, ["Date"]),
        "month": find_column(columns, ["Month"]),
        "time": find_column(columns, ["Time"]),
        "card": find_column(columns, ["Saheli Card Number"]),
        "name": find_column(columns, ["Name"]),
        "emergency_name": find_column(columns, ["Emergency contact Name", "Emergency Contact Name"]),
        "emergency_phone": find_column(columns, ["Emergency Number", "Emergency contact number"]),
        "risk": find_column(columns, ["Risk Stratification"]),
        "disability": find_column(columns, ["Disability", "Disabilities", "Learning Disabliity", "Learning Disability"]),
    }


def get_row_value(row: pd.Series, column_name: Optional[str]) -> Optional[Any]:
    if not column_name:
        return None

    if column_name not in row.index:
        return None

    return row[column_name]


def get_cell_value_from_ws(ws: "openpyxl.worksheet.worksheet.Worksheet", row: int, col: int) -> Optional[Any]:
    # Return the cell value, falling back to merged cell top-left if needed.
    try:
        cell = ws.cell(row=row, column=col)
    except Exception:
        return None

    if cell.value is not None:
        return cell.value

    # If empty, check merged ranges to see if this cell is inside one
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged:
            # top-left of merged range
            min_row, min_col = merged.min_row, merged.min_col
            return ws.cell(row=min_row, column=min_col).value

    return None


# ============================================================
# STATUS / MEMBER HELPERS
# ============================================================

def row_status(*values: Any) -> str:
    text = " ".join(
        clean_text(value) or ""
        for value in values
    ).lower()

    if any(word in text for word in CANCELLED_WORDS):
        return "cancelled"

    if any(word in text for word in NO_ATTENDANCE_WORDS):
        return "no_attendance"

    return "normal"


def extract_member(raw_card_value: Any, raw_name_value: Any) -> Tuple[Optional[str], Optional[str]]:
    raw_card = clean_text(raw_card_value)
    raw_name = clean_text(raw_name_value)

    card_number = None
    member_name = raw_name

    if raw_card:
        if re.fullmatch(r"\d+(?:\.0)?", raw_card):
            card_number = str(int(float(raw_card)))
        else:
            bracket_match = re.search(r"\((\d+)\)", raw_card)

            if bracket_match:
                card_number = bracket_match.group(1)
                possible_name = re.sub(r"\(\d+\)", "", raw_card).strip()

                if possible_name and not member_name:
                    member_name = possible_name
            elif not member_name:
                member_name = raw_card

    return card_number, member_name


def build_attendance_notes(
    source_sheet: str,
    source_row: int,
    raw_time: Optional[str],
    disability: Optional[str],
    original_card_value: Optional[str],
    attendance_member_kind: str,
) -> Optional[str]:
    note_parts = [
        f"Imported from {IMPORT_SOURCE_NAME}",
        f"Source sheet: {source_sheet}",
        f"Source row: {source_row}",
    ]

    if raw_time:
        note_parts.append(f"Raw time: {raw_time}")

    if disability:
        note_parts.append(f"Disability: {disability}")

    if original_card_value and attendance_member_kind == "LITE":
        note_parts.append(f"Original Excel card value: {original_card_value}")

    return "; ".join(note_parts)


# ============================================================
# SQL HELPERS
# ============================================================

def is_identity_column(cursor: pyodbc.Cursor, table_name: str, column_name: str) -> bool:
    sql = """
    SELECT c.is_identity
    FROM sys.columns c
    WHERE c.object_id = OBJECT_ID(?)
      AND c.name = ?;
    """

    row = cursor.execute(sql, f"dbo.{table_name}", column_name).fetchone()

    if not row:
        raise RuntimeError(f"Column not found: dbo.{table_name}.{column_name}")

    return bool(row.is_identity)


def get_next_manual_id(cursor: pyodbc.Cursor, table_name: str, column_name: str) -> int:
    sql = f"SELECT ISNULL(MAX({column_name}), 0) + 1 AS NextId FROM dbo.{table_name};"
    row = cursor.execute(sql).fetchone()
    return int(row.NextId)


def get_next_lite_display_number(cursor: pyodbc.Cursor) -> int:
    sql = """
    SELECT ISNULL(MAX(TRY_CONVERT(int, REPLACE(MemberDisplayId, 'LITE-', ''))), 0) + 1 AS NextLiteNumber
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND MemberDisplayId LIKE 'LITE-%';
    """

    row = cursor.execute(sql).fetchone()
    return int(row.NextLiteNumber)


def find_existing_session(
    cursor: pyodbc.Cursor,
    venue_name: str,
    activity_name: str,
    session_date: date,
    start_time: Optional[str],
    end_time: Optional[str],
) -> Optional[int]:
    sql = """
    SELECT TOP 1 SessionId
    FROM dbo.Sessions
    WHERE LOWER(LTRIM(RTRIM(ISNULL(VenueName, '')))) = LOWER(LTRIM(RTRIM(?)))
      AND LOWER(LTRIM(RTRIM(ISNULL(ActivityName, '')))) = LOWER(LTRIM(RTRIM(?)))
      AND CAST(SessionDate AS date) = CAST(? AS date)
      AND (
            (StartTime IS NULL AND ? IS NULL)
            OR CONVERT(varchar(8), StartTime, 108) = ?
          )
      AND (
            (EndTime IS NULL AND ? IS NULL)
            OR CONVERT(varchar(8), EndTime, 108) = ?
          )
    ORDER BY SessionId;
    """

    row = cursor.execute(
        sql,
        venue_name,
        activity_name,
        session_date.isoformat(),
        start_time,
        start_time,
        end_time,
        end_time,
    ).fetchone()

    if row:
        return int(row.SessionId)

    return None


def create_session(
    cursor: pyodbc.Cursor,
    session_identity: bool,
    manual_session_id: Optional[int],
    venue_name: str,
    activity_name: str,
    session_date: date,
    start_time: str,
    end_time: str,
    raw_time: Optional[str],
    is_cancelled: bool,
    created_at: datetime,
    source_sheet: str,
) -> int:
    columns = [
        "Frequency",
        "Category",
        "SubCategory",
        "ActivityCategory",
        "VenueName",
        "ActivityName",
        "Notes",
        "IsRecurringWeekly",
        "DayOfWeek",
        "SessionDate",
        "ArrivalTime",
        "StartTime",
        "EndTime",
        "Capacity",
        "IsBookingRequired",
        "IsCancelled",
        "CreatedAtUtc",
        "AssignedStaffId",
        "RecurringSeriesId",
    ]

    values = [
        DEFAULT_FREQUENCY,
        DEFAULT_CATEGORY,
        DEFAULT_SUBCATEGORY,
        DEFAULT_ACTIVITY_CATEGORY,
        venue_name,
        activity_name,
        f"Imported from {IMPORT_SOURCE_NAME}. Source sheet: {source_sheet}. Raw time: {raw_time}",
        0,
        None,
        session_date.isoformat(),
        None,
        start_time,
        end_time,
        None,
        0,
        1 if is_cancelled else 0,
        created_at,
        None,
        None,
    ]

    if not session_identity:
        if manual_session_id is None:
            raise RuntimeError("manual_session_id is required because SessionId is not an IDENTITY column.")

        columns.insert(0, "SessionId")
        values.insert(0, manual_session_id)

    column_sql = ", ".join(f"[{column}]" for column in columns)
    placeholder_sql = ", ".join("?" for _ in columns)

    sql = f"""
    INSERT INTO dbo.Sessions ({column_sql})
    OUTPUT INSERTED.SessionId
    VALUES ({placeholder_sql});
    """

    try:
        row = cursor.execute(sql, values).fetchone()
    except Exception:
        print("Error inserting session:")
        print(f"VenueName: {venue_name}")
        print(f"ActivityName: {activity_name}")
        print(f"SessionDate: {session_date}")
        print(f"StartTime: {start_time}")
        print(f"EndTime: {end_time}")
        print(f"Values: {values}")
        raise

    if not row:
        raise RuntimeError("Session insert failed. No SessionId was returned.")

    return int(row.SessionId)


def update_session_cancelled(cursor: pyodbc.Cursor, session_id: int) -> None:
    sql = """
    UPDATE dbo.Sessions
    SET IsCancelled = 1
    WHERE SessionId = ?;
    """

    cursor.execute(sql, session_id)


def find_existing_full_member_from_attendance(
    cursor: pyodbc.Cursor,
    saheli_card_number: Optional[str],
) -> Optional[dict]:
    if not saheli_card_number:
        return None

    sql = """
    SELECT TOP 1
        ParticipantId,
        SaheliCardNumber,
        MemberDisplayId,
        MemberName,
        Phone,
        EmergencyName,
        EmergencyPhone,
        AttendanceMemberKind
    FROM dbo.SessionAttendance
    WHERE SaheliCardNumber = ?
            AND AttendanceMemberKind = 'FULL'
            AND ParticipantId IS NOT NULL
    ORDER BY
        CASE WHEN ParticipantId IS NOT NULL THEN 0 ELSE 1 END,
        UpdatedAtUtc DESC,
        CreatedAtUtc DESC,
        AttendanceId DESC;
    """

    row = cursor.execute(sql, saheli_card_number).fetchone()

    if not row:
        return None

    return {
        "ParticipantId": row.ParticipantId,
        "SaheliCardNumber": row.SaheliCardNumber,
        "MemberDisplayId": row.MemberDisplayId,
        "MemberName": row.MemberName,
        "Phone": row.Phone,
        "EmergencyName": row.EmergencyName,
        "EmergencyPhone": row.EmergencyPhone,
        "AttendanceMemberKind": row.AttendanceMemberKind or "FULL",
    }


def find_existing_lite_member(
    cursor: pyodbc.Cursor,
    member_name: Optional[str],
    emergency_phone: Optional[str],
) -> Optional[dict]:
    if not member_name:
        return None

    if emergency_phone:
        sql = """
        SELECT TOP 1
            LiteMemberId,
            MemberDisplayId,
            MemberName,
            Phone,
            EmergencyName,
            EmergencyPhone
        FROM dbo.SessionAttendance
        WHERE AttendanceMemberKind = 'LITE'
          AND LiteMemberId IS NOT NULL
          AND LOWER(LTRIM(RTRIM(ISNULL(MemberName, '')))) = LOWER(LTRIM(RTRIM(?)))
          AND LTRIM(RTRIM(ISNULL(EmergencyPhone, ''))) = LTRIM(RTRIM(?))
        ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
        """

        row = cursor.execute(sql, member_name, emergency_phone).fetchone()

        if row:
            return {
                "LiteMemberId": row.LiteMemberId,
                "MemberDisplayId": row.MemberDisplayId,
                "MemberName": row.MemberName,
                "Phone": row.Phone,
                "EmergencyName": row.EmergencyName,
                "EmergencyPhone": row.EmergencyPhone,
            }

    sql = """
    SELECT TOP 1
        LiteMemberId,
        MemberDisplayId,
        MemberName,
        Phone,
        EmergencyName,
        EmergencyPhone
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND LiteMemberId IS NOT NULL
      AND LOWER(LTRIM(RTRIM(ISNULL(MemberName, '')))) = LOWER(LTRIM(RTRIM(?)))
    ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
    """

    row = cursor.execute(sql, member_name).fetchone()

    if not row:
        return None

    return {
        "LiteMemberId": row.LiteMemberId,
        "MemberDisplayId": row.MemberDisplayId,
        "MemberName": row.MemberName,
        "Phone": row.Phone,
        "EmergencyName": row.EmergencyName,
        "EmergencyPhone": row.EmergencyPhone,
    }


def find_existing_lite_member_by_original_card(
    cursor: pyodbc.Cursor,
    card_number: Optional[str],
) -> Optional[dict]:
    if not card_number:
        return None

    # Search Notes for the original Excel card indicator
    sql = """
    SELECT TOP 1
        LiteMemberId,
        MemberDisplayId,
        MemberName,
        Phone,
        EmergencyName,
        EmergencyPhone
    FROM dbo.SessionAttendance
    WHERE AttendanceMemberKind = 'LITE'
      AND Notes LIKE '%' + ? + '%'
    ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
    """

    note_search = f"Original Excel card value: {card_number}"

    row = cursor.execute(sql, note_search).fetchone()

    if not row:
        return None

    return {
        "LiteMemberId": row.LiteMemberId,
        "MemberDisplayId": row.MemberDisplayId,
        "MemberName": row.MemberName,
        "Phone": row.Phone,
        "EmergencyName": row.EmergencyName,
        "EmergencyPhone": row.EmergencyPhone,
    }


def find_name_for_card(cursor: pyodbc.Cursor, saheli_card_number: Optional[str]) -> Optional[str]:
    if not saheli_card_number:
        return None

    sql = """
    SELECT TOP 1
        MemberName
    FROM dbo.SessionAttendance
    WHERE SaheliCardNumber = ?
      AND LTRIM(RTRIM(ISNULL(MemberName, ''))) <> ''
    ORDER BY UpdatedAtUtc DESC, CreatedAtUtc DESC, AttendanceId DESC;
    """

    row = cursor.execute(sql, saheli_card_number).fetchone()

    if not row:
        return None

    return row.MemberName


def attendance_exists(
    cursor: pyodbc.Cursor,
    session_id: int,
    participant_id: Optional[int],
    saheli_card_number: Optional[str],
    member_name: Optional[str],
    lite_member_id: Optional[str],
    member_display_id: Optional[str],
) -> bool:
    conditions = []
    params: list[Any] = [session_id]

    if participant_id:
        conditions.append("ParticipantId = ?")
        params.append(participant_id)

    if saheli_card_number:
        conditions.append("SaheliCardNumber = ?")
        params.append(saheli_card_number)

    if lite_member_id:
        conditions.append("LiteMemberId = ?")
        params.append(lite_member_id)

    if member_display_id:
        conditions.append("MemberDisplayId = ?")
        params.append(member_display_id)

    if member_name:
        conditions.append("LOWER(LTRIM(RTRIM(ISNULL(MemberName, '')))) = LOWER(LTRIM(RTRIM(?)))")
        params.append(member_name)

    if not conditions:
        return False

    where_person = " OR ".join(f"({condition})" for condition in conditions)

    sql = f"""
    SELECT TOP 1 AttendanceId
    FROM dbo.SessionAttendance
    WHERE SessionId = ?
      AND ({where_person});
    """

    row = cursor.execute(sql, params).fetchone()
    return row is not None


def create_attendance(
    cursor: pyodbc.Cursor,
    attendance_identity: bool,
    manual_attendance_id: Optional[int],
    session_id: int,
    participant_id: Optional[int],
    activity_name: str,
    day_of_week: Optional[str],
    session_date: date,
    session_month: Optional[str],
    start_time: str,
    end_time: str,
    saheli_card_number: Optional[str],
    risk_stratification: Optional[str],
    notes: Optional[str],
    created_at: datetime,
    attendance_member_kind: str,
    lite_member_id: Optional[str],
    member_display_id: Optional[str],
    member_name: Optional[str],
    phone: Optional[str],
    emergency_name: Optional[str],
    emergency_phone: Optional[str],
) -> int:
    columns = [
        "SessionId",
        "ParticipantId",
        "SessionName",
        "SessionDay",
        "SessionDate",
        "SessionMonth",
        "SessionStartTime",
        "SessionEndTime",
        "SaheliCardNumber",
        "RiskStratification",
        "Attended",
        "CheckInTime",
        "CheckOutTime",
        "Notes",
        "CreatedAtUtc",
        "UpdatedAtUtc",
        "AttendanceMemberKind",
        "LiteMemberId",
        "MemberDisplayId",
        "MemberName",
        "Phone",
        "EmergencyName",
        "EmergencyPhone",
        "AF",
        "BP",
        "HeightCm",
        "WeightKg",
    ]

    values = [
        session_id,
        participant_id,
        activity_name,
        day_of_week,
        session_date.isoformat(),
        session_month,
        start_time,
        end_time,
        saheli_card_number,
        risk_stratification,
        1,
        start_time,
        end_time,
        notes,
        created_at,
        created_at,
        attendance_member_kind,
        lite_member_id,
        member_display_id,
        member_name,
        phone,
        emergency_name,
        emergency_phone,
        None,
        None,
        None,
        None,
    ]

    if not attendance_identity:
        if manual_attendance_id is None:
            raise RuntimeError("manual_attendance_id is required because AttendanceId is not an IDENTITY column.")

        columns.insert(0, "AttendanceId")
        values.insert(0, manual_attendance_id)

    column_sql = ", ".join(f"[{column}]" for column in columns)
    placeholder_sql = ", ".join("?" for _ in columns)

    sql = f"""
    INSERT INTO dbo.SessionAttendance ({column_sql})
    OUTPUT INSERTED.AttendanceId
    VALUES ({placeholder_sql});
    """

    try:
        row = cursor.execute(sql, values).fetchone()
    except Exception:
        print("Error inserting attendance:")
        print(f"SessionId: {session_id}")
        print(f"ParticipantId: {participant_id}")
        print(f"AttendanceMemberKind: {attendance_member_kind}")
        print(f"LiteMemberId: {lite_member_id}")
        print(f"SaheliCardNumber: {saheli_card_number}")
        print(f"MemberDisplayId: {member_display_id}")
        print(f"MemberName: {member_name}")
        print(f"Values: {values}")
        raise

    if not row:
        raise RuntimeError("Attendance insert failed. No AttendanceId was returned.")

    return int(row.AttendanceId)


# ============================================================
# MAIN IMPORT
# ============================================================

def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    if SQL_USER == "PUT_YOUR_SQL_USERNAME_HERE" or SQL_PASSWORD == "PUT_YOUR_SQL_PASSWORD_HERE":
        raise RuntimeError(
            "Please set SQL_USER and SQL_PASSWORD in the script, "
            "or set CRM_SQL_USER and CRM_SQL_PASSWORD environment variables."
        )

    created_at = datetime.now(timezone.utc).replace(microsecond=0, tzinfo=None)

    excel = pd.ExcelFile(INPUT_FILE)
    # Also open with openpyxl (data_only) for merged-cell lookup and richer cell inspection
    workbook = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    sessions_created = 0
    sessions_reused = 0
    sessions_cancelled = 0
    attendance_created = 0
    attendance_skipped_duplicate = 0
    attendance_failed = 0
    rows_skipped = 0
    rows_skipped_no_name = 0
    rows_skipped_missing_carried_datetime = 0
    invalid_time_rows = 0
    full_reused = 0
    full_created = 0
    lite_created = 0
    lite_reused = 0
    card_number_imported_as_lite = 0

    rows_imported_by_month: dict[str, int] = defaultdict(int)
    sessions_created_by_month: dict[str, int] = defaultdict(int)
    sessions_reused_by_month: dict[str, int] = defaultdict(int)

    session_id_cache: dict[tuple, int] = {}
    lite_member_cache: dict[str, dict] = {}

    with pyodbc.connect(SQL_CONNECTION_STRING) as conn:
        conn.autocommit = False
        cursor = conn.cursor()

        try:
            session_identity = is_identity_column(cursor, "Sessions", "SessionId")
            attendance_identity = is_identity_column(cursor, "SessionAttendance", "AttendanceId")

            next_manual_session_id = None
            next_manual_attendance_id = None

            if not session_identity:
                next_manual_session_id = get_next_manual_id(cursor, "Sessions", "SessionId")
                print(f"SessionId is not IDENTITY. Starting manual SessionId from {next_manual_session_id}")
            else:
                print("SessionId is IDENTITY. SQL Server will generate SessionId.")

            if not attendance_identity:
                next_manual_attendance_id = get_next_manual_id(cursor, "SessionAttendance", "AttendanceId")
                print(f"AttendanceId is not IDENTITY. Starting manual AttendanceId from {next_manual_attendance_id}")
            else:
                print("AttendanceId is IDENTITY. SQL Server will generate AttendanceId.")

            next_lite_number = get_next_lite_display_number(cursor)
            print(f"Next LITE MemberDisplayId starts from: LITE-{next_lite_number}")

            for sheet_name in excel.sheet_names:
                source_sheet_name = sheet_name.strip()
                sheet_activity_name = normalise_activity_from_sheet(source_sheet_name)

                if source_sheet_name in SKIP_SHEETS:
                    continue

                print(f"Processing sheet: {source_sheet_name}")

                df = pd.read_excel(
                    INPUT_FILE,
                    sheet_name=sheet_name,
                    header=0,
                    dtype=object,
                )

                # worksheet for merged-cell lookups
                try:
                    ws = workbook[sheet_name]
                except Exception:
                    ws = None

                if df.empty:
                    continue

                col_map = get_sheet_column_map(df)

                required_cols = ["session", "date", "time", "card"]

                if any(col_map.get(col) is None for col in required_cols):
                    print(f"Skipping sheet because required columns are missing: {source_sheet_name}")
                    print(f"Column map: {col_map}")
                    continue

                current_session_name = None
                current_day = None
                current_date = None
                current_month = None
                current_start_time = None
                current_end_time = None
                current_raw_time = None

                for idx, row in df.iterrows():
                    source_row = int(idx) + 2

                    raw_session = clean_text(get_row_value(row, col_map["session"]))
                    raw_day = clean_text(get_row_value(row, col_map["day"]))
                    raw_date = excel_date_to_date(get_row_value(row, col_map["date"]))
                    raw_month = clean_text(get_row_value(row, col_map["month"]))
                    # First attempt: value from the explicit time column
                    raw_time_value = get_row_value(row, col_map["time"])
                    raw_time_text = clean_text(raw_time_value)
                    parsed_start_time = parse_single_time(raw_time_value)

                    # If pandas did not read the merged Time cell, get the merged-cell value
                    # directly from openpyxl using the actual Time column.
                    if not parsed_start_time and ws is not None and col_map["time"] in df.columns:
                        time_col_index = list(df.columns).index(col_map["time"]) + 1
                        merged_time_value = get_cell_value_from_ws(ws, source_row, time_col_index)

                        if merged_time_value is not None:
                            parsed_start_time = parse_single_time(merged_time_value)
                            raw_time_text = clean_text(merged_time_value) or str(merged_time_value)

                    # If no time detected but sheet has configured default, use it
                    if not parsed_start_time and DEFAULT_SESSION_TIMES_BY_SHEET.get(source_sheet_name):
                        parsed_start_time = parse_single_time(DEFAULT_SESSION_TIMES_BY_SHEET[source_sheet_name])
                        raw_time_text = DEFAULT_SESSION_TIMES_BY_SHEET[source_sheet_name]

                    # IMPORTANT: define participant fields before any skip/debug checks
                    raw_card_value = get_row_value(row, col_map["card"])
                    raw_name_value = get_row_value(row, col_map["name"])

                    # Alum Rock Community Centre workbook has grouped rows.
                    # Session/date/time appear once at the top, then following participant rows are blank.
                    if raw_session:
                        current_session_name = raw_session

                    if raw_day:
                        current_day = raw_day

                    if raw_date:
                        current_date = raw_date

                    if raw_month:
                        current_month = raw_month

                    if parsed_start_time:
                        current_start_time = parsed_start_time
                        current_end_time = make_end_time(parsed_start_time)
                        current_raw_time = raw_time_text

                    # Diagnostic: if date is present but time is missing, for the first 20 occurrences per sheet
                    if current_date is not None and current_start_time is None:
                        # Use a per-sheet counter stored in sessions_reused_by_month keys area; use a local debug counter
                        debug_key = f"_debug_missing_time_{source_sheet_name}"
                        if debug_key not in locals():
                            locals()[debug_key] = 0
                        if locals()[debug_key] < 20:
                            # collect non-empty cells for this row and previous row
                            row_vals = []
                            prev_vals = []
                            for col_index, col_name in enumerate(df.columns, start=1):
                                v = get_row_value(row, col_name)
                                if v is None and ws is not None:
                                    v = get_cell_value_from_ws(ws, source_row, col_index)
                                if v is not None:
                                    row_vals.append((col_index, col_name, v))

                                # previous row
                                prev_row_idx = max(1, source_row - 1)
                                pv = None
                                if ws is not None:
                                    pv = get_cell_value_from_ws(ws, prev_row_idx, col_index)
                                if pv is not None:
                                    prev_vals.append((col_index, col_name, pv))

                            print("DEBUG missing time context:")
                            print(f"Sheet={source_sheet_name} Row={source_row}")
                            print(f"CurrentDate={current_date} CurrentTime={current_start_time}")
                            print("Row values:", ", ".join([f"{chr(64+c)}{source_row}={v}" for c,_,v in row_vals]))
                            print("Previous row values:", ", ".join([f"{chr(64+c)}{source_row-1}={v}" for c,_,v in prev_vals]))
                            locals()[debug_key] += 1

                    activity_name = sheet_activity_name
                    session_date = current_date
                    day_of_week = normalise_day(current_day, session_date)
                    session_month = month_name(current_month, session_date)
                    start_time = current_start_time
                    end_time = current_end_time
                    raw_time_for_notes = current_raw_time

                    # raw_card_value and raw_name_value were extracted earlier

                    if not activity_name or not session_date or not start_time or not end_time:
                        # If the row is completely blank (no card, no name, no time/session), don't spam logs
                        if not raw_card_value and not raw_name_value and not raw_session and not raw_time_text:
                            rows_skipped += 1
                            rows_skipped_missing_carried_datetime += 1
                            continue

                        rows_skipped += 1
                        rows_skipped_missing_carried_datetime += 1
                        print(
                            "Skipping row due to missing carried date/time:",
                            f"Sheet={source_sheet_name}",
                            f"Row={source_row}",
                            f"RawSession={raw_session}",
                            f"CurrentDate={session_date}",
                            f"CurrentTime={start_time}",
                            f"RawCard={clean_text(raw_card_value)}",
                            f"RawName={clean_text(raw_name_value)}",
                        )
                        continue

                    if end_time <= start_time:
                        invalid_time_rows += 1
                        rows_skipped += 1
                        continue

                    raw_emergency_name = get_row_value(row, col_map["emergency_name"])
                    raw_emergency_phone = get_row_value(row, col_map["emergency_phone"])
                    raw_risk = get_row_value(row, col_map["risk"])
                    raw_disability = get_row_value(row, col_map["disability"])

                    status = row_status(
                        raw_card_value,
                        raw_name_value,
                        raw_emergency_name,
                        raw_emergency_phone,
                    )

                    is_cancelled = status == "cancelled"

                    venue_name = DEFAULT_VENUE_NAME

                    session_key = (
                        venue_name.lower(),
                        activity_name.lower().strip(),
                        session_date.isoformat(),
                        start_time,
                        end_time,
                    )

                    if session_key in session_id_cache:
                        session_id = session_id_cache[session_key]
                    else:
                        existing_session_id = None

                        if REUSE_EXISTING_SESSIONS:
                            existing_session_id = find_existing_session(
                                cursor=cursor,
                                venue_name=venue_name,
                                activity_name=activity_name,
                                session_date=session_date,
                                start_time=start_time,
                                end_time=end_time,
                            )

                        if existing_session_id:
                            session_id = existing_session_id
                            session_id_cache[session_key] = session_id
                            sessions_reused += 1
                            # track reuse by month
                            month_key = session_month or (session_date.strftime("%B") if session_date else "Unknown")
                            sessions_reused_by_month[month_key] += 1

                            if is_cancelled:
                                update_session_cancelled(cursor, session_id)
                                sessions_cancelled += 1
                        else:
                            manual_session_id_to_use = None

                            if not session_identity:
                                manual_session_id_to_use = next_manual_session_id
                                next_manual_session_id += 1

                            session_id = create_session(
                                cursor=cursor,
                                session_identity=session_identity,
                                manual_session_id=manual_session_id_to_use,
                                venue_name=venue_name,
                                activity_name=activity_name,
                                session_date=session_date,
                                start_time=start_time,
                                end_time=end_time,
                                raw_time=raw_time_for_notes,
                                is_cancelled=is_cancelled,
                                created_at=created_at,
                                source_sheet=source_sheet_name,
                            )

                            session_id_cache[session_key] = session_id
                            sessions_created += 1

                            month_key = session_month or (session_date.strftime("%B") if session_date else "Unknown")
                            sessions_created_by_month[month_key] += 1

                            if is_cancelled:
                                sessions_cancelled += 1

                    if status in {"cancelled", "no_attendance"}:
                        continue

                    original_card_value = clean_text(raw_card_value)

                    card_number, member_name = extract_member(
                        raw_card_value,
                        raw_name_value,
                    )

                    # If there is no card and no name, this is not an attendance row.
                    if not card_number and not member_name:
                        rows_skipped += 1
                        rows_skipped_no_name += 1
                        continue

                    emergency_name_from_excel = clean_text(raw_emergency_name)
                    emergency_phone_from_excel = clean_text(raw_emergency_phone)
                    risk_stratification = clean_text(raw_risk)
                    disability = clean_text(raw_disability)

                    participant_id = None
                    phone = None
                    lite_member_id = None
                    member_display_id = None
                    attendance_member_kind = None
                    saheli_card_number_to_insert = None

                    existing_full_member = find_existing_full_member_from_attendance(
                        cursor=cursor,
                        saheli_card_number=card_number,
                    )

                    # ========================================================
                    # CASE 1: Existing FULL member found by SaheliCardNumber
                    # ========================================================
                    if existing_full_member:
                        attendance_member_kind = "FULL"
                        full_reused += 1

                        participant_id = existing_full_member.get("ParticipantId")
                        lite_member_id = None
                        saheli_card_number_to_insert = card_number

                        member_display_id = existing_full_member.get("MemberDisplayId") or card_number
                        member_name = member_name or existing_full_member.get("MemberName")
                        phone = existing_full_member.get("Phone")

                        emergency_name = emergency_name_from_excel or existing_full_member.get("EmergencyName")
                        emergency_phone = emergency_phone_from_excel or existing_full_member.get("EmergencyPhone")

                    # ========================================================
                    # FALLBACK: No existing FULL with ParticipantId -> treat as LITE
                    # ========================================================
                    else:
                        # If a card exists but no participant was found, try to get a name
                        # from any prior attendance row for this card. If still none,
                        # fall back to a Card {card_number} placeholder and import as LITE.
                        if card_number and not member_name:
                            name_from_card = find_name_for_card(cursor, card_number)
                            if name_from_card:
                                member_name = name_from_card
                            else:
                                member_name = f"Card {card_number}"

                        attendance_member_kind = "LITE"
                        participant_id = None
                        # Do not insert the Saheli card number for LITE rows
                        saheli_card_number_to_insert = None

                        # If this row originally had a card but no ParticipantId, count it
                        if card_number:
                            card_number_imported_as_lite += 1

                        # Determine cache key priority: prefer CARD-based reuse
                        if card_number:
                            lite_cache_key = f"CARD|{card_number}"
                        else:
                            lite_cache_key = f"NAME|{member_name.strip().lower()}|{(emergency_phone_from_excel or '').strip()}"

                        if lite_cache_key in lite_member_cache:
                            existing_lite = lite_member_cache[lite_cache_key]
                            lite_reused += 1
                        else:
                            existing_lite = None

                            # If we have an original card, try to find a LITE created previously for this card
                            if card_number:
                                existing_lite = find_existing_lite_member_by_original_card(cursor, card_number)

                            # Fallback to name+phone based lookup
                            if not existing_lite:
                                existing_lite = find_existing_lite_member(
                                    cursor=cursor,
                                    member_name=member_name,
                                    emergency_phone=emergency_phone_from_excel,
                                )

                            if existing_lite:
                                lite_reused += 1
                            else:
                                existing_lite = {
                                    "LiteMemberId": str(uuid.uuid4()).upper(),
                                    "MemberDisplayId": f"LITE-{next_lite_number}",
                                    "MemberName": member_name,
                                    "Phone": None,
                                    "EmergencyName": emergency_name_from_excel,
                                    "EmergencyPhone": emergency_phone_from_excel,
                                }

                                next_lite_number += 1
                                lite_created += 1

                            lite_member_cache[lite_cache_key] = existing_lite

                        lite_member_id = existing_lite.get("LiteMemberId")
                        member_display_id = existing_lite.get("MemberDisplayId")
                        member_name = member_name or existing_lite.get("MemberName")
                        phone = existing_lite.get("Phone")
                        emergency_name = emergency_name_from_excel or existing_lite.get("EmergencyName")
                        emergency_phone = emergency_phone_from_excel or existing_lite.get("EmergencyPhone")

                    if not member_display_id or not member_name:
                        print(
                            "Skipping attendance due to missing MemberDisplayId/MemberName:",
                            f"Sheet={source_sheet_name}",
                            f"Row={source_row}",
                            f"RawSession={raw_session}",
                            f"CurrentDate={session_date}",
                            f"CurrentTime={start_time}",
                            f"Card={card_number}",
                            f"Name={member_name}",
                        )
                        rows_skipped += 1
                        rows_skipped_no_name += 1
                        continue

                    if SKIP_DUPLICATE_ATTENDANCE and attendance_exists(
                        cursor=cursor,
                        session_id=session_id,
                        participant_id=participant_id,
                        saheli_card_number=saheli_card_number_to_insert,
                        member_name=member_name,
                        lite_member_id=lite_member_id,
                        member_display_id=member_display_id,
                    ):
                        attendance_skipped_duplicate += 1
                        continue

                    notes = build_attendance_notes(
                        source_sheet=source_sheet_name,
                        source_row=source_row,
                        raw_time=raw_time_for_notes,
                        disability=disability,
                        original_card_value=original_card_value,
                        attendance_member_kind=attendance_member_kind,
                    )

                    manual_attendance_id_to_use = None

                    if not attendance_identity:
                        manual_attendance_id_to_use = next_manual_attendance_id
                        next_manual_attendance_id += 1

                    try:
                        create_attendance(
                            cursor=cursor,
                            attendance_identity=attendance_identity,
                            manual_attendance_id=manual_attendance_id_to_use,
                            session_id=session_id,
                            participant_id=participant_id,
                            activity_name=activity_name,
                            day_of_week=day_of_week,
                            session_date=session_date,
                            session_month=session_month,
                            start_time=start_time,
                            end_time=end_time,
                            saheli_card_number=saheli_card_number_to_insert,
                            risk_stratification=risk_stratification,
                            notes=notes,
                            created_at=created_at,
                            attendance_member_kind=attendance_member_kind,
                            lite_member_id=lite_member_id,
                            member_display_id=member_display_id,
                            member_name=member_name,
                            phone=phone,
                            emergency_name=emergency_name,
                            emergency_phone=emergency_phone,
                        )

                        attendance_created += 1
                        if attendance_member_kind == "FULL":
                            full_created += 1
                        # count attendance by session month
                        month_key = session_month or (session_date.strftime("%B") if session_date else "Unknown")
                        rows_imported_by_month[month_key] += 1

                    except Exception as e:
                        attendance_failed += 1
                        print(
                            f"Warning: Could not insert attendance. "
                            f"Sheet={source_sheet_name}, Row={source_row}, "
                            f"SessionId={session_id}, Kind={attendance_member_kind}, "
                            f"Card={card_number}, LiteMemberId={lite_member_id}, "
                            f"DisplayId={member_display_id}, Name={member_name}, Error={e}"
                        )
                        continue

            conn.commit()

        except Exception:
            conn.rollback()
            raise

    print()
    print("Alum Rock Community Centre import completed successfully.")
    print("====================================")
    print(f"Sessions created: {sessions_created}")
    print(f"Sessions reused: {sessions_reused}")
    print(f"Cancelled sessions marked: {sessions_cancelled}")
    print(f"Attendance rows created: {attendance_created}")
    print(f"Duplicate attendance rows skipped: {attendance_skipped_duplicate}")
    print(f"Attendance rows failed: {attendance_failed}")
    print(f"FULL members reused from existing CRM attendance: {full_reused}")
    print(f"FULL rows created: {full_created}")
    print(f"LITE members created: {lite_created}")
    print(f"LITE members reused: {lite_reused}")
    print(f"Card-number rows converted to LITE (no ParticipantId): {card_number_imported_as_lite}")
    print(f"Rows skipped because no valid name: {rows_skipped_no_name}")
    print("\nRows imported by month:")
    for m, count in sorted(rows_imported_by_month.items()):
        print(f"{m}: {count}")

    print("\nSessions created by month:")
    for m, count in sorted(sessions_created_by_month.items()):
        print(f"{m}: {count}")

    print("\nSessions reused by month:")
    for m, count in sorted(sessions_reused_by_month.items()):
        print(f"{m}: {count}")

    print(f"Rows skipped missing carried date/time: {rows_skipped_missing_carried_datetime}")
    print(f"Invalid time rows skipped: {invalid_time_rows}")
    print(f"Rows skipped: {rows_skipped}")


if __name__ == "__main__":
    main()