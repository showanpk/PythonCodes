#!/usr/bin/env python3
"""
Safely migrate the cleaned ARCC 2025 staging workbooks into Saheli CRM.

FILES EXPECTED IN THE SAME FOLDER
---------------------------------
ARCC_2025_CRM_Sessions_Staging.xlsx
ARCC_2025_CRM_Participants_Attendance_Staging.xlsx

SAFETY
------
COMMIT_CHANGES = False by default.
The script will complete the migration inside a SQL transaction and then ROLLBACK.
It also creates a CSV audit file showing what would be inserted or skipped.

After reviewing the audit, change:
    COMMIT_CHANGES = True

FULL participants:
    Matched only by SaheliCardNumber in dbo.Participants.
    The script does not create new FULL participants.

LITE participants:
    Existing LiteMembers are matched by ProposedMembershipId first,
    then by one exact normalised name.
    New LiteMembers are only created when CREATE_MISSING_LITE_MEMBERS = True.

DUPLICATE PROTECTION
--------------------
Session:
    Venue + Activity + SessionDate + StartTime + EndTime

FULL attendance:
    SessionId + ParticipantId

LITE attendance:
    SessionId + LiteMemberId
"""

from __future__ import annotations

import csv
import re
import sys
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import pyodbc
from openpyxl import load_workbook


# ============================================================
# 1. SQL SERVER CONNECTION STRING
# ============================================================
# Replace the placeholder values below with your SQL Server details.

CONNECTION_STRING = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=tcp:sahelihub.database.windows.net,1433;"
    "DATABASE=SahelihubCRM;"
    "UID=sahelihubadmin;"
    "PWD=W7WZ7ZaG1YbMZ71gh%2xSFuR;"
    "Encrypt=yes;"
    "TrustServerCertificate=yes;"
)

# Windows Authentication alternative:
#
# CONNECTION_STRING = (
#     "DRIVER={ODBC Driver 18 for SQL Server};"
#     "SERVER=YOUR_SERVER_NAME;"
#     "DATABASE=YOUR_DATABASE_NAME;"
#     "Trusted_Connection=yes;"
#     "Encrypt=yes;"
#     "TrustServerCertificate=yes;"
# )


# ============================================================
# 2. MIGRATION SETTINGS
# ============================================================

SESSIONS_FILE = "ARCC_2025_CRM_Sessions_Staging.xlsx"
PARTICIPANTS_FILE = "ARCC_2025_CRM_Participants_Attendance_Staging.xlsx"

# Keep False for the first run.
COMMIT_CHANGES = False

# Keep False for the first preview.
# Change to True only after reviewing NEW_LITE_REQUIRED rows in the audit.
CREATE_MISSING_LITE_MEMBERS = False

# Optional User/Staff ID stored in LiteMembers.CreatedByUserId.
# Leave as None when not known.
CREATED_BY_USER_ID: int | None = None

EXPECTED_SESSIONS = 156
EXPECTED_PARTICIPANTS = 179
EXPECTED_ATTENDANCE = 1023

REPORT_START = date(2025, 4, 1)
REPORT_END = date(2025, 12, 31)


# ============================================================
# 3. NORMALISATION RULES
# ============================================================

VENUE_ALIASES = {
    "alum rock community centre": "alum rock community centre",
    "arcc": "alum rock community centre",
    "alum rock": "alum rock community centre",
}

ACTIVITY_ALIASES = {
    "chair exercise": "chair based exercise",
    "chair based exercise": "chair based exercise",
    "circuit": "circuit training",
    "circuit training": "circuit training",
    "mens exercise": "mens circuit exercise",
    "men s exercise": "mens circuit exercise",
    "mens circuit exercise": "mens circuit exercise",
    "men s circuit exercise": "mens circuit exercise",
    "strength and stretch": "strength & stretch",
    "strength stretch": "strength & stretch",
    "strength & stretch": "strength & stretch",
    "art social": "arts",
    "saheli social": "arts",
    "arts": "arts",
    "crochet knit": "crochet",
    "crochet": "crochet",
    "salsa": "salsa belly dancing",
    "salsa belly dancing": "salsa belly dancing",
    "self defence": "self defence",
    "body conditioning": "body conditioning",
    "walk talk": "walk & talk",
    "walk & talk": "walk & talk",
    "pilates": "pilates",
    "yoga": "yoga",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).strip()
    if value.lower() in {"", "none", "nan", "#n/a", "n/a"}:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def normalise_card(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return str(int(float(text)))
    return text


def normalise_name(value: Any) -> str:
    text = clean_text(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def canonical_venue(value: Any) -> str:
    name = normalise_name(value)
    return VENUE_ALIASES.get(name, name)


def canonical_activity(value: Any) -> str:
    name = normalise_name(value)
    return ACTIVITY_ALIASES.get(name, name)


def parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"Unable to parse date: {value!r}")


def parse_optional_date(value: Any) -> date | None:
    text = clean_text(value)
    if not text:
        return None

    # The staging workbook may contain a date as YYYY-MM-DD or DD/MM/YYYY.
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    return None


def parse_time(value: Any) -> time:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)

    text = clean_text(value)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass

    raise ValueError(f"Unable to parse time: {value!r}")


def to_int(value: Any, default: int | None = None) -> int | None:
    text = clean_text(value)
    if text == "":
        return default
    return int(float(text))


def first_semicolon_value(value: Any) -> str:
    text = clean_text(value)
    return text.split(";")[0].strip() if text else ""


# ============================================================
# 4. EXCEL READING
# ============================================================

def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Missing sheet {sheet_name!r} in {path}")

        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(iterator)]

        rows: list[dict[str, Any]] = []

        for values in iterator:
            if all(value is None or clean_text(value) == "" for value in values):
                continue

            row = {
                headers[index]: values[index] if index < len(values) else None
                for index in range(len(headers))
                if headers[index]
            }
            rows.append(row)

        return rows
    finally:
        workbook.close()


# ============================================================
# 5. DATABASE LOOKUPS
# ============================================================

def build_session_key(
    venue_name: Any,
    activity_name: Any,
    session_date: date,
    start_time: time,
    end_time: time,
) -> tuple[str, str, date, time, time]:
    return (
        canonical_venue(venue_name),
        canonical_activity(activity_name),
        session_date,
        start_time,
        end_time,
    )


def fetch_existing_sessions(
    cursor: pyodbc.Cursor,
) -> dict[tuple[str, str, date, time, time], list[dict[str, Any]]]:
    cursor.execute(
        """
        SELECT
            s.SessionId,
            s.VenueName,
            s.ActivityName,
            s.SessionDate,
            s.StartTime,
            s.EndTime,
            s.IsCancelled,
            COUNT(a.AttendanceId) AS AttendanceRows
        FROM dbo.Sessions s
        LEFT JOIN dbo.SessionAttendance a
            ON a.SessionId = s.SessionId
        WHERE s.SessionDate >= ?
          AND s.SessionDate <= ?
        GROUP BY
            s.SessionId,
            s.VenueName,
            s.ActivityName,
            s.SessionDate,
            s.StartTime,
            s.EndTime,
            s.IsCancelled
        """,
        REPORT_START,
        REPORT_END,
    )

    grouped: dict[
        tuple[str, str, date, time, time],
        list[dict[str, Any]],
    ] = defaultdict(list)

    for row in cursor.fetchall():
        if row.SessionDate is None or row.StartTime is None or row.EndTime is None:
            continue

        key = build_session_key(
            row.VenueName,
            row.ActivityName,
            row.SessionDate,
            row.StartTime,
            row.EndTime,
        )

        grouped[key].append(
            {
                "session_id": int(row.SessionId),
                "attendance_rows": int(row.AttendanceRows or 0),
                "is_cancelled": int(row.IsCancelled or 0),
                "venue": clean_text(row.VenueName),
                "activity": clean_text(row.ActivityName),
            }
        )

    return grouped


def choose_preferred_session(
    candidates: list[dict[str, Any]],
) -> dict[str, Any]:
    """
    When the CRM already contains duplicate sessions, keep the session
    with the most attendance, then the non-cancelled session, then the
    lowest SessionId.
    """
    return sorted(
        candidates,
        key=lambda item: (
            -item["attendance_rows"],
            item["is_cancelled"],
            item["session_id"],
        ),
    )[0]


def fetch_full_participants(
    cursor: pyodbc.Cursor,
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    cursor.execute(
        """
        SELECT
            ParticipantID,
            SaheliCardNumber,
            FullName,
            MobileNumber,
            Postcode,
            DateOfBirth
        FROM dbo.Participants
        """
    )

    by_card: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in cursor.fetchall():
        item = {
            "participant_id": int(row.ParticipantID),
            "card": normalise_card(row.SaheliCardNumber),
            "name": clean_text(row.FullName),
            "phone": clean_text(row.MobileNumber),
            "postcode": clean_text(row.Postcode),
            "dob": row.DateOfBirth,
        }

        if item["card"]:
            by_card[item["card"]] = item

        normalised = normalise_name(item["name"])
        if normalised:
            by_name[normalised].append(item)

    return by_card, by_name


def fetch_lite_members(
    cursor: pyodbc.Cursor,
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    cursor.execute(
        """
        SELECT
            Id,
            MembershipId,
            FirstName,
            LastName,
            Phone,
            Postcode,
            DateOfBirth,
            EmergencyName,
            EmergencyPhone
        FROM dbo.LiteMembers
        """
    )

    by_membership: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in cursor.fetchall():
        full_name = clean_text(
            f"{clean_text(row.FirstName)} {clean_text(row.LastName)}"
        )

        item = {
            "lite_member_id": row.Id,
            "membership_id": clean_text(row.MembershipId),
            "name": full_name,
            "phone": clean_text(row.Phone),
            "postcode": clean_text(row.Postcode),
            "dob": row.DateOfBirth,
            "emergency_name": clean_text(row.EmergencyName),
            "emergency_phone": clean_text(row.EmergencyPhone),
        }

        if item["membership_id"]:
            by_membership[item["membership_id"]] = item

        normalised = normalise_name(full_name)
        if normalised:
            by_name[normalised].append(item)

    return by_membership, by_name


# ============================================================
# 6. INSERT FUNCTIONS
# ============================================================

def insert_session(
    cursor: pyodbc.Cursor,
    row: dict[str, Any],
) -> int:
    session_date = parse_date(row["SessionDate"])
    start_time = parse_time(row["StartTime"])
    end_time = parse_time(row["EndTime"])

    if end_time <= start_time:
        raise ValueError(
            f"End time must be later than start time: "
            f"{row.get('ImportSessionKey')}"
        )

    cursor.execute(
        """
        INSERT INTO dbo.Sessions
        (
            Frequency,
            Category,
            SubCategory,
            ActivityCategory,
            VenueName,
            ActivityName,
            Notes,
            IsRecurringWeekly,
            DayOfWeek,
            SessionDate,
            ArrivalTime,
            StartTime,
            EndTime,
            Capacity,
            IsBookingRequired,
            IsCancelled,
            AssignedStaffId,
            RecurringSeriesId,
            SessionProviderId
        )
        OUTPUT INSERTED.SessionId
        VALUES
        (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )
        """,
        clean_text(row.get("Frequency")) or "WEEKLY",
        clean_text(row.get("Category")) or "Fitness",
        clean_text(row.get("SubCategory")) or None,
        clean_text(row.get("ActivityCategory")) or "Fitness",
        clean_text(row.get("VenueName")),
        clean_text(row.get("ActivityName")),
        clean_text(row.get("Notes")) or None,
        0,
        None,
        session_date,
        None,
        start_time,
        end_time,
        to_int(row.get("Capacity")),
        to_int(row.get("IsBookingRequired"), 0),
        to_int(row.get("IsCancelled"), 0),
        to_int(row.get("AssignedStaffId")),
        None,
        to_int(row.get("SessionProviderId")),
    )

    return int(cursor.fetchone()[0])


def create_lite_member(
    cursor: pyodbc.Cursor,
    row: dict[str, Any],
) -> dict[str, Any]:
    membership_id = clean_text(row.get("ProposedMembershipId"))
    if not membership_id:
        raise ValueError(
            f"Missing ProposedMembershipId for "
            f"{row.get('SourceParticipantKey')}"
        )

    lite_id = uuid.uuid5(
        uuid.NAMESPACE_URL,
        f"saheli-arcc-2025-lite:{membership_id}",
    )

    first_name = clean_text(row.get("FirstName"))
    last_name = clean_text(row.get("LastName")) or "(Not provided)"
    date_of_birth = parse_optional_date(row.get("DOB"))
    postcode = clean_text(row.get("Postcode"))
    emergency_name = clean_text(row.get("EmergencyName"))
    emergency_phone = clean_text(row.get("EmergencyPhone"))
    gender = clean_text(row.get("Gender"))

    cursor.execute(
        """
        INSERT INTO dbo.LiteMembers
        (
            Id,
            MembershipId,
            FirstName,
            LastName,
            DateOfBirth,
            Phone,
            Email,
            Address,
            Postcode,
            EmergencyName,
            EmergencyPhone,
            EmergencyRelation,
            HealthConditions,
            Gender,
            Ethnicity,
            CreatedByUserId
        )
        VALUES
        (
            ?, ?, ?, ?, ?, NULL, NULL, NULL, ?, ?, ?, NULL, NULL, ?, NULL, ?
        )
        """,
        str(lite_id),
        membership_id,
        first_name,
        last_name,
        date_of_birth,
        postcode or None,
        emergency_name or None,
        emergency_phone or None,
        gender or None,
        CREATED_BY_USER_ID,
    )

    return {
        "kind": "LITE",
        "lite_member_id": str(lite_id),
        "membership_id": membership_id,
        "name": clean_text(f"{first_name} {last_name}"),
        "phone": "",
        "postcode": postcode,
        "dob": date_of_birth,
    }


def insert_attendance(
    cursor: pyodbc.Cursor,
    row: dict[str, Any],
    session_id: int,
    member: dict[str, Any],
) -> int:
    session_date = parse_date(row["SessionDate"])
    start_time = parse_time(row["SessionStartTime"])
    end_time = parse_time(row["SessionEndTime"])

    if member["kind"] == "FULL":
        participant_id = member["participant_id"]
        lite_member_id = None
        card_number = member["card"]
        display_id = member["card"]
    else:
        participant_id = None
        lite_member_id = member["lite_member_id"]
        card_number = None
        display_id = member["membership_id"]

    source_note = (
        "Historical ARCC 2025 import"
        f" | workbook {clean_text(row.get('SourceWorkbook'))}"
        f" | sheet {clean_text(row.get('SourceSheet'))}"
        f" | source {clean_text(row.get('SourceRegisterType'))}"
        f" | activity {clean_text(row.get('SourceActivityLabel'))}"
        f" | key {clean_text(row.get('ImportAttendanceKey'))}"
    )

    cursor.execute(
        """
        INSERT INTO dbo.SessionAttendance
        (
            SessionId,
            ParticipantId,
            SessionName,
            SessionDay,
            SessionDate,
            SessionMonth,
            SessionStartTime,
            SessionEndTime,
            SaheliCardNumber,
            RiskStratification,
            Attended,
            CheckInTime,
            CheckOutTime,
            Notes,
            AttendanceMemberKind,
            LiteMemberId,
            MemberDisplayId,
            MemberName,
            Phone,
            EmergencyName,
            EmergencyPhone
        )
        OUTPUT INSERTED.AttendanceId
        VALUES
        (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )
        """,
        session_id,
        participant_id,
        clean_text(row.get("ActivityName")),
        clean_text(row.get("SessionDay"))
        or session_date.strftime("%A"),
        session_date,
        clean_text(row.get("SessionMonth"))
        or session_date.strftime("%B"),
        start_time,
        end_time,
        card_number,
        clean_text(row.get("RiskStratification")) or None,
        1,
        start_time,
        end_time,
        source_note[:2000],
        member["kind"],
        lite_member_id,
        display_id,
        member["name"]
        or clean_text(row.get("MemberName"))
        or display_id,
        member.get("phone") or None,
        clean_text(row.get("EmergencyName")) or None,
        clean_text(row.get("EmergencyPhone")) or None,
    )

    return int(cursor.fetchone()[0])


# ============================================================
# 7. AUDIT
# ============================================================

def add_audit(
    audit: list[dict[str, Any]],
    entity: str,
    source_key: str,
    action: str,
    database_id: Any = "",
    details: str = "",
) -> None:
    audit.append(
        {
            "Entity": entity,
            "SourceKey": source_key,
            "Action": action,
            "DatabaseId": database_id,
            "Details": details,
        }
    )


def save_audit(
    audit: list[dict[str, Any]],
    mode: str,
) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path.cwd() / f"arcc_2025_migration_audit_{timestamp}.csv"

    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "Mode",
                "Entity",
                "SourceKey",
                "Action",
                "DatabaseId",
                "Details",
            ],
        )
        writer.writeheader()

        for row in audit:
            writer.writerow({"Mode": mode, **row})

    return path


# ============================================================
# 8. MAIN MIGRATION
# ============================================================

def main() -> int:
    base_folder = Path(__file__).resolve().parent

    sessions_path = base_folder / SESSIONS_FILE
    participants_path = base_folder / PARTICIPANTS_FILE

    if not sessions_path.exists():
        raise FileNotFoundError(sessions_path)

    if not participants_path.exists():
        raise FileNotFoundError(participants_path)

    session_rows = read_sheet(sessions_path, "Sessions")
    participant_rows = read_sheet(participants_path, "Participants")
    attendance_rows = read_sheet(participants_path, "Attendance")

    actual_totals = (
        len(session_rows),
        len(participant_rows),
        len(attendance_rows),
    )

    expected_totals = (
        EXPECTED_SESSIONS,
        EXPECTED_PARTICIPANTS,
        EXPECTED_ATTENDANCE,
    )

    if actual_totals != expected_totals:
        raise RuntimeError(
            f"Staging totals do not match. "
            f"Expected {expected_totals}, received {actual_totals}."
        )

    audit: list[dict[str, Any]] = []
    summary: Counter[str] = Counter()

    connection = pyodbc.connect(
        CONNECTION_STRING,
        autocommit=False,
    )
    cursor = connection.cursor()
    cursor.execute("SET XACT_ABORT ON;")

    try:
        # ----------------------------------------------------
        # A. MATCH OR CREATE SESSIONS
        # ----------------------------------------------------
        existing_sessions = fetch_existing_sessions(cursor)
        resolved_session_ids: dict[str, int] = {}

        for row in session_rows:
            import_key = clean_text(row["ImportSessionKey"])

            exact_key = build_session_key(
                row["VenueName"],
                row["ActivityName"],
                parse_date(row["SessionDate"]),
                parse_time(row["StartTime"]),
                parse_time(row["EndTime"]),
            )

            candidates = existing_sessions.get(exact_key, [])

            if candidates:
                preferred = choose_preferred_session(candidates)
                session_id = preferred["session_id"]
                action = "EXISTING_SESSION"
                details = (
                    f"Matched {len(candidates)} CRM session(s). "
                    f"Selected SessionId {session_id} with "
                    f"{preferred['attendance_rows']} attendance row(s)."
                )
            else:
                session_id = insert_session(cursor, row)
                action = "NEW_SESSION"
                details = "Inserted inside the current SQL transaction."

                existing_sessions[exact_key] = [
                    {
                        "session_id": session_id,
                        "attendance_rows": 0,
                        "is_cancelled": 0,
                        "venue": clean_text(row["VenueName"]),
                        "activity": clean_text(row["ActivityName"]),
                    }
                ]

            resolved_session_ids[import_key] = session_id
            summary[action] += 1
            add_audit(
                audit,
                "SESSION",
                import_key,
                action,
                session_id,
                details,
            )

        # ----------------------------------------------------
        # B. MATCH PARTICIPANTS
        # ----------------------------------------------------
        participants_by_card, full_by_name = fetch_full_participants(cursor)
        lite_by_membership, lite_by_name = fetch_lite_members(cursor)

        resolved_members: dict[str, dict[str, Any]] = {}

        for row in participant_rows:
            source_key = clean_text(row["SourceParticipantKey"])
            expected_kind = clean_text(row["ExpectedMemberKind"]).upper()
            card = normalise_card(row.get("SaheliCardNumber"))
            source_name = normalise_name(
                row.get("NormalisedName")
                or row.get("SourceName")
            )

            member: dict[str, Any] | None = None
            action = ""
            details = ""

            if expected_kind == "FULL":
                full = participants_by_card.get(card)

                if full:
                    member = {"kind": "FULL", **full}
                    action = "MATCHED_FULL"
                else:
                    action = "UNMATCHED_CARD"
                    details = (
                        f"Saheli Card {card!r} was not found in "
                        "dbo.Participants. No FULL participant was created."
                    )

            else:
                membership_id = clean_text(
                    row.get("ProposedMembershipId")
                )

                existing_lite = lite_by_membership.get(membership_id)

                if existing_lite:
                    member = {"kind": "LITE", **existing_lite}
                    action = "MATCHED_LITE_MEMBERSHIP_ID"

                else:
                    lite_candidates = lite_by_name.get(source_name, [])
                    full_candidates = full_by_name.get(source_name, [])

                    if len(lite_candidates) == 1:
                        member = {"kind": "LITE", **lite_candidates[0]}
                        action = "MATCHED_LITE_NAME"

                    elif len(lite_candidates) > 1:
                        action = "AMBIGUOUS_LITE_NAME"
                        details = (
                            f"{len(lite_candidates)} LiteMembers have "
                            "the same normalised name."
                        )

                    elif len(full_candidates) == 1:
                        action = "REVIEW_FULL_NAME_MATCH"
                        details = (
                            "A FULL participant has the same name, but "
                            "the source record has no Saheli Card Number. "
                            "This row was held for review."
                        )

                    elif len(full_candidates) > 1:
                        action = "AMBIGUOUS_FULL_NAME"
                        details = (
                            f"{len(full_candidates)} FULL participants "
                            "have the same normalised name."
                        )

                    elif CREATE_MISSING_LITE_MEMBERS:
                        member = create_lite_member(cursor, row)
                        action = "CREATED_LITE"

                        lite_by_membership[
                            member["membership_id"]
                        ] = member
                        lite_by_name[source_name].append(member)

                    else:
                        action = "NEW_LITE_REQUIRED"
                        details = (
                            "No existing Lite Member match was found. "
                            "Set CREATE_MISSING_LITE_MEMBERS = True "
                            "after reviewing the preview audit."
                        )

            if member:
                resolved_members[source_key] = member

                database_id = (
                    member.get("participant_id")
                    if member["kind"] == "FULL"
                    else member.get("lite_member_id")
                )
            else:
                database_id = ""

            summary[action] += 1
            add_audit(
                audit,
                "PARTICIPANT",
                source_key,
                action,
                database_id,
                details,
            )

        # ----------------------------------------------------
        # C. LOAD EXISTING ATTENDANCE
        # ----------------------------------------------------
        unique_session_ids = sorted(
            set(resolved_session_ids.values())
        )

        existing_attendance: dict[
            tuple[int, str, str],
            int,
        ] = {}

        if unique_session_ids:
            placeholders = ",".join("?" for _ in unique_session_ids)

            cursor.execute(
                f"""
                SELECT
                    AttendanceId,
                    SessionId,
                    AttendanceMemberKind,
                    ParticipantId,
                    LiteMemberId
                FROM dbo.SessionAttendance
                WHERE SessionId IN ({placeholders})
                """,
                *unique_session_ids,
            )

            for row in cursor.fetchall():
                kind = clean_text(
                    row.AttendanceMemberKind
                ).upper()

                if kind == "FULL" and row.ParticipantId is not None:
                    member_id = str(int(row.ParticipantId))

                elif kind == "LITE" and row.LiteMemberId is not None:
                    member_id = str(row.LiteMemberId).lower()

                else:
                    continue

                existing_attendance[
                    (int(row.SessionId), kind, member_id)
                ] = int(row.AttendanceId)

        # ----------------------------------------------------
        # D. INSERT ONLY MISSING ATTENDANCE
        # ----------------------------------------------------
        for row in attendance_rows:
            attendance_key = clean_text(
                row["ImportAttendanceKey"]
            )
            import_session_key = clean_text(
                row["ImportSessionKey"]
            )
            source_participant_key = clean_text(
                row["SourceParticipantKey"]
            )

            session_id = resolved_session_ids.get(
                import_session_key
            )
            member = resolved_members.get(
                source_participant_key
            )

            if not session_id:
                action = "UNMATCHED_SESSION"
                details = (
                    f"No SessionId was resolved for "
                    f"{import_session_key}."
                )
                summary[action] += 1
                add_audit(
                    audit,
                    "ATTENDANCE",
                    attendance_key,
                    action,
                    "",
                    details,
                )
                continue

            if not member:
                action = "UNMATCHED_PARTICIPANT"
                details = (
                    f"No safe participant match was resolved for "
                    f"{source_participant_key}."
                )
                summary[action] += 1
                add_audit(
                    audit,
                    "ATTENDANCE",
                    attendance_key,
                    action,
                    "",
                    details,
                )
                continue

            if member["kind"] == "FULL":
                member_id = str(member["participant_id"])
            else:
                member_id = str(
                    member["lite_member_id"]
                ).lower()

            duplicate_key = (
                session_id,
                member["kind"],
                member_id,
            )

            existing_attendance_id = existing_attendance.get(
                duplicate_key
            )

            if existing_attendance_id:
                action = "ALREADY_IN_CRM"
                details = (
                    "Skipped because this participant already has "
                    "attendance for the resolved session."
                )
                summary[action] += 1
                add_audit(
                    audit,
                    "ATTENDANCE",
                    attendance_key,
                    action,
                    existing_attendance_id,
                    details,
                )
                continue

            attendance_id = insert_attendance(
                cursor,
                row,
                session_id,
                member,
            )

            existing_attendance[
                duplicate_key
            ] = attendance_id

            action = "NEW_ATTENDANCE"
            summary[action] += 1
            add_audit(
                audit,
                "ATTENDANCE",
                attendance_key,
                action,
                attendance_id,
                "Inserted inside the current SQL transaction.",
            )

        # ----------------------------------------------------
        # E. COMMIT OR ROLLBACK
        # ----------------------------------------------------
        if COMMIT_CHANGES:
            connection.commit()
            mode = "COMMITTED"
        else:
            connection.rollback()
            mode = "PREVIEW_ROLLED_BACK"

    except Exception:
        connection.rollback()
        raise

    finally:
        cursor.close()
        connection.close()

    audit_path = save_audit(audit, mode)

    print()
    print("=" * 68)
    print("ARCC 2025 CRM MIGRATION")
    print("=" * 68)
    print("Mode:", mode)
    print("Sessions file:", sessions_path)
    print("Participants file:", participants_path)
    print("Audit file:", audit_path)
    print()
    print("Action summary")
    print("-" * 68)

    for action, count in sorted(summary.items()):
        print(f"{action:38s} {count:>6}")

    print()

    if not COMMIT_CHANGES:
        print("No database changes were saved.")
        print("Review the audit CSV before setting COMMIT_CHANGES = True.")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print()
        print("MIGRATION FAILED:", error, file=sys.stderr)
        raise
