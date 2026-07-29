#!/usr/bin/env python3
"""
Duplicate-safe Calthorpe 2025/26 migration into SahelihubCRM.

Expected files in the same folder:
    Calthorpe_2025_26_CRM_Sessions_Staging.xlsx
    Calthorpe_2025_26_CRM_Participants_Attendance_Staging.xlsx

The migration package contains:
    430 sessions
    386 source participants
    3,598 unique participant/session attendance records

Safety:
    COMMIT_CHANGES = False by default.

The first run completes the matching and insertion logic inside a SQL
transaction, creates an audit CSV, and rolls everything back.

After reviewing the audit:
    1. Set CREATE_MISSING_LITE_MEMBERS = True and preview again.
    2. Set COMMIT_CHANGES = True only when the preview is correct.
"""

from __future__ import annotations

import csv
import re
import sys
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import pyodbc
from openpyxl import load_workbook


# ============================================================
# 1. SQL SERVER CONNECTION STRING
# ============================================================
# The server and database are filled from your SQL Server setup.
# Replace only YOUR_SQL_USERNAME and YOUR_SQL_PASSWORD.

CONNECTION_STRING = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=tcp:sahelihub.database.windows.net,1433;"
    "DATABASE=SahelihubCRM;"
    "UID=sahelihubadmin;"
    "PWD=W7WZ7ZaG1YbMZ71gh%2xSFuR;"
    "Encrypt=yes;"
    "TrustServerCertificate=yes;"
)

# Windows Authentication alternative for a local/on-premises SQL Server:
#
# CONNECTION_STRING = (
#     "DRIVER={ODBC Driver 18 for SQL Server};"
#     "SERVER=YOUR_SERVER_NAME;"
#     "DATABASE=SahelihubCRM;"
#     "Trusted_Connection=yes;"
#     "Encrypt=yes;"
#     "TrustServerCertificate=yes;"
# )


# ============================================================
# 2. MIGRATION SETTINGS
# ============================================================

SESSIONS_FILE = "Calthorpe_2025_26_CRM_Sessions_Staging.xlsx"
PARTICIPANTS_FILE = (
    "Calthorpe_2025_26_CRM_Participants_Attendance_Staging.xlsx"
)

# First run: keep both False.
COMMIT_CHANGES = False
CREATE_MISSING_LITE_MEMBERS = True

# Optional application user ID for LiteMembers.CreatedByUserId.
CREATED_BY_USER_ID: int | None = None

EXPECTED_SESSIONS = 430
EXPECTED_PARTICIPANTS = 386
EXPECTED_ATTENDANCE = 3598

REPORT_START = date(2025, 11, 3)
REPORT_END = date(2026, 6, 9)


# ============================================================
# 3. TEXT, DATE AND TIME HELPERS
# ============================================================

def clean_text(value: Any) -> str:
    if value is None:
        return ""

    text = re.sub(r"\s+", " ", str(value).strip())

    if text.lower() in {
        "",
        "none",
        "nan",
        "#n/a",
        "#ref!",
        "#value!",
        "#name?",
        "#div/0!",
        "#null!",
        "#num!",
        "n/a",
        "null",
    }:
        return ""

    return text


def normalise_card(value: Any) -> str:
    text = clean_text(value)

    if not text:
        return ""

    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return str(int(float(text)))

    return text


def normalise_name(value: Any) -> str:
    text = clean_text(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = clean_text(value)

    for date_format in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            pass

    raise ValueError(f"Unable to parse date: {value!r}")


def parse_optional_date(value: Any) -> date | None:
    text = clean_text(value)

    if not text:
        return None

    for date_format in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            pass

    return None


def parse_time(value: Any) -> time:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)

    if isinstance(value, time):
        return value.replace(microsecond=0)

    text = clean_text(value)

    for time_format in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, time_format).time()
        except ValueError:
            pass

    raise ValueError(f"Unable to parse time: {value!r}")


def to_int(value: Any, default: int | None = None) -> int | None:
    text = clean_text(value)

    if not text:
        return default

    return int(float(text))


def normalise_phone_for_database(value: Any) -> str | None:
    """
    Accept only realistic telephone numbers.

    Values containing names, notes or other letters are rejected and
    inserted as NULL rather than being truncated.
    """
    text = clean_text(value)

    if not text:
        return None

    if re.search(r"[A-Za-z]", text):
        return None

    has_plus = text.startswith("+")
    digits = re.sub(r"\D", "", text)

    if not 7 <= len(digits) <= 15:
        return None

    return ("+" if has_plus else "") + digits


def safe_database_text(
    value: Any,
    maximum_length: int | None,
) -> str | None:
    text = clean_text(value)

    if not text:
        return None

    if maximum_length is None:
        return text

    return text[:maximum_length]


# ============================================================
# 4. VENUE AND ACTIVITY NORMALISATION
# ============================================================

VENUE_ALIASES = {
    "calthorpe": "calthorpe wellbeing hub",
    "calthorpe wellbeing hub": "calthorpe wellbeing hub",
}

ACTIVITY_ALIASES = {
    "a c": "arts",
    "arts and craft": "arts",
    "arts": "arts",

    "social knit and crochet": "crochet",
    "crochet": "crochet",

    "chair exercise": "chair based exercise",
    "chair based exercise": "chair based exercise",

    "circuit": "circuit training",
    "circuits class": "circuit training",
    "circuit training": "circuit training",

    "strength and stretch": "strength & stretch",
    "strength stretch": "strength & stretch",
    "strength & stretch": "strength & stretch",

    "pilate": "pilates",
    "pilate floor base": "pilates",
    "pilates": "pilates",

    "salsa": "salsa belly dancing",
    "belly dancing salsa": "salsa belly dancing",
    "salsa belly dancing": "salsa belly dancing",

    "mens multisports": "mens multi sports",
    "mens multi sports": "mens multi sports",
    "men s multi sports": "mens multi sports",

    "workshop": "workshops",
    "workshops": "workshops",

    "walk with abi": "walk with abi",
    "body conditioning": "body conditioning",
    "aerobics": "aerobics",
    "hiit": "hiit",
    "yoga": "yoga",
    "zumba": "zumba",
    "tennis": "tennis",
    "esol": "esol",
}


def canonical_venue(value: Any) -> str:
    name = normalise_name(value)
    return VENUE_ALIASES.get(name, name)


def canonical_activity(value: Any) -> str:
    name = normalise_name(value)
    return ACTIVITY_ALIASES.get(name, name)


def build_session_key(
    venue_name: Any,
    activity_name: Any,
    session_date: date,
    start_time: time,
    end_time: time,
) -> tuple[str, str, date, time, time]:
    return (
        canonical_venue(venue_name),
        canonical_activity(activity_name),
        session_date,
        start_time,
        end_time,
    )


# ============================================================
# 5. EXCEL READING
# ============================================================

def read_sheet(
    path: Path,
    sheet_name: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(
                f"Missing sheet {sheet_name!r} in {path}"
            )

        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(iterator)]

        rows: list[dict[str, Any]] = []

        for values in iterator:
            if all(
                value is None or clean_text(value) == ""
                for value in values
            ):
                continue

            row = {
                headers[index]: (
                    values[index]
                    if index < len(values)
                    else None
                )
                for index in range(len(headers))
                if headers[index]
            }

            rows.append(row)

        return rows

    finally:
        workbook.close()


# ============================================================
# 6. SQL METADATA AND LOOKUPS
# ============================================================

def fetch_text_column_lengths(
    cursor: pyodbc.Cursor,
) -> dict[tuple[str, str], int | None]:
    cursor.execute(
        """
        SELECT
            t.name AS TableName,
            c.name AS ColumnName,
            ty.name AS DataType,
            c.max_length AS MaxLength
        FROM sys.tables t
        INNER JOIN sys.columns c
            ON c.object_id = t.object_id
        INNER JOIN sys.types ty
            ON ty.user_type_id = c.user_type_id
        WHERE t.name IN ('SessionAttendance', 'LiteMembers')
          AND ty.name IN ('nvarchar', 'varchar', 'nchar', 'char')
        ORDER BY t.name, c.column_id
        """
    )

    result: dict[tuple[str, str], int | None] = {}

    for row in cursor.fetchall():
        max_length = int(row.MaxLength)

        if max_length == -1:
            character_length = None
        elif row.DataType in ("nvarchar", "nchar"):
            character_length = max_length // 2
        else:
            character_length = max_length

        result[(str(row.TableName), str(row.ColumnName))] = (
            character_length
        )

    return result


def fetch_existing_sessions(
    cursor: pyodbc.Cursor,
) -> dict[
    tuple[str, str, date, time, time],
    list[dict[str, Any]],
]:
    cursor.execute(
        """
        SELECT
            s.SessionId,
            s.VenueName,
            s.ActivityName,
            s.SessionDate,
            s.StartTime,
            s.EndTime,
            s.IsCancelled,
            COUNT(a.AttendanceId) AS AttendanceRows
        FROM dbo.Sessions s
        LEFT JOIN dbo.SessionAttendance a
            ON a.SessionId = s.SessionId
        WHERE s.SessionDate >= ?
          AND s.SessionDate <= ?
        GROUP BY
            s.SessionId,
            s.VenueName,
            s.ActivityName,
            s.SessionDate,
            s.StartTime,
            s.EndTime,
            s.IsCancelled
        """,
        REPORT_START,
        REPORT_END,
    )

    grouped: dict[
        tuple[str, str, date, time, time],
        list[dict[str, Any]],
    ] = defaultdict(list)

    for row in cursor.fetchall():
        if (
            row.SessionDate is None
            or row.StartTime is None
            or row.EndTime is None
        ):
            continue

        key = build_session_key(
            row.VenueName,
            row.ActivityName,
            row.SessionDate,
            row.StartTime,
            row.EndTime,
        )

        grouped[key].append(
            {
                "session_id": int(row.SessionId),
                "attendance_rows": int(
                    row.AttendanceRows or 0
                ),
                "is_cancelled": int(
                    row.IsCancelled or 0
                ),
                "venue": clean_text(row.VenueName),
                "activity": clean_text(row.ActivityName),
            }
        )

    return grouped


def choose_preferred_session(
    candidates: list[dict[str, Any]],
) -> dict[str, Any]:
    """
    Where duplicate CRM sessions already exist, use:
      1. the record with the most attendance,
      2. a non-cancelled record,
      3. the lowest SessionId.
    """
    return sorted(
        candidates,
        key=lambda item: (
            -item["attendance_rows"],
            item["is_cancelled"],
            item["session_id"],
        ),
    )[0]


def fetch_full_participants(
    cursor: pyodbc.Cursor,
) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, list[dict[str, Any]]],
]:
    cursor.execute(
        """
        SELECT
            ParticipantID,
            SaheliCardNumber,
            FullName,
            MobileNumber,
            Postcode,
            DateOfBirth
        FROM dbo.Participants
        """
    )

    by_card: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in cursor.fetchall():
        item = {
            "participant_id": int(row.ParticipantID),
            "card": normalise_card(row.SaheliCardNumber),
            "name": clean_text(row.FullName),
            "phone": clean_text(row.MobileNumber),
            "postcode": clean_text(row.Postcode),
            "dob": row.DateOfBirth,
        }

        if item["card"]:
            by_card[item["card"]] = item

        name_key = normalise_name(item["name"])

        if name_key:
            by_name[name_key].append(item)

    return by_card, by_name


def fetch_lite_members(
    cursor: pyodbc.Cursor,
) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, list[dict[str, Any]]],
]:
    cursor.execute(
        """
        SELECT
            Id,
            MembershipId,
            FirstName,
            LastName,
            Phone,
            Postcode,
            DateOfBirth,
            EmergencyName,
            EmergencyPhone
        FROM dbo.LiteMembers
        """
    )

    by_membership: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in cursor.fetchall():
        full_name = clean_text(
            f"{clean_text(row.FirstName)} "
            f"{clean_text(row.LastName)}"
        )

        item = {
            "lite_member_id": row.Id,
            "membership_id": clean_text(row.MembershipId),
            "name": full_name,
            "phone": clean_text(row.Phone),
            "postcode": clean_text(row.Postcode),
            "dob": row.DateOfBirth,
            "emergency_name": clean_text(row.EmergencyName),
            "emergency_phone": clean_text(row.EmergencyPhone),
        }

        if item["membership_id"]:
            by_membership[item["membership_id"]] = item

        name_key = normalise_name(full_name)

        if name_key:
            by_name[name_key].append(item)

    return by_membership, by_name


# ============================================================
# 7. AUDIT
# ============================================================

def add_audit(
    audit: list[dict[str, Any]],
    entity: str,
    source_key: str,
    action: str,
    database_id: Any = "",
    details: str = "",
) -> None:
    audit.append(
        {
            "Entity": entity,
            "SourceKey": source_key,
            "Action": action,
            "DatabaseId": database_id,
            "Details": details,
        }
    )


def save_audit(
    audit: list[dict[str, Any]],
    mode: str,
) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    path = (
        Path.cwd()
        / f"calthorpe_migration_audit_{timestamp}.csv"
    )

    with path.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "Mode",
                "Entity",
                "SourceKey",
                "Action",
                "DatabaseId",
                "Details",
            ],
        )

        writer.writeheader()

        for row in audit:
            writer.writerow({"Mode": mode, **row})

    return path


# ============================================================
# 8. INSERT FUNCTIONS
# ============================================================

def insert_session(
    cursor: pyodbc.Cursor,
    row: dict[str, Any],
) -> int:
    session_date = parse_date(row["SessionDate"])
    start_time = parse_time(row["StartTime"])
    end_time = parse_time(row["EndTime"])

    if end_time <= start_time:
        raise ValueError(
            "Session EndTime must be later than StartTime: "
            f"{row.get('ImportSessionKey')}"
        )

    cursor.execute(
        """
        INSERT INTO dbo.Sessions
        (
            Frequency,
            Category,
            SubCategory,
            ActivityCategory,
            VenueName,
            ActivityName,
            Notes,
            IsRecurringWeekly,
            DayOfWeek,
            SessionDate,
            ArrivalTime,
            StartTime,
            EndTime,
            Capacity,
            IsBookingRequired,
            IsCancelled,
            AssignedStaffId,
            RecurringSeriesId,
            SessionProviderId
        )
        OUTPUT INSERTED.SessionId
        VALUES
        (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )
        """,
        clean_text(row.get("Frequency")) or "WEEKLY",
        clean_text(row.get("Category")) or "Fitness",
        clean_text(row.get("SubCategory")) or None,
        clean_text(row.get("ActivityCategory")) or "Fitness",
        clean_text(row.get("VenueName")),
        clean_text(row.get("ActivityName")),
        clean_text(row.get("Notes"))[:1000] or None,
        0,
        None,
        session_date,
        None,
        start_time,
        end_time,
        to_int(row.get("Capacity")),
        to_int(row.get("IsBookingRequired"), 0),
        to_int(row.get("IsCancelled"), 0),
        to_int(row.get("AssignedStaffId")),
        None,
        to_int(row.get("SessionProviderId")),
    )

    return int(cursor.fetchone()[0])


def create_lite_member(
    cursor: pyodbc.Cursor,
    row: dict[str, Any],
    column_lengths: dict[tuple[str, str], int | None],
) -> dict[str, Any]:
    membership_id = safe_database_text(
        row.get("ProposedMembershipId"),
        column_lengths.get(("LiteMembers", "MembershipId")),
    )

    if not membership_id:
        raise ValueError(
            "Missing ProposedMembershipId for "
            f"{row.get('SourceParticipantKey')}"
        )

    first_name = safe_database_text(
        row.get("FirstName"),
        column_lengths.get(("LiteMembers", "FirstName")),
    )

    last_name = safe_database_text(
        row.get("LastName") or "(Not provided)",
        column_lengths.get(("LiteMembers", "LastName")),
    )

    if not first_name:
        raise ValueError(
            "Cannot create a LiteMember without FirstName: "
            f"{row.get('SourceParticipantKey')}"
        )

    lite_id = uuid.uuid5(
        uuid.NAMESPACE_URL,
        f"saheli-calthorpe-2025-26:{membership_id}",
    )

    emergency_phone = normalise_phone_for_database(
        row.get("EmergencyPhone")
    )

    cursor.execute(
        """
        INSERT INTO dbo.LiteMembers
        (
            Id,
            MembershipId,
            FirstName,
            LastName,
            DateOfBirth,
            Phone,
            Email,
            Address,
            Postcode,
            EmergencyName,
            EmergencyPhone,
            EmergencyRelation,
            HealthConditions,
            Gender,
            Ethnicity,
            CreatedByUserId
        )
        VALUES
        (
            ?, ?, ?, ?, NULL, NULL, NULL, NULL, NULL, ?, ?, NULL, NULL,
            NULL, NULL, ?
        )
        """,
        str(lite_id),
        membership_id,
        first_name,
        last_name,
        safe_database_text(
            row.get("EmergencyName"),
            column_lengths.get(("LiteMembers", "EmergencyName")),
        ),
        safe_database_text(
            emergency_phone,
            column_lengths.get(("LiteMembers", "EmergencyPhone")),
        ),
        CREATED_BY_USER_ID,
    )

    return {
        "kind": "LITE",
        "lite_member_id": str(lite_id),
        "membership_id": membership_id,
        "name": clean_text(f"{first_name} {last_name}"),
        "phone": "",
    }


def insert_attendance(
    cursor: pyodbc.Cursor,
    row: dict[str, Any],
    session_id: int,
    member: dict[str, Any],
    column_lengths: dict[tuple[str, str], int | None],
    audit: list[dict[str, Any]],
) -> int:
    session_date = parse_date(row["SessionDate"])
    start_time = parse_time(row["SessionStartTime"])
    end_time = parse_time(row["SessionEndTime"])

    source_key = clean_text(row["SourceParticipantKey"])

    if member["kind"] == "FULL":
        participant_id = member["participant_id"]
        lite_member_id = None
        card_number = member["card"]
        display_id = member["card"]
    else:
        participant_id = None
        lite_member_id = member["lite_member_id"]
        card_number = None
        display_id = member["membership_id"]

    raw_phone = clean_text(member.get("phone"))
    safe_phone = normalise_phone_for_database(raw_phone)

    if raw_phone and safe_phone is None:
        add_audit(
            audit,
            "ATTENDANCE",
            source_key,
            "INVALID_PHONE_SKIPPED",
            "",
            f"Field=Phone; raw value={raw_phone!r}",
        )

    raw_emergency_phone = clean_text(
        row.get("EmergencyPhone")
    )
    safe_emergency_phone = normalise_phone_for_database(
        raw_emergency_phone
    )

    if (
        raw_emergency_phone
        and safe_emergency_phone is None
    ):
        add_audit(
            audit,
            "ATTENDANCE",
            source_key,
            "INVALID_PHONE_SKIPPED",
            "",
            (
                "Field=EmergencyPhone; "
                f"raw value={raw_emergency_phone!r}"
            ),
        )

    notes = (
        "Historical Calthorpe register import"
        f" | workbook={clean_text(row.get('SourceWorkbook'))}"
        f" | sheet={clean_text(row.get('SourceSheet'))}"
        f" | row={clean_text(row.get('SourceRow'))}"
        f" | sourceActivity={clean_text(row.get('SourceActivityLabel'))}"
        f" | sourceTime={clean_text(row.get('SourceTimeText'))}"
        f" | key={clean_text(row.get('ImportAttendanceKey'))}"
    )

    cursor.execute(
        """
        INSERT INTO dbo.SessionAttendance
        (
            SessionId,
            ParticipantId,
            SessionName,
            SessionDay,
            SessionDate,
            SessionMonth,
            SessionStartTime,
            SessionEndTime,
            SaheliCardNumber,
            RiskStratification,
            Attended,
            CheckInTime,
            CheckOutTime,
            Notes,
            AttendanceMemberKind,
            LiteMemberId,
            MemberDisplayId,
            MemberName,
            Phone,
            EmergencyName,
            EmergencyPhone
        )
        OUTPUT INSERTED.AttendanceId
        VALUES
        (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )
        """,
        session_id,
        participant_id,
        safe_database_text(
            row.get("ActivityName"),
            column_lengths.get(
                ("SessionAttendance", "SessionName")
            ),
        ),
        safe_database_text(
            row.get("SessionDay")
            or session_date.strftime("%A"),
            column_lengths.get(
                ("SessionAttendance", "SessionDay")
            ),
        ),
        session_date,
        safe_database_text(
            row.get("SessionMonth")
            or session_date.strftime("%B"),
            column_lengths.get(
                ("SessionAttendance", "SessionMonth")
            ),
        ),
        start_time,
        end_time,
        safe_database_text(
            card_number,
            column_lengths.get(
                ("SessionAttendance", "SaheliCardNumber")
            ),
        ),
        safe_database_text(
            row.get("RiskStratification"),
            column_lengths.get(
                ("SessionAttendance", "RiskStratification")
            ),
        ),
        1,
        start_time,
        end_time,
        safe_database_text(
            notes,
            column_lengths.get(
                ("SessionAttendance", "Notes")
            ),
        ),
        safe_database_text(
            member["kind"],
            column_lengths.get(
                ("SessionAttendance", "AttendanceMemberKind")
            ),
        ),
        lite_member_id,
        safe_database_text(
            display_id,
            column_lengths.get(
                ("SessionAttendance", "MemberDisplayId")
            ),
        ),
        safe_database_text(
            member.get("name")
            or row.get("MemberName")
            or display_id,
            column_lengths.get(
                ("SessionAttendance", "MemberName")
            ),
        ),
        safe_database_text(
            safe_phone,
            column_lengths.get(
                ("SessionAttendance", "Phone")
            ),
        ),
        safe_database_text(
            row.get("EmergencyName"),
            column_lengths.get(
                ("SessionAttendance", "EmergencyName")
            ),
        ),
        safe_database_text(
            safe_emergency_phone,
            column_lengths.get(
                ("SessionAttendance", "EmergencyPhone")
            ),
        ),
    )

    return int(cursor.fetchone()[0])


# ============================================================
# 9. MAIN MIGRATION
# ============================================================

def main() -> int:
    base_folder = Path(__file__).resolve().parent

    sessions_path = base_folder / SESSIONS_FILE
    participants_path = base_folder / PARTICIPANTS_FILE

    if not sessions_path.exists():
        raise FileNotFoundError(sessions_path)

    if not participants_path.exists():
        raise FileNotFoundError(participants_path)

    session_rows = read_sheet(
        sessions_path,
        "Sessions",
    )
    participant_rows = read_sheet(
        participants_path,
        "Participants",
    )
    attendance_rows = read_sheet(
        participants_path,
        "Attendance",
    )

    actual_totals = (
        len(session_rows),
        len(participant_rows),
        len(attendance_rows),
    )

    expected_totals = (
        EXPECTED_SESSIONS,
        EXPECTED_PARTICIPANTS,
        EXPECTED_ATTENDANCE,
    )

    if actual_totals != expected_totals:
        raise RuntimeError(
            "Staging totals do not match. "
            f"Expected {expected_totals}, "
            f"received {actual_totals}."
        )

    # Refuse to run if the staging file contains duplicate
    # participant/session combinations.
    source_attendance_keys = [
        (
            clean_text(row["ImportSessionKey"]),
            clean_text(row["SourceParticipantKey"]),
        )
        for row in attendance_rows
    ]

    duplicate_source_keys = [
        key
        for key, count in Counter(source_attendance_keys).items()
        if count > 1
    ]

    if duplicate_source_keys:
        raise RuntimeError(
            "The Attendance sheet contains duplicate "
            "session/participant records. First examples: "
            f"{duplicate_source_keys[:10]}"
        )

    audit: list[dict[str, Any]] = []
    summary: Counter[str] = Counter()

    connection = pyodbc.connect(
        CONNECTION_STRING,
        autocommit=False,
    )

    cursor = connection.cursor()
    cursor.execute("SET XACT_ABORT ON;")

    try:
        column_lengths = fetch_text_column_lengths(
            cursor
        )

        print()
        print("Runtime SQL text column lengths (characters)")
        print("-" * 68)

        for (
            table_name,
            column_name,
        ), length in sorted(column_lengths.items()):
            length_text = "MAX" if length is None else str(length)
            print(
                f"dbo.{table_name}.{column_name}: "
                f"{length_text}"
            )

        participants_by_card, full_by_name = (
            fetch_full_participants(cursor)
        )
        lite_by_membership, lite_by_name = (
            fetch_lite_members(cursor)
        )

        # ----------------------------------------------------
        # A. PREFLIGHT PARTICIPANT CHECKS
        # ----------------------------------------------------
        print()
        print("=" * 68)
        print("PREFLIGHT VALIDATION")
        print("=" * 68)

        preflight: Counter[str] = Counter()

        for row in participant_rows:
            source_key = clean_text(
                row["SourceParticipantKey"]
            )
            kind = clean_text(
                row["ExpectedMemberKind"]
            ).upper()

            if kind == "FULL":
                card = normalise_card(
                    row.get("SaheliCardNumber")
                )

                if card not in participants_by_card:
                    preflight["PREFLIGHT_UNMATCHED_CARD"] += 1
                    add_audit(
                        audit,
                        "PREFLIGHT",
                        source_key,
                        "PREFLIGHT_UNMATCHED_CARD",
                        "",
                        f"Card {card!r} is not in dbo.Participants.",
                    )

            elif not clean_text(row.get("SourceName")):
                preflight["PREFLIGHT_BLANK_LITE_NAME"] += 1
                add_audit(
                    audit,
                    "PREFLIGHT",
                    source_key,
                    "PREFLIGHT_BLANK_LITE_NAME",
                    "",
                    "A LiteMember cannot be created without a name.",
                )

            emergency_phone = clean_text(
                row.get("EmergencyPhone")
            )

            if (
                emergency_phone
                and normalise_phone_for_database(
                    emergency_phone
                ) is None
            ):
                preflight["INVALID_PHONE_SKIPPED"] += 1
                add_audit(
                    audit,
                    "PREFLIGHT",
                    source_key,
                    "INVALID_PHONE_SKIPPED",
                    "",
                    (
                        "Field=EmergencyPhone; "
                        f"raw value={emergency_phone!r}"
                    ),
                )

        if preflight:
            for action, count in sorted(preflight.items()):
                print(f"{action:38s} {count:>6}")
                summary[action] += count
        else:
            print("No preflight problems detected.")

        # ----------------------------------------------------
        # B. MATCH OR CREATE SESSIONS
        # ----------------------------------------------------
        existing_sessions = fetch_existing_sessions(
            cursor
        )

        resolved_session_ids: dict[str, int] = {}

        for row in session_rows:
            import_key = clean_text(
                row["ImportSessionKey"]
            )

            exact_key = build_session_key(
                row["VenueName"],
                row["ActivityName"],
                parse_date(row["SessionDate"]),
                parse_time(row["StartTime"]),
                parse_time(row["EndTime"]),
            )

            candidates = existing_sessions.get(
                exact_key,
                [],
            )

            if candidates:
                preferred = choose_preferred_session(
                    candidates
                )
                session_id = preferred["session_id"]
                action = "EXISTING_SESSION"
                details = (
                    f"Matched {len(candidates)} CRM session(s). "
                    f"Selected SessionId {session_id}; "
                    f"existing attendance="
                    f"{preferred['attendance_rows']}."
                )
            else:
                session_id = insert_session(
                    cursor,
                    row,
                )
                action = "NEW_SESSION"
                details = (
                    "Inserted inside the current SQL transaction."
                )

                existing_sessions[exact_key] = [
                    {
                        "session_id": session_id,
                        "attendance_rows": 0,
                        "is_cancelled": to_int(
                            row.get("IsCancelled"), 0
                        ),
                    }
                ]

            resolved_session_ids[
                import_key
            ] = session_id

            summary[action] += 1

            add_audit(
                audit,
                "SESSION",
                import_key,
                action,
                session_id,
                details,
            )

        # ----------------------------------------------------
        # C. MATCH OR CREATE PARTICIPANTS
        # ----------------------------------------------------
        resolved_members: dict[
            str,
            dict[str, Any],
        ] = {}

        for row in participant_rows:
            source_key = clean_text(
                row["SourceParticipantKey"]
            )
            expected_kind = clean_text(
                row["ExpectedMemberKind"]
            ).upper()
            card = normalise_card(
                row.get("SaheliCardNumber")
            )
            source_name_key = normalise_name(
                row.get("NormalisedName")
                or row.get("SourceName")
            )

            member: dict[str, Any] | None = None
            action = ""
            details = ""

            if expected_kind == "FULL":
                full = participants_by_card.get(card)

                if full:
                    member = {
                        "kind": "FULL",
                        **full,
                    }
                    action = "MATCHED_FULL"
                else:
                    action = "UNMATCHED_CARD"
                    details = (
                        f"Card {card!r} was not found in "
                        "dbo.Participants. No FULL participant "
                        "was created."
                    )

            else:
                membership_id = clean_text(
                    row.get("ProposedMembershipId")
                )

                existing_lite = lite_by_membership.get(
                    membership_id
                )

                if existing_lite:
                    member = {
                        "kind": "LITE",
                        **existing_lite,
                    }
                    action = "MATCHED_LITE_MEMBERSHIP_ID"

                else:
                    lite_candidates = lite_by_name.get(
                        source_name_key,
                        [],
                    )
                    full_candidates = full_by_name.get(
                        source_name_key,
                        [],
                    )

                    if len(lite_candidates) == 1:
                        member = {
                            "kind": "LITE",
                            **lite_candidates[0],
                        }
                        action = "MATCHED_LITE_NAME"

                    elif len(lite_candidates) > 1:
                        action = "AMBIGUOUS_LITE_NAME"
                        details = (
                            f"{len(lite_candidates)} LiteMembers "
                            "have the same normalised name."
                        )

                    elif len(full_candidates) == 1:
                        action = "REVIEW_FULL_NAME_MATCH"
                        details = (
                            "One FULL participant has the same "
                            "name, but the source has no card number."
                        )

                    elif len(full_candidates) > 1:
                        action = "AMBIGUOUS_FULL_NAME"
                        details = (
                            f"{len(full_candidates)} FULL "
                            "participants have the same name."
                        )

                    elif CREATE_MISSING_LITE_MEMBERS:
                        member = create_lite_member(
                            cursor,
                            row,
                            column_lengths,
                        )
                        action = "CREATED_LITE"

                        lite_by_membership[
                            member["membership_id"]
                        ] = member

                        lite_by_name[
                            source_name_key
                        ].append(member)

                    else:
                        action = "NEW_LITE_REQUIRED"
                        details = (
                            "No existing LiteMember was found. "
                            "Set CREATE_MISSING_LITE_MEMBERS = True "
                            "after reviewing the first preview."
                        )

            if member:
                resolved_members[
                    source_key
                ] = member

                database_id = (
                    member.get("participant_id")
                    if member["kind"] == "FULL"
                    else member.get("lite_member_id")
                )
            else:
                database_id = ""

            summary[action] += 1

            add_audit(
                audit,
                "PARTICIPANT",
                source_key,
                action,
                database_id,
                details,
            )

        # ----------------------------------------------------
        # D. LOAD EXISTING ATTENDANCE
        # ----------------------------------------------------
        session_ids = sorted(
            set(resolved_session_ids.values())
        )

        existing_attendance: dict[
            tuple[int, str, str],
            int,
        ] = {}

        if session_ids:
            placeholders = ",".join(
                "?" for _ in session_ids
            )

            cursor.execute(
                f"""
                SELECT
                    AttendanceId,
                    SessionId,
                    ParticipantId,
                    LiteMemberId
                FROM dbo.SessionAttendance
                WHERE SessionId IN ({placeholders})
                """,
                *session_ids,
            )

            for row in cursor.fetchall():
                if row.ParticipantId is not None:
                    kind = "FULL"
                    member_id = str(int(row.ParticipantId))

                elif row.LiteMemberId is not None:
                    kind = "LITE"
                    member_id = str(row.LiteMemberId).lower()

                else:
                    continue

                existing_attendance[
                    (
                        int(row.SessionId),
                        kind,
                        member_id,
                    )
                ] = int(row.AttendanceId)

        # ----------------------------------------------------
        # E. INSERT ONLY MISSING ATTENDANCE
        # ----------------------------------------------------
        for row in attendance_rows:
            attendance_key = clean_text(
                row["ImportAttendanceKey"]
            )
            import_session_key = clean_text(
                row["ImportSessionKey"]
            )
            source_participant_key = clean_text(
                row["SourceParticipantKey"]
            )

            session_id = resolved_session_ids.get(
                import_session_key
            )
            member = resolved_members.get(
                source_participant_key
            )

            if not session_id:
                action = "UNMATCHED_SESSION"
                summary[action] += 1

                add_audit(
                    audit,
                    "ATTENDANCE",
                    attendance_key,
                    action,
                    "",
                    (
                        "No session was resolved for "
                        f"{import_session_key}."
                    ),
                )
                continue

            if not member:
                action = "UNMATCHED_PARTICIPANT"
                summary[action] += 1

                add_audit(
                    audit,
                    "ATTENDANCE",
                    attendance_key,
                    action,
                    "",
                    (
                        "No safe participant match was resolved for "
                        f"{source_participant_key}."
                    ),
                )
                continue

            if member["kind"] == "FULL":
                member_id = str(
                    member["participant_id"]
                )
            else:
                member_id = str(
                    member["lite_member_id"]
                ).lower()

            duplicate_key = (
                session_id,
                member["kind"],
                member_id,
            )

            existing_attendance_id = (
                existing_attendance.get(
                    duplicate_key
                )
            )

            if existing_attendance_id:
                action = "ALREADY_IN_CRM"
                summary[action] += 1

                add_audit(
                    audit,
                    "ATTENDANCE",
                    attendance_key,
                    action,
                    existing_attendance_id,
                    (
                        "Skipped because this participant already "
                        "has attendance for the resolved session."
                    ),
                )
                continue

            attendance_id = insert_attendance(
                cursor,
                row,
                session_id,
                member,
                column_lengths,
                audit,
            )

            existing_attendance[
                duplicate_key
            ] = attendance_id

            action = "NEW_ATTENDANCE"
            summary[action] += 1

            add_audit(
                audit,
                "ATTENDANCE",
                attendance_key,
                action,
                attendance_id,
                (
                    "Inserted inside the current SQL transaction."
                ),
            )

        # ----------------------------------------------------
        # F. COMMIT OR ROLLBACK
        # ----------------------------------------------------
        if COMMIT_CHANGES:
            connection.commit()
            mode = "COMMITTED"
        else:
            connection.rollback()
            mode = "PREVIEW_ROLLED_BACK"

    except Exception:
        connection.rollback()
        raise

    finally:
        cursor.close()
        connection.close()

    audit_path = save_audit(
        audit,
        mode,
    )

    print()
    print("=" * 68)
    print("CALTHORPE 2025/26 CRM MIGRATION")
    print("=" * 68)
    print("Mode:", mode)
    print("Sessions file:", sessions_path)
    print("Participants file:", participants_path)
    print("Audit file:", audit_path)
    print()
    print("Action summary")
    print("-" * 68)

    for action, count in sorted(
        summary.items()
    ):
        print(f"{action:38s} {count:>6}")

    print()

    if not COMMIT_CHANGES:
        print("No database changes were saved.")
        print(
            "Review the audit CSV before setting "
            "COMMIT_CHANGES = True."
        )

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print()
        print(
            "MIGRATION FAILED:",
            error,
            file=sys.stderr,
        )
        raise
