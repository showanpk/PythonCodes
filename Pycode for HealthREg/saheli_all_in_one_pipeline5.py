# ============================================================
# saheli_pipeline_flat_header.py
# IDENTICAL to original pipeline EXCEPT:
#   - APPLY_GROUPED_HEADER = False  (single row header)
#   - Column names match Master "Full Register" exactly:
#       Personal: "Full Name:", "Date of Birth:", "Address:", etc. (no leading space)
#       Assessment: "1st Assessment  Weight (KG):" (double space separator)
# ============================================================

from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
REG_FILE    = r"C:\Users\shonk\Downloads\Main Registration Form(1-1143).xlsx"
HEALTH_FILE = r"C:\Users\shonk\Downloads\Saheli Hub Health Assessment(1-1478).xlsx"

REG_OUTPUT_FILE    = r"C:\Users\shonk\source\PythonCodes\New folder\Registrations_Cleaned.xlsx"
HEALTH_OUTPUT_FILE = r"C:\Users\shonk\source\PythonCodes\New folder\Healthassessments_Prepared.xlsx"
FINAL_OUTPUT_FILE  = r"C:\Users\shonk\source\PythonCodes\New folder\Saheli_Master_Wide_Output.xlsx"

REG_SHEET_NAME    = None
HEALTH_SHEET_NAME = None
MAX_ASSESSMENTS   = 9   # master has up to 9th Assessment

# ── CHANGED: no grouped header ────────────────────────────────────────────────
APPLY_GROUPED_HEADER = False


# =========================
# HELPERS  (unchanged)
# =========================
def normalize_header(h) -> str:
    if h is None: return ""
    s = str(h)
    s = s.replace("\r","").replace("\n","")
    s = s.replace('"',"").replace("\u201c","").replace("\u201d","")
    s = s.strip().lower()
    for ch in [" ",":","/"," ","?","(",")",".",",","-","&","'","\u2019"]:
        s = s.replace(ch,"")
    return s

def keep_digits_only(v):
    if pd.isna(v): return pd.NA
    if isinstance(v,(int,float)):
        try:
            fv=float(v)
            if fv.is_integer(): return str(int(fv))
            digits=re.sub(r"\D+","",format(fv,"f"))
            return digits if digits else pd.NA
        except: pass
    s=str(v).strip()
    if re.fullmatch(r"\d+\.0+",s): return s.split(".")[0]
    digits=re.sub(r"\D+","",s)
    return digits if digits else pd.NA

def safe_saheli_key_series(series):
    s=series.apply(keep_digits_only)
    s=pd.to_numeric(s,errors="coerce").astype("Int64")
    return s.astype("string")

def read_excel_flexible(path,sheet_name=None):
    p=Path(path)
    if not p.exists(): raise FileNotFoundError(f"File not found: {p}")
    return pd.read_excel(p) if sheet_name is None else pd.read_excel(p,sheet_name=sheet_name)

def write_excel(df,out_path):
    out=Path(out_path)
    out.parent.mkdir(parents=True,exist_ok=True)
    with pd.ExcelWriter(out,engine="openpyxl",
                        datetime_format="dd/mm/yyyy",date_format="dd/mm/yyyy") as writer:
        df.to_excel(writer,index=False)
    return out

def build_normalized_col_map(df):
    m={}
    for c in df.columns:
        k=normalize_header(c)
        m.setdefault(k,[]).append(c)
    return m

def pick_col(norm_map,*candidates,occurrence=1):
    for cand in candidates:
        k=normalize_header(cand)
        if k in norm_map and len(norm_map[k])>=occurrence:
            return norm_map[k][occurrence-1]
    return None

def parse_date(series):
    return pd.to_datetime(series,errors="coerce",dayfirst=True).dt.date

def compute_age(dob_series,ref_series=None):
    dob=pd.to_datetime(dob_series,errors="coerce",dayfirst=True)
    if ref_series is None:
        ref=pd.Series([pd.Timestamp.today().normalize()]*len(dob),index=dob.index)
    else:
        ref=pd.to_datetime(ref_series,errors="coerce",dayfirst=True)
    age=ref.dt.year-dob.dt.year
    before=((ref.dt.month<dob.dt.month)|
            ((ref.dt.month==dob.dt.month)&(ref.dt.day<dob.dt.day)))
    return (age-before.astype(int)).astype("Int64")

def ordinal(n):
    suf="th" if 10<=n%100<=20 else {1:"st",2:"nd",3:"rd"}.get(n%10,"th")
    return f"{n}{suf}"

def clean_subheader_label(label):
    s=str(label).strip()
    if s=="WEMBS":          return "WEMWBS"
    if s=="SOCIAL ISOLATION": return "SOCIAL ISOLATION"
    return s

def build_health_field_order_map():
    m={}
    for i,f in enumerate(HEALTH_FIELDS):
        d=clean_subheader_label(f)
        if d not in m: m[d]=i
    return m


# =========================
# ── CHANGED: Output labels match Master exactly (no leading spaces) ──────────
# =========================
REG_OUTPUT_LABELS = [
    "No of assessment completed",
    "Registration Date",
    "Saheli Card Number",
    "Full Name:",
    "Date of Birth:",
    "AGE",
    "Address:",
    "Postcode:",
    "Email:",
    "Mobile/Home No:",
    "Emergency Contact Name:",
    "Emergency No:",
    "Emergency Relation To You:",
    "Gender:",
    "Is your gender the same as assigned at birth?",
    "Health Conditions/Disability:",
    "Ethnicity:",
    "Preferred spoken language:",
    "Religion:",
    "Relationship status:",
    "Caring responsibilities:",
    "Living alone:",
    "Sexuality:",
    "Occupation:",
    "Referral reason",
    "How heard about Saheli Hub?",
    "GP Surgery Name:",
    "Consent to store information:",
    "Health declaration:",
    "Permission to be added to Saheli WhatsApp group?",
    "Permission to be in photos and videos? (Media consent)",
    "Notes:",
    "Staff Member:",
    "Site:",
]

# Health field display labels — must match Master subfield names exactly
HEALTH_FIELDS = [
    "Weight (KG):",
    "Height (CM):",
    "BMI:",
    "BMI Results:",
    "Waist (CM):",
    "Hip (CM):",
    "Waist to Hip Ratio (CM):",
    "Body Fat Percentage Result:",
    "Body Fat Percentage Score:",
    "Visceral Fat Level Result:",
    "Visceral Fat Level Score:",
    "Skeletal Muscle Percentage:",
    "Skeletal Muscle Score:",
    "Resting Metabolism:",
    "Do You Have Any Health Condition?",
    "When did you last measure your blood pressure?",
    "Have you recorded your blood pressure measurement and registered it with a GP or Pharmacist? (Yes/No/Not sure)",
    "What is a healthy blood pressure for an adult?",
    "Why is a high blood pressure dangerous?",
    "How can you help reduce your blood pressure?",
    "Blood Pressure (Systolic/Diastolic):",
    "Blood Pressure Level:",
    "Do You Have a Heart Condition?",
    "Heart Rate (BPM):",
    "Atrial Fibrillation Result:",
    "Heart Age:",
    "Did Your Doctor Advise You Not to Exercise?",
    "Do You Feel Pain in Chest at Rest/During Activity?",
    "Do You Have Shortness of Breath?",
    "Do You Have Diabetes?",
    "Diabetes Risk:",
    "Glucose Level ( mg/dL):",
    "HbA1c:",
    "Do You Take Sugary Drinks, Including Chai?",
    "Do You Have High Cholesterol? (Total/HDL)",
    "Do You Experience The Following Health Issues?",
    "Do You Have a Bone / joint Condition?",
    "Do You Take Any Prescribed Medication?",
    "Referred to doctor for any concerning results?",
    "Risk Stratification Score",
    "Comments:",
    "How well do you manage your health/condition(s)? (Rating out of 10)",
    "In the past week, on how many days have you done a total of 30 mins or more of physical activity, which was enough to raise your breathing rate?",
    "Physical Activity Level:",
    "Comments:PA",
    "I've been feeling optimistic about the future",
    "I've been feeling useful",
    "I've been feeling relaxed",
    "I've been feeling interested in other people",
    "I've had energy to spare",
    "I've been dealing with problems well",
    "I've been thinking clearly",
    "I've been feeling good about myself",
    "I've been feeling close to other people",
    "I've been feeling confident",
    "I've been able to make up my own mind about things",
    "I've been feeling loved",
    "I've been interested in new things",
    "I've been feeling cheerful",
    "WEMWBS",       # displayed as WEMWBS (computed from WEMBS)
    "Comments:2",
    "Nourishment: Rate the quality of the food you put into your body on a daily basis",
    "Movement: Rate how often and for how long you move your body on a daily basis",
    "Connectedness: Rate how well you stay connected with family, friends and your higher power",
    "Sleep: Rate the quality of your sleep",
    "Happy self: Rate how often and for how long you perform positive practices (gratitude, virtue awareness, meditation, prayer, etc.)",
    "Resilience: Rate how well you are able to manage stress in your life",
    "Green and Blue: Rate how often and how long you spend in nature or outdoors",
    "Screen time: Rate how happy you are with your current amount of screen time",
    "Substance use: Rate how comfortable you are with any current substance use (smoking, alcohol, drugs)",
    "Purpose: Rate how well you feel you are fulfilling your passion, purpose or vocation in life",
    "Comments:3",
    "How often do you feel that you lack companionship?",
    "How often do you feel left out?",
    "How often do you feel isolated from others?",
    "SOCIAL ISOLATION",
    "Comments:4",
    "How confident are you to join activities?",
    "How many hobbies and passions do you have?",
    "How involved you feel in your community?",
    "How much you know about local support/services?",
    "What are your aims & goals?",
    "Comments:5",
    "What reasons stop you from joining activities?",
    "Comments:6",
    "What are your preferred activities?",
    "Comments:7",
    "Date of next review appointment:",
]


# =========================
# STEP 1 & 2: unchanged
# =========================
def prepare_registration_df(df_reg):
    reg_map=build_normalized_col_map(df_reg)
    col_saheli=pick_col(reg_map,"Saheli Card No","SaheliCardNo","Saheli Card Number")
    if not col_saheli: raise KeyError("Registration Saheli Card No column not found.")
    df=df_reg.copy()
    df[col_saheli]=df[col_saheli].apply(keep_digits_only)
    df[col_saheli]=pd.to_numeric(df[col_saheli],errors="coerce").astype("Int64")
    return df

def prepare_health_df(df_health):
    health_map=build_normalized_col_map(df_health)
    col_completion=pick_col(health_map,"Completion time")
    col_saheli=pick_col(health_map,"Saheli Card No","SaheliCardNo","Saheli Card Number")
    if not col_completion: raise KeyError("Health 'Completion time' column not found.")
    if not col_saheli:     raise KeyError("Health 'Saheli Card No' column not found.")
    df=df_health.copy()
    df[col_saheli]=df[col_saheli].apply(keep_digits_only)
    df[col_saheli]=pd.to_numeric(df[col_saheli],errors="coerce").astype("Int64")
    df[col_completion]=pd.to_datetime(df[col_completion],errors="coerce",dayfirst=True).dt.date
    cols=list(df.columns)
    cols.remove(col_saheli)
    cols.insert(cols.index(col_completion)+1,col_saheli)
    df=df[cols].copy()
    health_map2=build_normalized_col_map(df)
    col_completion=pick_col(health_map2,"Completion time")
    col_saheli=pick_col(health_map2,"Saheli Card No","SaheliCardNo","Saheli Card Number")
    df=df.sort_values(by=[col_saheli,col_completion],ascending=[True,True],
                      na_position="last",kind="mergesort").reset_index(drop=True)
    assess_num=(df.groupby(col_saheli,dropna=False).cumcount()+1)
    assess_num=assess_num.where(df[col_saheli].notna(),pd.NA).astype("Int64")
    pos=df.columns.get_loc(col_saheli)
    df.insert(pos+1,"AssessmentNumber",assess_num)
    return df


# =========================
# STEP 3: WIDE — flat single-row header matching Master exactly
# =========================
def create_final_wide_df(df_reg_clean,df_health_prepared):
    reg_map=build_normalized_col_map(df_reg_clean)
    health_map=build_normalized_col_map(df_health_prepared)

    reg_saheli_col   =pick_col(reg_map,"Saheli Card No","SaheliCardNo","Saheli Card Number")
    health_saheli_col=pick_col(health_map,"Saheli Card No","SaheliCardNo","Saheli Card Number")
    completion_col   =pick_col(health_map,"Completion time")
    assess_col       =pick_col(health_map,"AssessmentNumber")

    if not reg_saheli_col:    raise KeyError("Reg Saheli col not found.")
    if not health_saheli_col: raise KeyError("Health Saheli col not found.")
    if not completion_col:    raise KeyError("Completion time not found.")
    if not assess_col:        raise KeyError("AssessmentNumber not found.")

    df_reg=df_reg_clean.copy()
    df_health=df_health_prepared.copy()

    df_reg["_SaheliKey"]   =safe_saheli_key_series(df_reg[reg_saheli_col])
    df_health["_SaheliKey"]=safe_saheli_key_series(df_health[health_saheli_col])
    df_health["_CompletionDate"]=parse_date(df_health[completion_col])
    df_health[assess_col]=pd.to_numeric(df_health[assess_col],errors="coerce").astype("Int64")

    # Compute WEMWBS
    wem_items=[pick_col(health_map,f) for f in [
        "I've been feeling optimistic about the future","I've been feeling useful",
        "I've been feeling relaxed","I've been feeling interested in other people",
        "I've had energy to spare","I've been dealing with problems well",
        "I've been thinking clearly","I've been feeling good about myself",
        "I've been feeling close to other people","I've been feeling confident",
        "I've been able to make up my own mind about things","I've been feeling loved",
        "I've been interested in new things","I've been feeling cheerful",
    ]]
    wem_items=[c for c in wem_items if c]
    for c in wem_items: df_health[c]=pd.to_numeric(df_health[c],errors="coerce")
    df_health["WEMBS"]=df_health[wem_items].sum(axis=1,min_count=1) if wem_items else pd.NA

    soc_items=[pick_col(health_map,f) for f in [
        "How often do you feel that you lack companionship?",
        "How often do you feel left out?",
        "How often do you feel isolated from others?",
    ]]
    soc_items=[c for c in soc_items if c]
    for c in soc_items: df_health[c]=pd.to_numeric(df_health[c],errors="coerce")
    df_health["SOCIAL ISOLATION"]=df_health[soc_items].sum(axis=1,min_count=1) if soc_items else pd.NA

    # Registration base
    reg_date_col=(pick_col(reg_map,"Date") or pick_col(reg_map,"Completion time")
                  or pick_col(reg_map,"Start time"))
    dob_col=pick_col(reg_map,"Date of Birth")
    age_col=pick_col(reg_map,"Age")

    if reg_date_col:
        df_reg["_RegDateParsed"]=parse_date(df_reg[reg_date_col])
    else:
        df_reg["_RegDateParsed"]=pd.NaT

    df_reg_base =df_reg[df_reg["_SaheliKey"].notna()].copy()
    df_reg_base =df_reg_base.sort_values(["_SaheliKey","_RegDateParsed"],
                                          ascending=[True,True],na_position="last",kind="mergesort")
    df_reg_first=df_reg_base.drop_duplicates(subset=["_SaheliKey"],keep="first").copy()
    reg_first_map=build_normalized_col_map(df_reg_first)
    def rp(*c,occurrence=1): return pick_col(reg_first_map,*c,occurrence=occurrence)

    # ── CHANGED: output labels match Master exactly (no leading spaces) ───────
    reg_out=pd.DataFrame({"_SaheliKey":df_reg_first["_SaheliKey"].astype("string")})
    reg_out["Registration Date"]=df_reg_first["_RegDateParsed"]
    reg_out["Saheli Card Number"]=df_reg_first["_SaheliKey"]

    fn=rp("Full Name") or rp("Name")
    reg_out["Full Name:"]    =df_reg_first[fn] if fn else pd.NA
    reg_out["Date of Birth:"]=parse_date(df_reg_first[dob_col]) if dob_col else pd.NA

    if age_col:
        reg_out["AGE"]=pd.to_numeric(df_reg_first[age_col],errors="coerce").astype("Int64")
        if dob_col:
            miss=reg_out["AGE"].isna()
            reg_out.loc[miss,"AGE"]=compute_age(df_reg_first[dob_col],df_reg_first["_RegDateParsed"])[miss]
    else:
        reg_out["AGE"]=compute_age(df_reg_first[dob_col],df_reg_first["_RegDateParsed"]) if dob_col else pd.NA

    # Exact Master personal column names
    reg_assignments=[
        ("Address:",                                            rp("Address")),
        ("Postcode:",                                           rp("Postcode")),
        ("Email:",                                              rp("Email",occurrence=2) or rp("Email",occurrence=1)),
        ("Mobile/Home No:",                                     rp("Mobile/Home No")),
        ("Emergency Contact Name:",                             rp("Emergency Contact Name")),
        ("Emergency No:",                                       rp("Emergency No")),
        ("Emergency Relation To You:",                          rp("Emergency Relation To You")),
        ("Gender:",                                             rp("Gender")),
        ("Is your gender the same as assigned at birth?",       rp("Is your gender the same as assigned at birth")),
        ("Health Conditions/Disability:",                       rp("Health Conditions/Disability")),
        ("Ethnicity:",                                          rp("Ethnicity")),
        ("Preferred spoken language:",                          rp("Preferred spoken language")),
        ("Religion:",                                           rp("Religion")),
        ("Relationship status:",                                rp("Relationship status")),
        ("Caring responsibilities:",                            rp("Caring responsibilities")),
        ("Living alone:",                                       rp("Living alone")),
        ("Sexuality:",                                          rp("Sexuality")),
        ("Occupation:",                                         rp("Occupation")),
        ("Referral reason",                                     rp("Referral reason")),
        ("How heard about Saheli Hub?",                         rp("How heard about Saheli Hub")),
        ("GP Surgery Name:",                                    rp("GP Surgery Name")),
        ("Consent to store information:",                       rp("Consent to store information")),
        ("Health declaration:",                                 rp("Health declaration")),
        ("Permission to be added to Saheli WhatsApp group?",    rp("Permission to be added to Saheli WhatsApp group")),
        ("Permission to be in photos and videos? (Media consent)", rp("Permission to be in photos and videos","Media consent")),
        ("Notes:",                                              rp("Notes")),
        ("Staff Member:",                                       rp("Staff Member")),
        ("Site:",                                               rp("Site")),
    ]
    for out_label,src_col in reg_assignments:
        reg_out[out_label]=df_reg_first[src_col] if src_col else pd.NA

    # Append extra REG cols
    done=set(reg_out.columns)|{"_SaheliKey","_RegDateParsed",reg_saheli_col}
    for c in df_reg_first.columns:
        if c not in done:
            reg_out[c]=df_reg_first[c]

    # Health subset
    df_health=df_health[df_health[assess_col].notna()].copy()
    df_health[assess_col]=pd.to_numeric(df_health[assess_col],errors="coerce").astype("Int64")
    df_health=df_health[df_health[assess_col]<=MAX_ASSESSMENTS].copy()
    df_health[assess_col]=df_health[assess_col].astype(int)

    assess_counts=(
        df_health.dropna(subset=["_SaheliKey"])
        .groupby("_SaheliKey")[assess_col].max()
        .rename("No of assessment completed").astype("Int64")
    )

    def health_source_col(label):
        if label=="Comments:":    return pick_col(health_map,"Comments",occurrence=1)
        if label=="Comments:PA":  return pick_col(health_map,"Comments",occurrence=2)
        if label in ("Comments:2","Comments:3","Comments:4","Comments:5","Comments:6","Comments:7"):
            return (pick_col(health_map,label)
                    or pick_col(health_map,label.replace(":","").replace(" ","")))
        if label=="WEMWBS":          return "WEMBS"   # computed col
        if label=="SOCIAL ISOLATION": return "SOCIAL ISOLATION"
        return (pick_col(health_map,label)
                or pick_col(health_map,label.replace(":",""))
                or pick_col(health_map,label.replace("?","")))

    reserved_health={
        "_SaheliKey","_CompletionDate",assess_col,completion_col,
        health_saheli_col,"WEMBS","SOCIAL ISOLATION"
    }
    extra_health_cols=[c for c in df_health.columns if c not in reserved_health]

    health_wide=pd.DataFrame({"_SaheliKey":reg_out["_SaheliKey"].astype("string")})

    for n in range(1,MAX_ASSESSMENTS+1):
        block=f"{ordinal(n)} Assessment"
        h_n=df_health[df_health[assess_col]==n].copy()
        if h_n.empty: continue

        tmp=pd.DataFrame({"_SaheliKey":h_n["_SaheliKey"].astype("string").values})

        # ── CHANGED: flat col name "1st Assessment  AssessmentDate" ───────────
        tmp[f"{block}  AssessmentDate"]=h_n["_CompletionDate"].values
        added={f"{block}  AssessmentDate"}

        for fld in HEALTH_FIELDS:
            src =health_source_col(fld)
            disp=clean_subheader_label(fld)
            flat_name=f"{block}  {disp}"
            if flat_name in added: continue
            added.add(flat_name)

            if src and src in h_n.columns:
                vals=h_n[src]
                if normalize_header(fld)==normalize_header("Date of next review appointment"):
                    vals=parse_date(vals)
                tmp[flat_name]=vals.values
            else:
                tmp[flat_name]=pd.NA

        # Extra health cols
        for c in extra_health_cols:
            flat_name=f"{block}  {str(c).strip()}"
            if flat_name in added or flat_name in tmp.columns: continue
            added.add(flat_name)
            tmp[flat_name]=h_n[c].values

        tmp=tmp.drop_duplicates(subset=["_SaheliKey"],keep="first")
        health_wide=health_wide.merge(tmp,on="_SaheliKey",how="left")

    # Final merge
    final_df=reg_out.merge(health_wide,on="_SaheliKey",how="left")
    final_df=final_df.merge(assess_counts.reset_index(),on="_SaheliKey",how="left")
    final_df["No of assessment completed"]=final_df["No of assessment completed"].fillna(0).astype("Int64")
    final_df["Saheli Card Number"]=final_df["Saheli Card Number"].fillna(final_df["_SaheliKey"])

    # Column order
    first_cols=list(REG_OUTPUT_LABELS)
    for c in first_cols:
        if c not in final_df.columns: final_df[c]=pd.NA

    rem_cols=[c for c in final_df.columns if c not in first_cols+["_SaheliKey"]]

    health_order_map=build_health_field_order_map()
    def sort_key(c):
        m=re.match(r"^(\d+)(st|nd|rd|th)\sAssessment(?:\s{2,}(.*))?$",str(c))
        if m:
            bn=int(m.group(1))
            sf=(m.group(3) or "").strip()
            if not sf or normalize_header(sf)in("assessmentdate","completiondate"):
                return(0,bn,-1,"")
            d=clean_subheader_label(sf)
            return(0,bn,health_order_map.get(d,9999),d.lower())
        return(1,999,9999,str(c).lower())

    rem_cols=sorted(rem_cols,key=sort_key)
    final_df=final_df[first_cols+rem_cols].copy()

    final_df["_s"]=pd.to_numeric(final_df["Saheli Card Number"],errors="coerce")
    final_df=final_df.sort_values("_s",kind="mergesort").drop(columns=["_s"])

    reg_keys =set(reg_out["_SaheliKey"].dropna().astype(str))
    final_keys=set(final_df["Saheli Card Number"].dropna().astype(str))
    missing=sorted(reg_keys-final_keys)
    print(f"[CHECK] REG keys:   {len(reg_keys)}")
    print(f"[CHECK] FINAL keys: {len(final_keys)}")
    print(f"[CHECK] Missing:    {len(missing)}")
    if missing: print("[CHECK] Sample:",missing[:10])

    return final_df


# =========================
# STEP 4: Write flat Excel (single header row, freeze, autofilter)
# =========================
def write_excel_flat(df,out_path):
    out=Path(out_path)
    out.parent.mkdir(parents=True,exist_ok=True)
    with pd.ExcelWriter(out,engine="openpyxl",
                        datetime_format="dd/mm/yyyy",date_format="dd/mm/yyyy") as writer:
        df.to_excel(writer,index=False)

    wb=load_workbook(out_path)
    ws=wb.active
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(ws.max_column)}1"
    ws.row_dimensions[1].height=30
    for ci in range(1,ws.max_column+1):
        h=str(ws.cell(row=1,column=ci).value or "")
        ws.column_dimensions[get_column_letter(ci)].width=max(12,min(40,len(h)*0.85))
    wb.save(out_path)


# =========================
# MAIN
# =========================
def main():
    print("=== STEP 1: Read source files ===")
    df_reg_raw   =read_excel_flexible(REG_FILE,REG_SHEET_NAME)
    df_health_raw=read_excel_flexible(HEALTH_FILE,HEALTH_SHEET_NAME)
    print(f"Registration rows : {len(df_reg_raw)}")
    print(f"Health rows       : {len(df_health_raw)}")

    print("\n=== STEP 2: Prepare registration ===")
    df_reg_clean=prepare_registration_df(df_reg_raw)
    write_excel(df_reg_clean,REG_OUTPUT_FILE)
    print(f"Saved: {REG_OUTPUT_FILE}")

    print("\n=== STEP 3: Prepare health assessments ===")
    df_health_prepared=prepare_health_df(df_health_raw)
    write_excel(df_health_prepared,HEALTH_OUTPUT_FILE)
    print(f"Saved: {HEALTH_OUTPUT_FILE}")

    print("\n=== STEP 4: Build final wide output ===")
    final_df=create_final_wide_df(df_reg_clean,df_health_prepared)

    if "_SaheliKey" in final_df.columns:
        final_df=final_df.drop(columns=["_SaheliKey"])

    write_excel_flat(final_df,FINAL_OUTPUT_FILE)

    print(f"\n=== DONE ===")
    print(f"Final file : {FINAL_OUTPUT_FILE}")
    print(f"Rows       : {len(final_df)}")
    print(f"Columns    : {len(final_df.columns)}")
    print(f"First cols : {list(final_df.columns[:6])}")
    print(f"Assess cols: {list(final_df.columns[34:37])}")


if __name__=="__main__":
    main()