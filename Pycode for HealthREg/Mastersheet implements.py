# ============================================================
# Q FULL FILE: saheli_upsert_fullregister_from_generated_bestcol.py
# ------------------------------------------------------------
# Fixes the "headers match but values still blank" issue by:
#  - handling duplicate columns in GENERATED (same normalized name)
#  - choosing the GENERATED column with the MOST non-blank values
#    for each normalized header key
#
# RULES:
#  - do not add columns in MASTER
#  - fill blanks only (no overwrite)
#  - add new rows for new Saheli numbers
#  - WEMWBS in MASTER is filled ONLY from Comments:2 in GENERATED (per block)
#    IMPORTANT: do NOT map MASTER WEMWBS from GENERATED WEMWBS
# ============================================================

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


MASTER_FILE = r"C:\Users\shonk\Downloads\Full Registration for SAHELI.xlsx"
MASTER_SHEET = "Full Register"

GENERATED_FILE = r"C:\Users\shonk\source\PythonCodes\New folder\Saheli_Master_Wide_Output.xlsx"
GENERATED_SHEET = 0

MASTER_UPDATED_FILE = r"C:\Users\shonk\Downloads\Full Registration for SAHELI_UPDATED.xlsx"
CHANGELOG_FILE = r"C:\Users\shonk\Downloads\Master_Upsert_CHANGELOG.xlsx"

ALLOW_OVERWRITE = False


def clean_text(x) -> str:
    if x is None:
        return ""
    s = str(x).replace("\r", " ").replace("\n", " ")
    s = s.replace('"', "").replace("“", "").replace("”", "")
    s = re.sub(r"\s+", " ", s).strip()
    if s == "WEMBS":
        s = "WEMWBS"
    if s == "SOCIAL":
        s = "SOCIAL ISOLATION"
    return s


def strip_dup_suffix(s: str) -> str:
    return re.sub(r"__dup\d+$", "", s or "")


def normalize_key(s: str) -> str:
    s = strip_dup_suffix(clean_text(s)).lower().replace("&", "and")
    s = re.sub(r"[\s:\/\?\(\)\-\,\.\'’\"]+", "", s)
    return s


def is_blank(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and pd.isna(x):
        return True
    if isinstance(x, str) and x.strip() == "":
        return True
    return False


def norm_value(x: Any) -> str:
    if is_blank(x):
        return ""
    if isinstance(x, pd.Timestamp):
        return str(x.date())
    return str(x).strip()


def keep_digits_only(v):
    if pd.isna(v):
        return pd.NA
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            if fv.is_integer():
                return str(int(fv))
            digits = re.sub(r"\D+", "", format(fv, "f"))
            return digits if digits else pd.NA
        except Exception:
            pass
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    digits = re.sub(r"\D+", "", s)
    return digits if digits else pd.NA


def safe_saheli_key(v: Any) -> str | None:
    if is_blank(v):
        return None
    d = keep_digits_only(v)
    if is_blank(d):
        return None
    try:
        return str(int(str(d)))
    except Exception:
        s = str(d).strip()
        return s if s else None


def find_saheli_col(cols: List[str]) -> str:
    best = []
    for c in cols:
        nk = normalize_key(c)
        if nk in ("sahelicardnumber", "sahelicardno", "sahelicardno:", "sahelicardno"):
            return c
        if "saheli" in nk and ("card" in nk or "number" in nk or "no" in nk or "num" in nk):
            best.append(c)
    if len(best) == 1:
        return best[0]
    if len(best) > 1:
        return sorted(best, key=lambda x: (len(clean_text(x)), clean_text(x)))[0]
    raise KeyError("Saheli Card Number column not found.")


def flatten_multiindex_columns(cols: pd.Index) -> List[str]:
    out = []
    for a, b in cols:
        a = clean_text(a)
        b = clean_text(b)
        a_is_unnamed = a.lower().startswith("unnamed:")
        b_is_unnamed = b.lower().startswith("unnamed:")
        # In this workbook, level-1 often contains data examples, not true headers.
        # Prefer level-0 header text when present.
        if a and not a_is_unnamed:
            out.append(a)
        elif b and not b_is_unnamed:
            out.append(b)
        else:
            out.append("")
    return [clean_text(c) for c in out]


def make_unique_columns(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in cols:
        c = clean_text(c)
        if not c:
            out.append("")
            continue
        n = seen.get(c, 0)
        out.append(c if n == 0 else f"{c}__dup{n}")
        seen[c] = n + 1
    return out


def read_headers(ws: Worksheet, header_row: int = 1) -> Tuple[List[str], Dict[str, int]]:
    headers = []
    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        h = clean_text(ws.cell(row=header_row, column=col).value)
        headers.append(h)
        if h and h not in header_to_col:
            header_to_col[h] = col
    return headers, header_to_col


def build_norm_map(headers: List[str]) -> Dict[str, str]:
    m = {}
    for h in headers:
        if not h:
            continue
        nk = normalize_key(h)
        if nk and nk not in m:
            m[nk] = h
    return m


def parse_block_and_field(col: str) -> Tuple[str, str]:
    col = clean_text(col)
    if "  " in col:
        a, b = col.split("  ", 1)
        return clean_text(a), clean_text(b)
    return "", col


def is_master_wemwbs_header(h: str) -> bool:
    return bool(re.match(r"^\d+(st|nd|rd|th)\s+Assessment\s{2,}WEMWBS$", clean_text(h)))


def is_comments2_field(field: str) -> bool:
    nk = normalize_key(field)
    return ("comment" in nk) and ("2" in nk)


def nonblank_count(series: pd.Series) -> int:
    # Count values that are not blank / not NaN
    def ok(v):
        if is_blank(v):
            return False
        s = str(v).strip().lower()
        return s not in ("nan", "nat")
    return int(series.map(ok).sum())


def build_best_generated_map(df_gen: pd.DataFrame) -> Tuple[Dict[str, str], List[Dict[str, Any]]]:
    """
    Returns:
      best_map: normalized_key -> chosen_column_name (max non-blank count)
      report_rows: list describing duplicates and chosen winner
    """
    cols_by_nk: Dict[str, List[str]] = {}
    for c in df_gen.columns:
        nk = normalize_key(str(c))
        if not nk:
            continue
        cols_by_nk.setdefault(nk, []).append(str(c))

    best_map: Dict[str, str] = {}
    report_rows: List[Dict[str, Any]] = []

    for nk, cols in cols_by_nk.items():
        if len(cols) == 1:
            best_map[nk] = cols[0]
            continue

        scored = []
        for c in cols:
            try:
                cnt = nonblank_count(df_gen[c])
            except Exception:
                cnt = 0
            scored.append((cnt, c))

        scored.sort(reverse=True, key=lambda x: (x[0], -len(x[1]), x[1]))
        best = scored[0][1]
        best_map[nk] = best

        report_rows.append(
            {
                "NormalizedKey": nk,
                "CandidateColumns": " | ".join(cols),
                "Counts": " | ".join([f"{c}={cnt}" for cnt, c in scored]),
                "Chosen": best,
            }
        )

    return best_map, report_rows


def main():
    master_path = Path(MASTER_FILE)
    gen_path = Path(GENERATED_FILE)
    if not master_path.exists():
        raise FileNotFoundError(master_path)
    if not gen_path.exists():
        raise FileNotFoundError(gen_path)

    # Load master
    wb = load_workbook(master_path, data_only=False, keep_links=False)
    if MASTER_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{MASTER_SHEET}' not found.")
    ws = wb[MASTER_SHEET]

    master_headers, master_header_to_col = read_headers(ws, 1)
    master_norm = build_norm_map(master_headers)
    master_saheli = find_saheli_col([h for h in master_headers if h])
    master_saheli_idx = master_header_to_col[master_saheli]

    print("=== MASTER ===")
    print("Cols:", len([h for h in master_headers if h]))
    print("Saheli:", master_saheli)

    # Read generated (try 2-row, else 1-row)
    print("\n=== GENERATED ===")
    try:
        df_gen = pd.read_excel(gen_path, sheet_name=GENERATED_SHEET, header=[0, 1])
        df_gen.columns = flatten_multiindex_columns(df_gen.columns)
    except Exception:
        df_gen = pd.read_excel(gen_path, sheet_name=GENERATED_SHEET, header=0)
        df_gen.columns = [clean_text(c) for c in df_gen.columns]

    df_gen = df_gen[[c for c in df_gen.columns if clean_text(c)]].copy()
    df_gen.columns = make_unique_columns([str(c) for c in df_gen.columns])
    print("Rows:", len(df_gen), "Cols:", len(df_gen.columns))

    gen_saheli = find_saheli_col(list(df_gen.columns))
    print("Generated Saheli:", gen_saheli)

    # Build BEST map for duplicates (this is the fix)
    gen_best_map, dup_report = build_best_generated_map(df_gen)

    # Build Comments2-by-block mapping (block -> "<block>  Comments:2" column)
    gen_comments2_by_block: Dict[str, str] = {}
    for c in df_gen.columns:
        block, field = parse_block_and_field(str(c))
        if block and is_comments2_field(field):
            gen_comments2_by_block[normalize_key(block)] = str(c)

    # WEMWBS pairs: MASTER "<block>  WEMWBS" <- GENERATED "<block>  Comments:2"
    wemwbs_pairs: List[Tuple[str, str]] = []
    for mcol in master_headers:
        if not mcol or not is_master_wemwbs_header(mcol):
            continue
        block, _ = parse_block_and_field(mcol)
        g_c2 = gen_comments2_by_block.get(normalize_key(block))
        if g_c2:
            wemwbs_pairs.append((mcol, g_c2))

    # Normal sync pairs: intersection only, EXCLUDING master WEMWBS
    # AND also excluding any generated WEMWBS column (never map WEMWBS <- WEMWBS)
    sync_pairs: List[Tuple[str, str]] = []
    for nk, mcol in master_norm.items():
        if is_master_wemwbs_header(mcol):
            continue
        gcol = gen_best_map.get(nk)
        if not gcol:
            continue
        # Do not allow a generated WEMWBS column to be used in normal mapping
        if normalize_key(gcol).endswith("wemwbs"):
            continue
        sync_pairs.append((mcol, gcol))

    print("[INFO] Sync pairs:", len(sync_pairs))
    print("[INFO] WEMWBS pairs:", len(wemwbs_pairs))
    if len(wemwbs_pairs) == 0:
        print("[WARN] No WEMWBS pairs found. Ensure Generated has '<block>  Comments:2' columns.")

    # Index master keys
    master_key_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        k = safe_saheli_key(ws.cell(row=r, column=master_saheli_idx).value)
        if k and k not in master_key_to_row:
            master_key_to_row[k] = r
    print("[INFO] Master keys:", len(master_key_to_row))

    cell_changes: List[Dict[str, Any]] = []
    new_rows: List[str] = []

    def mcol_idx(h: str) -> int | None:
        return master_header_to_col.get(h)

    # UPSERT
    for i in range(len(df_gen)):
        key = safe_saheli_key(df_gen.at[i, gen_saheli])
        if not key:
            continue

        if key in master_key_to_row:
            r = master_key_to_row[key]
            # fill blanks
            for mcol, gcol in sync_pairs:
                mci = mcol_idx(mcol)
                if not mci:
                    continue
                nv = df_gen.at[i, gcol]
                if is_blank(nv):
                    continue
                ov = ws.cell(row=r, column=mci).value
                if is_blank(ov):
                    ws.cell(row=r, column=mci).value = nv
                    cell_changes.append(
                        {
                            "Saheli Card Number": key,
                            "RowType": "Existing",
                            "MasterRow": r,
                            "MasterColumn": mcol,
                            "GeneratedColumn": gcol,
                            "OldValue": ov,
                            "NewValue": nv,
                            "ChangeType": "FilledBlank",
                        }
                    )
                elif ALLOW_OVERWRITE and norm_value(ov) != norm_value(nv):
                    ws.cell(row=r, column=mci).value = nv
                    cell_changes.append(
                        {
                            "Saheli Card Number": key,
                            "RowType": "Existing",
                            "MasterRow": r,
                            "MasterColumn": mcol,
                            "GeneratedColumn": gcol,
                            "OldValue": ov,
                            "NewValue": nv,
                            "ChangeType": "Overwritten",
                        }
                    )

            for m_w, g_c2 in wemwbs_pairs:
                mci = mcol_idx(m_w)
                if not mci:
                    continue
                nv = df_gen.at[i, g_c2]
                if is_blank(nv):
                    continue
                ov = ws.cell(row=r, column=mci).value
                if is_blank(ov):
                    ws.cell(row=r, column=mci).value = nv
                    cell_changes.append(
                        {
                            "Saheli Card Number": key,
                            "RowType": "Existing",
                            "MasterRow": r,
                            "MasterColumn": m_w,
                            "GeneratedColumn": g_c2,
                            "OldValue": ov,
                            "NewValue": nv,
                            "ChangeType": "FilledBlank(WEMWBS<-Comments2)",
                        }
                    )

        else:
            new_rows.append(key)
            new_r = ws.max_row + 1
            ws.cell(row=new_r, column=master_saheli_idx).value = key

            for mcol, gcol in sync_pairs:
                mci = mcol_idx(mcol)
                if not mci:
                    continue
                v = df_gen.at[i, gcol]
                if is_blank(v):
                    continue
                ov = ws.cell(row=new_r, column=mci).value
                if is_blank(ov):
                    ws.cell(row=new_r, column=mci).value = v
                    cell_changes.append(
                        {
                            "Saheli Card Number": key,
                            "RowType": "New",
                            "MasterRow": new_r,
                            "MasterColumn": mcol,
                            "GeneratedColumn": gcol,
                            "OldValue": ov,
                            "NewValue": v,
                            "ChangeType": "InsertedRowValue",
                        }
                    )

            for m_w, g_c2 in wemwbs_pairs:
                mci = mcol_idx(m_w)
                if not mci:
                    continue
                v = df_gen.at[i, g_c2]
                if is_blank(v):
                    continue
                ov = ws.cell(row=new_r, column=mci).value
                if is_blank(ov):
                    ws.cell(row=new_r, column=mci).value = v
                    cell_changes.append(
                        {
                            "Saheli Card Number": key,
                            "RowType": "New",
                            "MasterRow": new_r,
                            "MasterColumn": m_w,
                            "GeneratedColumn": g_c2,
                            "OldValue": ov,
                            "NewValue": v,
                            "ChangeType": "InsertedRowValue(WEMWBS<-Comments2)",
                        }
                    )

            master_key_to_row[key] = new_r

    # Save updated master
    out_master = Path(MASTER_UPDATED_FILE)
    out_master.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_master)

    # Changelog
    ch_path = Path(CHANGELOG_FILE)
    ch_path.parent.mkdir(parents=True, exist_ok=True)

    df_changes = pd.DataFrame(cell_changes)
    df_summary = pd.DataFrame(
        {
            "MasterSheet": [MASTER_SHEET],
            "GeneratedRows": [len(df_gen)],
            "GeneratedCols": [len(df_gen.columns)],
            "NewRowsAdded": [len(new_rows)],
            "CellsFilledOrInserted": [int(len(df_changes))],
            "SyncPairs": [len(sync_pairs)],
            "WEMWBSPairs": [len(wemwbs_pairs)],
            "OverwriteEnabled": [ALLOW_OVERWRITE],
            "WEMWBSRule": ["MASTER <block WEMWBS> <- GENERATED <block Comments:2> only"],
            "DuplicateKeysResolved": [len(dup_report)],
        }
    )

    # ColumnMapping output (what master maps to)
    mapping_rows: List[Dict[str, Any]] = []
    for mcol in [h for h in master_headers if h]:
        if is_master_wemwbs_header(mcol):
            block, _ = parse_block_and_field(mcol)
            g = gen_comments2_by_block.get(normalize_key(block))
            mapping_rows.append(
                {"MasterColumn": mcol, "GeneratedColumn": g if g else "NOT FOUND", "Rule": "SPECIAL: WEMWBS<-Comments2"}
            )
        else:
            g = gen_best_map.get(normalize_key(mcol))
            # show what would map, but note we intentionally skip generated WEMWBS in sync_pairs
            mapping_rows.append(
                {"MasterColumn": mcol, "GeneratedColumn": g if g else "NOT FOUND", "Rule": "Best nonblank duplicate"}
            )

    with pd.ExcelWriter(ch_path, engine="openpyxl") as w:
        df_summary.to_excel(w, index=False, sheet_name="Summary")
        pd.DataFrame({"NewSaheliCardNumber": new_rows}).to_excel(w, index=False, sheet_name="NewRows")
        pd.DataFrame(mapping_rows).to_excel(w, index=False, sheet_name="ColumnMapping")
        pd.DataFrame(dup_report).to_excel(w, index=False, sheet_name="DuplicateResolution")
        df_changes.to_excel(w, index=False, sheet_name="CellChanges")

    print("\n=== DONE ===")
    print("Updated master:", str(out_master))
    print("Changelog:", str(ch_path))
    print("New rows:", len(new_rows))
    print("Cells filled/inserted:", len(df_changes))


if __name__ == "__main__":
    main()