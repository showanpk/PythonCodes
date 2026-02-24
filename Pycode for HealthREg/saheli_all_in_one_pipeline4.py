# ============================================================
# Q FULL FILE: saheli_all_in_one_pipeline.py
# ------------------------------------------------------------
# ALL-IN-ONE pipeline (NO COLUMN DROPS):
#   1) Prepare Registrations file
#   2) Prepare Healthassessments file
#   3) Create final WIDE output (LEFT JOIN from REGISTRATION)
#
# GUARANTEE:
#   - Do NOT miss any columns from REG_FILE or HEALTH_FILE
#   - REG is the PRIMARY table (LEFT JOIN)
#   - If Saheli exists in REG but no assessment, still included
#
# Includes:
#   - Safe Saheli ID cleaning (prevents 1 -> 10 bug)
#   - Flexible header matching (spaces/newlines/colons/quotes)
#   - AssessmentNumber generation
#   - Assessment date stored in "1st Assessment", "2nd Assessment", ...
#   - Health columns ordered by HEALTH_FIELDS first, THEN any extras appended (per assessment)
#   - Registration columns ordered by REG_OUTPUT_LABELS first, THEN any extras appended
#   - Fix: DO NOT collapse Comments:PA to Comments: (prevents overwrite)
#
# Install:
#   pip install pandas openpyxl
# ============================================================

from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
REG_FILE = r"C:\Users\shonk\Downloads\Main Registration Form(1-1143).xlsx"
HEALTH_FILE = r"C:\Users\shonk\Downloads\Saheli Hub Health Assessment(1-1478).xlsx"

REG_OUTPUT_FILE = r"C:\Users\shonk\source\PythonCodes\New folder\Registrations_Cleaned.xlsx"
HEALTH_OUTPUT_FILE = r"C:\Users\shonk\source\PythonCodes\New folder\Healthassessments_Prepared.xlsx"
FINAL_OUTPUT_FILE = r"C:\Users\shonk\source\PythonCodes\New folder\Saheli_Master_Wide_Output.xlsx"

# If needed set explicit sheet names, otherwise None = first sheet
REG_SHEET_NAME = None
HEALTH_SHEET_NAME = None

MAX_ASSESSMENTS = 8

# Set True if you want the 2-row grouped header in Excel output
APPLY_GROUPED_HEADER = True

# Header styling (Excel grouped header)
COLOR_FIRST_ASSESSMENT = "7030A0"   # purple
COLOR_SECOND_ASSESSMENT = "00B0F0"  # blue
COLOR_OTHER_ASSESSMENT = "D9E2F3"   # light blue/grey
COLOR_REG_HEADER = "FFFFFF"         # white


# =========================
# HELPERS
# =========================
def normalize_header(h) -> str:
    """Flexible header normalizer (handles newlines, quotes, colons, spaces, etc.)."""
    if h is None:
        return ""
    s = str(h)

    # Remove Excel/Forms line breaks and quotes
    s = s.replace("\r", "").replace("\n", "")
    s = s.replace('"', "").replace("“", "").replace("”", "")

    s = s.strip().lower()
    s = s.replace(" ", "").replace(":", "")
    s = s.replace("/", "")
    s = s.replace("?", "")
    s = s.replace("(", "").replace(")", "")
    s = s.replace("-", "")
    s = s.replace(",", "")
    s = s.replace(".", "")
    s = s.replace("&", "and")
    s = s.replace("’", "").replace("'", "")
    return s


def keep_digits_only(v):
    """
    Safe digit cleaner for IDs.
    FIX: prevents 1.0 -> '10'
    """
    if pd.isna(v):
        return pd.NA

    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            if fv.is_integer():
                return str(int(fv))
            s_num = format(fv, "f")
            digits = re.sub(r"\D+", "", s_num)
            return digits if digits else pd.NA
        except Exception:
            pass

    s = str(v).strip()

    # Handle "14.0", "27.000"
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]

    digits = re.sub(r"\D+", "", s)
    return digits if digits else pd.NA


def safe_saheli_key_series(series: pd.Series) -> pd.Series:
    """Standardize Saheli key for joining."""
    s = series.apply(keep_digits_only)
    s = pd.to_numeric(s, errors="coerce").astype("Int64")
    return s.astype("string")


def read_excel_flexible(path: str, sheet_name=None) -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {p}")
    return pd.read_excel(p) if sheet_name is None else pd.read_excel(p, sheet_name=sheet_name)


def write_excel(df: pd.DataFrame, out_path: str):
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl", datetime_format="dd/mm/yyyy", date_format="dd/mm/yyyy") as writer:
        df.to_excel(writer, index=False)
    return out


def build_normalized_col_map(df: pd.DataFrame):
    """normalized_header -> [actual columns in order]"""
    m = {}
    for c in df.columns:
        k = normalize_header(c)
        m.setdefault(k, []).append(c)
    return m


def pick_col(norm_map, *candidates, occurrence=1):
    """Pick column by normalized candidate names."""
    for cand in candidates:
        k = normalize_header(cand)
        if k in norm_map and len(norm_map[k]) >= occurrence:
            return norm_map[k][occurrence - 1]
    return None


def parse_date(series):
    dt = pd.to_datetime(series, errors="coerce", dayfirst=True)
    return dt.dt.date


def compute_age(dob_series, ref_series=None):
    dob = pd.to_datetime(dob_series, errors="coerce", dayfirst=True)
    if ref_series is None:
        ref = pd.Series([pd.Timestamp.today().normalize()] * len(dob), index=dob.index)
    else:
        ref = pd.to_datetime(ref_series, errors="coerce", dayfirst=True)

    age = ref.dt.year - dob.dt.year
    before_bday = (ref.dt.month < dob.dt.month) | (
        (ref.dt.month == dob.dt.month) & (ref.dt.day < dob.dt.day)
    )
    return (age - before_bday.astype(int)).astype("Int64")


def ordinal(n):
    if 10 <= n % 100 <= 20:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"


def clean_subheader_label(label: str) -> str:
    """
    Convert internal column labels to expected display labels WITHOUT COLLISIONS.
    IMPORTANT:
      - DO NOT map Comments:PA -> Comments: (this causes overwriting)
    """
    s = str(label).strip()

    # Keep these consistent
    if s == "WEMBS":
        return "WEMWBS"
    if s == "SOCIAL ISOLATION":
        return "SOCIAL ISOLATION"

    return s


def build_health_field_order_map():
    """Order index for assessment subheaders based on HEALTH_FIELDS (preserves exact order)."""
    order_map = {}
    for i, fld in enumerate(HEALTH_FIELDS):
        disp = clean_subheader_label(fld)
        if disp not in order_map:
            order_map[disp] = i
    return order_map


# =========================
# OUTPUT LABELS (preferred ordering)
# NOTE: We will still append ANY extra columns found in source, so nothing is dropped.
# =========================
REG_OUTPUT_LABELS = [
    "Registration Date",
    "Saheli Card Number",
    " Full Name:",
    " Date of Birth:",
    "AGE",
    " Address:",
    " Postcode:",
    " Email:",
    " Mobile/Home No:",
    " Emergency Contact Name:",
    " Emergency No:",
    " Emergency Relation To You:",
    " Gender:",
    " Is your gender the same as assigned at birth?",
    " Health Conditions/Disability:",
    " Ethnicity:",
    " Preferred spoken language:",
    " Religion:",
    " Relationship status:",
    " Caring responsibilities:",
    " Living alone:",
    " Sexuality:",
    " Occupation:",
    "Referral reason",
    " How heard about Saheli Hub?",
    "GP Surgery Name:",
    " Consent to store information:",
    " Health declaration:",
    " Permission to be added to Saheli WhatsApp group?",
    " Permission to be in photos and videos? (Media consent)",
    "Notes:",
    " Staff Member:",
    "Site:",
]

HEALTH_FIELDS = [
    " Weight (KG):",
    " Height (CM):",
    " BMI:",
    " BMI Results:",
    " Waist (CM):",
    " Hip (CM):",
    " Waist to Hip Ratio (CM):",
    " Body Fat Percentage Result:",
    " Body Fat Percentage Score:",
    " Visceral Fat Level Result:",
    " Visceral Fat Level Score:",
    " Skeletal Muscle Percentage:",
    " Skeletal Muscle Score:",
    " Resting Metabolism:",
    " Do You Have Any Health Condition?",
    "When did you last measure your blood pressure?",
    "Have you recorded your blood pressure measurement and registered it with a GP or Pharmacist? (Yes/No/Not sure)",
    "What is a healthy blood pressure for an adult?",
    "Why is a high blood pressure dangerous?",
    "How can you help reduce your blood pressure?",
    " Blood Pressure (Systolic/Diastolic):",
    " Blood Pressure Level:",
    " Do You Have a Heart Condition?",
    " Heart Rate (BPM):",
    " Atrial Fibrillation Result:",
    " Heart Age:",
    " Did Your Doctor Advise You Not to Exercise?",
    " Do You Feel Pain in Chest at Rest/During Activity?",
    " Do You Have Shortness of Breath?",
    " Do You Have Diabetes?",
    " Diabetes Risk:",
    " Glucose Level ( mg/dL):",
    " HbA1c:",
    " Do You Take Sugary Drinks, Including Chai?",
    " Do You Have High Cholesterol? (Total/HDL)",
    " Do You Experience The Following Health Issues?",
    " Do You Have a Bone / joint Condition?",
    " Do You Take Any Prescribed Medication?",
    " Referred to doctor for any concerning results?",
    " Risk Stratification Score",
    " Comments:",       # Risk stratification comments (1st "Comments:")
    " How well do you manage your health/condition(s)? (Rating out of 10)",
    " In the past week, on how many days have you done a total of 30 mins or more of physical activity, which was enough to raise your breathing rate?",
    " Physical Activity Level:",
    "Comments:PA",      # Physical activity comments (2nd "Comments:" in source)
    "I’ve been feeling optimistic about the future",
    "I’ve been feeling useful",
    "I’ve been feeling relaxed",
    "I’ve been feeling interested in other people",
    "I’ve had energy to spare",
    "I’ve been dealing with problems well",
    "I’ve been thinking clearly",
    "I’ve been feeling good about myself",
    "I’ve been feeling close to other people",
    "I’ve been feeling confident",
    "I’ve been able to make up my own mind about things",
    "I’ve been feeling loved",
    "I’ve been interested in new things",
    "I’ve been feeling cheerful",
    "WEMBS",            # computed -> displayed as WEMWBS
    "Comments:2",
    " Nourishment: Rate the quality of the food you put into your body on a daily basis",
    " Movement: Rate how often and for how long you move your body on a daily basis",
    " Connectedness: Rate how well you stay connected with family, friends and your higher power",
    " Sleep: Rate the quality of your sleep",
    " Happy self: Rate how often and for how long you perform positive practices (gratitude, virtue awareness, meditation, prayer, etc.)",
    " Resilience: Rate how well you are able to manage stress in your life",
    " Green and Blue: Rate how often and how long you spend in nature or outdoors",
    " Screen time: Rate how happy you are with your current amount of screen time",
    " Substance use: Rate how comfortable you are with any current substance use (smoking, alcohol, drugs)",
    " Purpose: Rate how well you feel you are fulfilling your passion, purpose or vocation in life",
    "Comments:3",
    " How often do you feel that you lack companionship?",
    "How often do you feel left out?",
    " How often do you feel isolated from others?",
    "SOCIAL ISOLATION",  # computed -> displayed as SOCIAL ISOLATION (or you can rename later)
    "Comments:4",
    " How confident are you to join activities?",
    " How many hobbies and passions do you have?",
    " How involved you feel in your community?",
    " How much you know about local support/services?",
    "What are your aims & goals?",
    "Comments:5",
    " What reasons stop you from joining activities?",
    "Comments:6",
    "What are your preferred activities?",
    "Comments:7",
    " Date of next review appointment:",
]


# =========================
# STEP 1: PREP REGISTRATION
# =========================
def prepare_registration_df(df_reg: pd.DataFrame) -> pd.DataFrame:
    reg_map = build_normalized_col_map(df_reg)
    col_saheli = pick_col(reg_map, "Saheli Card No", "SaheliCardNo", "Saheli Card Number")
    if not col_saheli:
        raise KeyError("Registration Saheli Card No column not found.")

    df = df_reg.copy()
    df[col_saheli] = df[col_saheli].apply(keep_digits_only)
    df[col_saheli] = pd.to_numeric(df[col_saheli], errors="coerce").astype("Int64")
    return df


# =========================
# STEP 2: PREP HEALTH
# =========================
def prepare_health_df(df_health: pd.DataFrame) -> pd.DataFrame:
    health_map = build_normalized_col_map(df_health)
    col_completion = pick_col(health_map, "Completion time")
    col_saheli = pick_col(health_map, "Saheli Card No", "SaheliCardNo", "Saheli Card Number")

    if not col_completion:
        raise KeyError("Health 'Completion time' column not found.")
    if not col_saheli:
        raise KeyError("Health 'Saheli Card No' column not found.")

    df = df_health.copy()

    # Clean Saheli
    df[col_saheli] = df[col_saheli].apply(keep_digits_only)
    df[col_saheli] = pd.to_numeric(df[col_saheli], errors="coerce").astype("Int64")

    # Convert Completion time to date
    df[col_completion] = pd.to_datetime(df[col_completion], errors="coerce", dayfirst=True).dt.date

    # Move Saheli next to Completion
    cols = list(df.columns)
    cols.remove(col_saheli)
    idx_completion = cols.index(col_completion)
    cols.insert(idx_completion + 1, col_saheli)
    df = df[cols].copy()

    # Re-find after move
    health_map2 = build_normalized_col_map(df)
    col_completion = pick_col(health_map2, "Completion time")
    col_saheli = pick_col(health_map2, "Saheli Card No", "SaheliCardNo", "Saheli Card Number")

    # Sort by Saheli + Completion
    df = df.sort_values(
        by=[col_saheli, col_completion],
        ascending=[True, True],
        na_position="last",
        kind="mergesort",
    ).reset_index(drop=True)

    # Add AssessmentNumber
    assessment_num = (df.groupby(col_saheli, dropna=False).cumcount() + 1)
    assessment_num = assessment_num.where(df[col_saheli].notna(), pd.NA).astype("Int64")

    pos_saheli = df.columns.get_loc(col_saheli)
    df.insert(pos_saheli + 1, "AssessmentNumber", assessment_num)

    return df


# =========================
# STEP 3: BUILD FINAL WIDE (NO COLUMN DROPS)
# =========================
def create_final_wide_df(df_reg_clean: pd.DataFrame, df_health_prepared: pd.DataFrame) -> pd.DataFrame:
    reg_map = build_normalized_col_map(df_reg_clean)
    health_map = build_normalized_col_map(df_health_prepared)

    reg_saheli_col = pick_col(reg_map, "Saheli Card No", "SaheliCardNo", "Saheli Card Number")
    health_saheli_col = pick_col(health_map, "Saheli Card No", "SaheliCardNo", "Saheli Card Number")
    completion_col = pick_col(health_map, "Completion time")
    assess_col = pick_col(health_map, "AssessmentNumber")

    if not reg_saheli_col:
        raise KeyError("Registration Saheli Card No column not found in cleaned registration file.")
    if not health_saheli_col:
        raise KeyError("Health Saheli Card No column not found in prepared health file.")
    if not completion_col:
        raise KeyError("Health Completion time column not found in prepared health file.")
    if not assess_col:
        raise KeyError("AssessmentNumber column not found in prepared health file.")

    df_reg = df_reg_clean.copy()
    df_health = df_health_prepared.copy()

    # Strong standardized keys
    df_reg["_SaheliKey"] = safe_saheli_key_series(df_reg[reg_saheli_col])
    df_health["_SaheliKey"] = safe_saheli_key_series(df_health[health_saheli_col])

    # Health completion date
    df_health["_CompletionDate"] = parse_date(df_health[completion_col])

    # Ensure assessment number numeric
    df_health[assess_col] = pd.to_numeric(df_health[assess_col], errors="coerce").astype("Int64")

    # Compute WEMWBS (WEMBS label) and Social Isolation
    wem_items = [
        pick_col(health_map, "I’ve been feeling optimistic about the future"),
        pick_col(health_map, "I’ve been feeling useful"),
        pick_col(health_map, "I’ve been feeling relaxed"),
        pick_col(health_map, "I’ve been feeling interested in other people"),
        pick_col(health_map, "I’ve had energy to spare"),
        pick_col(health_map, "I’ve been dealing with problems well"),
        pick_col(health_map, "I’ve been thinking clearly"),
        pick_col(health_map, "I’ve been feeling good about myself"),
        pick_col(health_map, "I’ve been feeling close to other people"),
        pick_col(health_map, "I’ve been feeling confident"),
        pick_col(health_map, "I’ve been able to make up my own mind about things"),
        pick_col(health_map, "I’ve been feeling loved"),
        pick_col(health_map, "I’ve been interested in new things"),
        pick_col(health_map, "I’ve been feeling cheerful"),
    ]
    wem_items = [c for c in wem_items if c]
    for c in wem_items:
        df_health[c] = pd.to_numeric(df_health[c], errors="coerce")
    df_health["WEMBS"] = df_health[wem_items].sum(axis=1, min_count=1) if wem_items else pd.NA

    social_items = [
        pick_col(health_map, "How often do you feel that you lack companionship?"),
        pick_col(health_map, "How often do you feel left out?"),
        pick_col(health_map, "How often do you feel isolated from others?"),
    ]
    social_items = [c for c in social_items if c]
    for c in social_items:
        df_health[c] = pd.to_numeric(df_health[c], errors="coerce")
    df_health["SOCIAL ISOLATION"] = df_health[social_items].sum(axis=1, min_count=1) if social_items else pd.NA

    # ----------------------------------------
    # REGISTRATION BASE (REG is primary)
    # ----------------------------------------
    reg_date_col = pick_col(reg_map, "Date") or pick_col(reg_map, "Completion time") or pick_col(reg_map, "Start time")
    dob_col = pick_col(reg_map, "Date of Birth")
    age_col = pick_col(reg_map, "Age")

    if reg_date_col:
        df_reg["_RegDateParsed"] = parse_date(df_reg[reg_date_col])
    else:
        df_reg["_RegDateParsed"] = pd.NaT

    # Keep only usable Saheli keys for joining, dedupe earliest reg
    df_reg_base = df_reg[df_reg["_SaheliKey"].notna()].copy()
    df_reg_base = df_reg_base.sort_values(
        by=["_SaheliKey", "_RegDateParsed"],
        ascending=[True, True],
        na_position="last",
        kind="mergesort",
    )
    df_reg_first = df_reg_base.drop_duplicates(subset=["_SaheliKey"], keep="first").copy()
    reg_first_map = build_normalized_col_map(df_reg_first)

    def reg_pick(*cands, occurrence=1):
        return pick_col(reg_first_map, *cands, occurrence=occurrence)

    # Build ordered REG output (preferred labels first)
    reg_out = pd.DataFrame({"_SaheliKey": df_reg_first["_SaheliKey"].astype("string")})
    reg_out["Registration Date"] = df_reg_first["_RegDateParsed"]
    reg_out["Saheli Card Number"] = df_reg_first["_SaheliKey"]

    # Preferred mapping for core fields
    if reg_pick("Full Name"):
        reg_out[" Full Name:"] = df_reg_first[reg_pick("Full Name")]
    else:
        # fallback: maybe "Name" exists
        reg_out[" Full Name:"] = df_reg_first[reg_pick("Name")] if reg_pick("Name") else pd.NA

    reg_out[" Date of Birth:"] = parse_date(df_reg_first[dob_col]) if dob_col else pd.NA

    if age_col:
        reg_out["AGE"] = pd.to_numeric(df_reg_first[age_col], errors="coerce").astype("Int64")
        if dob_col:
            miss = reg_out["AGE"].isna()
            calc_age = compute_age(df_reg_first[dob_col], df_reg_first["_RegDateParsed"])
            reg_out.loc[miss, "AGE"] = calc_age[miss]
    else:
        reg_out["AGE"] = compute_age(df_reg_first[dob_col], df_reg_first["_RegDateParsed"]) if dob_col else pd.NA

    reg_assignments = [
        (" Address:", reg_pick("Address")),
        (" Postcode:", reg_pick("Postcode")),
        (" Email:", reg_pick("Email", occurrence=2) or reg_pick("Email", occurrence=1)),
        (" Mobile/Home No:", reg_pick("Mobile/Home No")),
        (" Emergency Contact Name:", reg_pick("Emergency Contact Name")),
        (" Emergency No:", reg_pick("Emergency No")),
        (" Emergency Relation To You:", reg_pick("Emergency Relation To You")),
        (" Gender:", reg_pick("Gender")),
        (" Is your gender the same as assigned at birth?", reg_pick("Is your gender the same as assigned at birth")),
        (" Health Conditions/Disability:", reg_pick("Health Conditions/Disability")),
        (" Ethnicity:", reg_pick("Ethnicity")),
        (" Preferred spoken language:", reg_pick("Preferred spoken language")),
        (" Religion:", reg_pick("Religion")),
        (" Relationship status:", reg_pick("Relationship status")),
        (" Caring responsibilities:", reg_pick("Caring responsibilities")),
        (" Living alone:", reg_pick("Living alone")),
        (" Sexuality:", reg_pick("Sexuality")),
        (" Occupation:", reg_pick("Occupation")),
        ("Referral reason", reg_pick("Referral reason")),
        (" How heard about Saheli Hub?", reg_pick("How heard about Saheli Hub")),
        ("GP Surgery Name:", reg_pick("GP Surgery Name")),
        (" Consent to store information:", reg_pick("Consent to store information")),
        (" Health declaration:", reg_pick("Health declaration")),
        (" Permission to be added to Saheli WhatsApp group?", reg_pick("Permission to be added to Saheli WhatsApp group")),
        (" Permission to be in photos and videos? (Media consent)", reg_pick("Permission to be in photos and videos", "Media consent")),
        ("Notes:", reg_pick("Notes")),
        (" Staff Member:", reg_pick("Staff Member")),
        ("Site:", reg_pick("Site")),
    ]
    for out_label, src_col in reg_assignments:
        reg_out[out_label] = df_reg_first[src_col] if src_col else pd.NA

    # IMPORTANT: Append ANY extra REG columns not already included (so nothing is missed)
    preferred_reg_cols = set(reg_out.columns)
    excluded_reg = {
        "_SaheliKey", "_RegDateParsed", reg_saheli_col
    }
    # Keep original source column order
    for c in df_reg_first.columns:
        if c in excluded_reg:
            continue
        # If the source column name is already represented in our outputs, skip
        # Otherwise include it (using original label)
        if c not in reg_out.columns and c not in preferred_reg_cols:
            reg_out[c] = df_reg_first[c]

    # ----------------------------------------
    # HEALTH subset (valid assessment numbers)
    # ----------------------------------------
    df_health = df_health[df_health[assess_col].notna()].copy()
    df_health[assess_col] = pd.to_numeric(df_health[assess_col], errors="coerce").astype("Int64")
    df_health = df_health[df_health[assess_col].notna()].copy()
    df_health[assess_col] = df_health[assess_col].astype(int)
    df_health = df_health[df_health[assess_col] <= MAX_ASSESSMENTS].copy()

    # Count assessments
    assess_counts = (
        df_health.dropna(subset=["_SaheliKey"])
        .groupby("_SaheliKey")[assess_col]
        .max()
        .rename("No of assessment completed")
        .astype("Int64")
    )

    # Map health labels to source columns (including duplicate Comments)
    def health_source_col(label):
        # 1st Comments (risk stratification)
        if label == " Comments:":
            return pick_col(health_map, "Comments", occurrence=1)

        # 2nd Comments (physical activity) -> map to the 2nd "Comments" in source
        if label == "Comments:PA":
            return pick_col(health_map, "Comments", occurrence=2)

        # if label == "Comments:2":
        #     return pick_col(health_map, "Comments:2") or pick_col(health_map, "Comments2")
        if label == "Comments:3":
            return pick_col(health_map, "Comments:3") or pick_col(health_map, "Comments3")
        if label == "Comments:4":
            return pick_col(health_map, "Comments:4") or pick_col(health_map, "Comments4")
        if label == "Comments:5":
            return pick_col(health_map, "Comments:5") or pick_col(health_map, "Comments5")
        if label == "Comments:6":
            return pick_col(health_map, "Comments:6") or pick_col(health_map, "Comments6")
        if label == "Comments:7":
            return pick_col(health_map, "Comments:7") or pick_col(health_map, "Comments7")

        if label == "WEMBS":
            return "WEMBS"
        if label == "SOCIAL ISOLATION":
            return "SOCIAL ISOLATION"

        return (
            pick_col(health_map, label)
            or pick_col(health_map, label.replace(":", ""))
            or pick_col(health_map, label.replace("?", ""))
        )

    # Build health wide table ONLY for keys in registration (REG primary)
    health_wide = pd.DataFrame({"_SaheliKey": reg_out["_SaheliKey"].astype("string")})

    # Identify "extra" health columns to append per assessment so nothing is missed
    reserved_health_cols = {
        "_SaheliKey", "_CompletionDate", assess_col, completion_col, health_saheli_col
    }
    # We'll also reserve computed columns so they are not duplicated
    reserved_health_cols.update({"WEMBS", "SOCIAL ISOLATION"})

    # Anything else in df_health is "extra"
    extra_health_cols = [c for c in df_health.columns if c not in reserved_health_cols]

    for n in range(1, MAX_ASSESSMENTS + 1):
        block = f"{ordinal(n)} Assessment"
        h_n = df_health[df_health[assess_col] == n].copy()
        if h_n.empty:
            continue

        tmp = pd.DataFrame({"_SaheliKey": h_n["_SaheliKey"].astype("string").values})

        # Assessment date
        tmp["AssessmentDate"] = h_n["_CompletionDate"].values

        # Add columns in exact HEALTH_FIELDS order first
        added_disp = set()
        for fld in HEALTH_FIELDS:
            src = health_source_col(fld)
            disp = clean_subheader_label(fld)
            added_disp.add(disp)

            if src and src in h_n.columns:
                vals = h_n[src]
                if normalize_header(fld) == normalize_header("Date of next review appointment"):
                    vals = parse_date(vals)
                tmp[disp] = vals.values
            else:
                tmp[disp] = pd.NA

        # Append ANY extra health columns not already included (so nothing is missed)
        for c in extra_health_cols:
            # skip if it would collide with existing display names
            disp = str(c).strip()
            if disp in tmp.columns or disp in added_disp:
                continue
            tmp[disp] = h_n[c].values

        # Rename to wide:
        rename_map = {}
        for c in tmp.columns:
            if c == "_SaheliKey":
                continue
            if c == block:
                rename_map[c] = c
            else:
                rename_map[c] = f"{block}  {c}"  # double-space for grouped header parsing
        tmp = tmp.rename(columns=rename_map)

        tmp = tmp.drop_duplicates(subset=["_SaheliKey"], keep="first")
        health_wide = health_wide.merge(tmp, on="_SaheliKey", how="left")

    # FINAL LEFT JOIN (REG PRIMARY)
    final_df = reg_out.merge(health_wide, on="_SaheliKey", how="left")
    final_df = final_df.merge(assess_counts.reset_index(), on="_SaheliKey", how="left")
    final_df["No of assessment completed"] = final_df["No of assessment completed"].fillna(0).astype("Int64")

    # Ensure Saheli Card Number always present
    final_df["Saheli Card Number"] = final_df["Saheli Card Number"].fillna(final_df["_SaheliKey"])

    # Put preferred first columns, then everything else
    first_cols = ["No of assessment completed"] + REG_OUTPUT_LABELS
    for c in first_cols:
        if c not in final_df.columns:
            final_df[c] = pd.NA

    rem_cols = [c for c in final_df.columns if c not in first_cols + ["_SaheliKey"]]

    # Sort assessment columns by:
    #  - assessment number
    #  - date column first
    #  - HEALTH_FIELDS order
    health_order_map = build_health_field_order_map()

    def assessment_sort_key(c):
        m = re.match(r"^(\d+)(st|nd|rd|th)\sAssessment(?:\s{2,}(.*))?$", str(c))
        if m:
            block_n = int(m.group(1))
            suffix = (m.group(3) or "").strip()
            # Date column first (support both styles)
            if suffix == "" or normalize_header(suffix) in ("assessmentdate", "completiondate"):
                return (0, block_n, -1, "assessmentdate")
            suffix_disp = clean_subheader_label(suffix)
            field_pos = health_order_map.get(suffix_disp, 9999)
            return (0, block_n, field_pos, suffix_disp.lower())
        return (1, 999, 9999, str(c).lower())

    rem_cols = sorted(rem_cols, key=assessment_sort_key)
    final_df = final_df[first_cols + rem_cols].copy()

    # Sort by Saheli number
    final_df["_sort_num"] = pd.to_numeric(final_df["Saheli Card Number"], errors="coerce")
    final_df = final_df.sort_values(["_sort_num", "Saheli Card Number"], kind="mergesort").drop(columns=["_sort_num"])

    # Validation
    reg_keys = set(reg_out["_SaheliKey"].dropna().astype(str))
    final_keys = set(final_df["Saheli Card Number"].dropna().astype(str))
    missing_from_final = sorted(reg_keys - final_keys)
    print(f"[CHECK] REG unique Saheli keys:   {len(reg_keys)}")
    print(f"[CHECK] FINAL unique Saheli keys: {len(final_keys)}")
    print(f"[CHECK] Missing in FINAL:         {len(missing_from_final)}")
    if missing_from_final:
        print("[CHECK] Sample missing keys:", missing_from_final[:20])

    return final_df


# =========================
# STEP 4: FORMAT FINAL FILE (GROUPED HEADER)
# =========================
def apply_grouped_header_format(output_path: str):
    """
    Convert single-row header into a 2-row grouped header for assessment columns:
      Row1: 1st Assessment | 1st Assessment | ... | 2nd Assessment | ...
      Row2: (blank for date col) or subheader labels
    Keeps registration columns on row 1 and blank row 2 for them.
    """
    wb = load_workbook(output_path)
    ws = wb.active

    ws.insert_rows(1)
    max_col = ws.max_column
    original_headers = [ws.cell(row=2, column=c).value for c in range(1, max_col + 1)]

    row1_vals = []
    row2_vals = []

    for h in original_headers:
        h = "" if h is None else str(h)
        m = re.match(r"^(\d+)(st|nd|rd|th)\sAssessment(?:\s{2,}(.*))?$", h)
        if m:
            block_label = f"{m.group(1)}{m.group(2)} Assessment"
            sub = (m.group(3) or "").strip()
            row1_vals.append(block_label)
            row2_vals.append(sub if sub else "")
        else:
            row1_vals.append(h)
            row2_vals.append("")

    for c, v in enumerate(row1_vals, start=1):
        ws.cell(row=1, column=c, value=v)
    for c, v in enumerate(row2_vals, start=1):
        ws.cell(row=2, column=c, value=v)

    # Merges
    c = 1
    while c <= max_col:
        v1 = ws.cell(row=1, column=c).value
        if v1 is None:
            c += 1
            continue

        v1s = str(v1)
        is_assessment = bool(re.match(r"^\d+(st|nd|rd|th)\sAssessment$", v1s))

        if is_assessment:
            start = c
            end = c
            while end + 1 <= max_col and str(ws.cell(row=1, column=end + 1).value) == v1s:
                end += 1
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            c = end + 1
        else:
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
            c += 1

    # Styling
    purple_fill = PatternFill(fill_type="solid", fgColor=COLOR_FIRST_ASSESSMENT)
    blue_fill = PatternFill(fill_type="solid", fgColor=COLOR_SECOND_ASSESSMENT)
    other_fill = PatternFill(fill_type="solid", fgColor=COLOR_OTHER_ASSESSMENT)
    reg_fill = PatternFill(fill_type="solid", fgColor=COLOR_REG_HEADER)

    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="000000", bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx in range(1, max_col + 1):
        top_val = ws.cell(row=1, column=col_idx).value
        top_text = "" if top_val is None else str(top_val)

        fill = reg_fill
        font = black_font
        align1 = center
        align2 = left_wrap

        m = re.match(r"^(\d+)(st|nd|rd|th)\sAssessment$", top_text)
        if m:
            n = int(m.group(1))
            if n == 1:
                fill = purple_fill
                font = white_font
            elif n == 2:
                fill = blue_fill
                font = white_font
            else:
                fill = other_fill
                font = black_font

        for r in (1, 2):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = fill
            cell.font = font
            cell.alignment = align1 if r == 1 else align2

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{ws.max_row}"

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 44

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        top = ws.cell(row=1, column=col_idx).value or ""
        sub = ws.cell(row=2, column=col_idx).value or ""

        if re.match(r"^\d+(st|nd|rd|th)\sAssessment$", str(top)):
            if sub == "":
                ws.column_dimensions[col_letter].width = 14
            else:
                text_len = len(str(sub))
                ws.column_dimensions[col_letter].width = max(18, min(42, text_len * 0.9))
        else:
            text_len = len(str(top))
            ws.column_dimensions[col_letter].width = max(14, min(32, text_len * 0.9))

    wb.save(output_path)


# =========================
# MAIN RUNNER
# =========================
def main():
    print("=== STEP 1: Read source files ===")
    df_reg_raw = read_excel_flexible(REG_FILE, REG_SHEET_NAME)
    df_health_raw = read_excel_flexible(HEALTH_FILE, HEALTH_SHEET_NAME)
    print(f"Registration rows: {len(df_reg_raw)}")
    print(f"Health rows:       {len(df_health_raw)}")

    print("\n=== STEP 2: Prepare registration ===")
    df_reg_clean = prepare_registration_df(df_reg_raw)
    write_excel(df_reg_clean, REG_OUTPUT_FILE)
    print("Registration cleaned saved:", REG_OUTPUT_FILE)

    print("\n=== STEP 3: Prepare healthassessment ===")
    df_health_prepared = prepare_health_df(df_health_raw)
    write_excel(df_health_prepared, HEALTH_OUTPUT_FILE)
    print("Health prepared saved:", HEALTH_OUTPUT_FILE)

    print("\n=== STEP 4: Create final WIDE output (LEFT JOIN from registration) ===")
    final_df = create_final_wide_df(df_reg_clean, df_health_prepared)
    write_excel(final_df, FINAL_OUTPUT_FILE)

    if APPLY_GROUPED_HEADER:
        print("\n=== STEP 5: Apply grouped Excel header formatting ===")
        apply_grouped_header_format(FINAL_OUTPUT_FILE)

    print("Final wide output saved:", FINAL_OUTPUT_FILE)
    print("\n=== DONE ===")
    print(f"Registration cleaned: {REG_OUTPUT_FILE}")
    print(f"Health prepared:      {HEALTH_OUTPUT_FILE}")
    print(f"Final wide file:      {FINAL_OUTPUT_FILE}")
    print(f"Final rows: {len(final_df)} | Final columns: {len(final_df.columns)}")


if __name__ == "__main__":
    main()